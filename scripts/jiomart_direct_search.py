"""
Jiomart Direct Search — Search each NA product directly on Jiomart, get SP/MRP.

No pool, no intermediate matching. For each NA product:
  1. Search product name on Jiomart
  2. Scroll + paginate to find all results
  3. Match by brand + name + weight from search results
  4. Get SP/MRP directly
  5. Save all results to JSON + Excel

Usage:
    backend/venv/bin/python scripts/jiomart_direct_search.py --pincode 834002 \
        --csv /path/to/assortment.csv --state JH
"""
import asyncio
import contextlib
import csv
import io
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import quote

sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import normalize, parse_num, UNIT_ALIASES, to_base_unit

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA = PROJECT_ROOT / "data"

try:
    from rapidfuzz import fuzz
    def token_set_ratio(a, b):
        return fuzz.token_set_ratio(a, b)
except ImportError:
    def token_set_ratio(a, b):
        a_s = " ".join(sorted(a.lower().split()))
        b_s = " ".join(sorted(b.lower().split()))
        return SequenceMatcher(None, a_s, b_s).ratio() * 100

PINCODE_LOCATION = {
    "834002": {"pincode": "834002", "city": "RANCHI", "state": "JHARKHAND", "lat": "23.3441", "lon": "85.3096"},
    "712232": {"pincode": "712232", "city": "KOLKATA", "state": "WEST BENGAL", "lat": "22.5726", "lon": "88.3639"},
    "492001": {"pincode": "492001", "city": "RAIPUR", "state": "CHHATTISGARH", "lat": "21.2514", "lon": "81.6296"},
    "825301": {"pincode": "825301", "city": "HAZARIBAGH", "state": "JHARKHAND", "lat": "23.9925", "lon": "85.3637"},
    "495001": {"pincode": "495001", "city": "BILASPUR", "state": "CHHATTISGARH", "lat": "22.0797", "lon": "82.1409"},
    "831001": {"pincode": "831001", "city": "JAMSHEDPUR", "state": "JHARKHAND", "lat": "22.8046", "lon": "86.2029"},
}

# Weight parsing
UNIT_RE = r'(g|gm|gms|gram|grams|kg|kgs|kilo|ml|mls|l|ltr|ltrs|litre|liter|pc|pcs|piece|pieces|unit|units|n|nos)'


def parse_weight(text):
    """Parse weight from text → (base_value, base_unit) in g/ml/pc."""
    if not text:
        return None, None
    s = str(text).strip().lower()
    pack_n = 1
    pm = re.search(r'(?:pack|set|box)\s+(?:of\s+)?(\d+)', s)
    if pm:
        pack_n = int(pm.group(1))
    m = re.search(rf'(\d+\.?\d*)\s*[x×]\s*(\d+\.?\d*)\s*{UNIT_RE}\b', s)
    if m:
        v = float(m.group(1)) * float(m.group(2))
        u = UNIT_ALIASES.get(m.group(3), m.group(3))
        return to_base_unit(v, u)
    m = re.search(rf'(\d+\.?\d*)\s*{UNIT_RE}\s*[x×]\s*(\d+)', s)
    if m:
        v = float(m.group(1)) * int(m.group(3))
        u = UNIT_ALIASES.get(m.group(2), m.group(2))
        return to_base_unit(v, u)
    m = re.search(rf'(\d+\.?\d*)\s*{UNIT_RE}\b', s)
    if m:
        v = float(m.group(1)) * pack_n
        u = UNIT_ALIASES.get(m.group(2), m.group(2))
        return to_base_unit(v, u)
    return None, None


def build_search_term(am):
    """Build a good search term from AM product. Short, specific."""
    name = am.get("display_name", "")
    brand = (am.get("brand") or "").strip()

    # For loose/ASM items: search product name directly (e.g., "Sugar 1 kg", "Chana Dal 500g")
    is_loose = "loose" in name.lower() or "asm" in brand.lower() or \
               brand.strip().upper() in ("ASM", "LOOSE CONTAINER")
    if is_loose:
        # Strip "Loose" prefix and brand, keep product + weight
        clean = re.sub(r'\s*\|.*$', '', name).strip()
        clean = re.sub(r'(?i)^(loose\s+)', '', clean).strip()
        clean = re.sub(r'\s*[-–]\s*$', '', clean).strip()
        # Replace hyphens
        clean = re.sub(r'[-–]', ' ', clean).strip()
        clean = re.sub(r'\s+', ' ', clean)
        return clean[:60]  # Keep full name for loose (need weight in search)

    # For branded items: product name without weight suffix
    clean = re.sub(r'\s*\|.*$', '', name).strip()
    clean = re.sub(r'\s*[-–]\s*\d+\.?\d*\s*(g|gm|kg|ml|l|ltr|pcs?)\b.*$', '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'[-–]', ' ', clean).strip()
    clean = re.sub(r'\s+', ' ', clean)
    words = clean.split()[:5]
    term = " ".join(words)
    if brand and brand.lower().replace('-', ' ') not in term.lower():
        brand_clean = re.sub(r'[-–]', ' ', brand)
        term = f"{brand_clean} {term}"
    return term


def search_results_to_pool(search_results):
    """Convert Jiomart search results to engine-compatible pool format."""
    pool = []
    for p in search_results:
        name = p.get("name", "")
        if not name:
            continue
        # Parse unit from name
        unit = ""
        m = re.search(
            r"(\d+\.?\d*)\s*(g|gm|gms|kg|kgs|ml|mls|l|ltr|ltrs|pc|pcs)\b",
            name.lower(),
        )
        if m:
            unit = f"{m.group(1)} {m.group(2)}"
        pool.append({
            "product_id": p.get("product_id", ""),
            "product_url": f"https://www.jiomart.com/product/{p['slug']}" if p.get("slug") else "",
            "product_name": name,
            "brand": p.get("brand", ""),
            "price": p.get("sp"),
            "mrp": p.get("mrp"),
            "unit": unit,
            "in_stock": True,
            "barcode": p.get("barcode", ""),
            "slug": p.get("slug", ""),
        })
    return pool


def set_engine_pool(engine, search_results):
    """Set search results as engine pool. Call once per search, not per item."""
    if not search_results:
        return
    pool = search_results_to_pool(search_results)
    with contextlib.redirect_stdout(io.StringIO()):
        engine.set_pool(pool)


def match_with_engine(engine, am, am_mrp):
    """
    Match AM product against engine's current pool.
    Pool must be set beforehand via set_engine_pool().
    Returns (matched_product, score, status, reason, flags).
    """
    if not engine._pool:
        return None, 0, "NA", "no_search_results", []

    result = engine.match(am, am_mrp)

    matched_p = None
    if result.product:
        matched_p = {
            "name": result.product.get("product_name", ""),
            "brand": result.product.get("brand", ""),
            "sp": result.product.get("price"),
            "mrp": result.product.get("mrp"),
            "product_id": result.product.get("product_id", ""),
            "slug": result.product.get("slug", ""),
        }

    return matched_p, result.score, result.status, result.reason, result.flags


async def init_browser(pw, pincode):
    """Init Firefox with Jiomart location."""
    loc = PINCODE_LOCATION.get(pincode, PINCODE_LOCATION["834002"])
    browser = await pw.firefox.launch(headless=True)
    context = await browser.new_context(
        user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:128.0) Gecko/20100101 Firefox/128.0",
        viewport={"width": 1366, "height": 768},
        locale="en-IN", timezone_id="Asia/Kolkata",
    )
    loc_json = json.dumps({
        "country": "INDIA", "country_iso_code": "IN",
        "city": loc["city"], "pincode": loc["pincode"], "state": loc["state"],
    })
    await context.add_cookies([
        {"name": "pincode", "value": loc["pincode"], "domain": ".jiomart.com", "path": "/"},
        {"name": "app_location_details", "value": quote(loc_json), "domain": "www.jiomart.com", "path": "/"},
    ])
    await context.add_init_script(f"""
        localStorage.setItem('pin', JSON.stringify({{
            country: "INDIA", pincode: "{loc['pincode']}", city: "{loc['city']}", state: "{loc['state']}"
        }}));
        localStorage.setItem('jio_lat_long', JSON.stringify({{
            latitude: "{loc['lat']}", longitude: "{loc['lon']}"
        }}));
        Object.defineProperty(navigator, 'webdriver', {{get: () => undefined}});
    """)
    return browser, context


async def search_jiomart(page, search_term, max_pages=3):
    """Search on Jiomart, scroll + paginate, return all products found."""
    all_products = []
    seen = set()

    # Clean search term: replace hyphens/special chars with spaces
    clean_term = re.sub(r'[-–/|&]', ' ', search_term)
    clean_term = re.sub(r'\s+', ' ', clean_term).strip()

    for pg in range(1, max_pages + 1):
        url = f"https://www.jiomart.com/products?q={quote(clean_term)}"
        if pg > 1:
            url += f"&page={pg}"

        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=25000)
        except Exception:
            # Retry once on timeout
            try:
                await asyncio.sleep(2)
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            except Exception:
                break
        await asyncio.sleep(2.5)

        # Scroll fully
        prev = 0
        for _ in range(12):
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(1)
            count = await page.evaluate("document.querySelectorAll('[class*=productCard]').length")
            if count == prev:
                break
            prev = count

        # Extract from dataLayer (has product_id, slug, brand)
        dl = await page.evaluate('''() => {
            const p = [], s = new Set();
            if (window.dataLayer) for (const e of window.dataLayer)
                for (const i of (e?.ecommerce?.items || []))
                    if (i.product_url && !s.has(i.product_url)) {
                        s.add(i.product_url);
                        p.push({slug: i.product_url, product_id: String(i.item_id || ''),
                                name: i.item_name || '', brand: i.item_brand || '', sp: i.price || null});
                    }
            return p;
        }''')

        # Extract from DOM (has MRP)
        dom = await page.evaluate(r'''() => {
            const p = [];
            document.querySelectorAll('.productCard__productDescription').forEach(card => {
                const nameEl = card.querySelector('.productCard__productTitle, h3');
                const spEl = card.querySelector('[class*=currentPrice]');
                const mrpEl = card.querySelector('[class*=originalPrice]');
                const name = nameEl ? nameEl.textContent.trim() : '';
                const sp = spEl ? parseFloat(spEl.textContent.replace(/[₹,\s]/g, '')) : null;
                const mrp = mrpEl ? parseFloat(mrpEl.textContent.replace(/[₹,\s]/g, '')) : null;
                if (name && sp) p.push({name, sp, mrp: mrp || sp});
            });
            return p;
        }''')

        # Merge dataLayer + DOM (dataLayer has pid/slug/brand, DOM has MRP)
        for dl_p in dl:
            for dom_p in dom:
                if dl_p["name"] == dom_p["name"] and dl_p.get("sp") == dom_p.get("sp"):
                    dl_p["mrp"] = dom_p.get("mrp")
                    break
            if "mrp" not in dl_p:
                dl_p["mrp"] = dl_p.get("sp")

        # Add unique products
        page_new = 0
        for p in dl:
            key = p.get("name", "")
            if key and key not in seen:
                seen.add(key)
                all_products.append(p)
                page_new += 1
        # DOM-only products (no dataLayer entry)
        for p in dom:
            key = p.get("name", "")
            if key and key not in seen:
                seen.add(key)
                all_products.append({"slug": "", "product_id": "", "name": key,
                                     "brand": "", "sp": p["sp"], "mrp": p["mrp"]})
                page_new += 1

        if page_new == 0:
            break  # Empty page — stop

    return all_products


async def main():
    args = sys.argv[1:]
    pincode = "834002"
    csv_path = None
    state_filter = None

    for i, a in enumerate(args):
        if a == "--pincode" and i + 1 < len(args):
            pincode = args[i + 1]
        elif a == "--csv" and i + 1 < len(args):
            csv_path = args[i + 1]
        elif a == "--state" and i + 1 < len(args):
            state_filter = args[i + 1]

    print(f"\n{'=' * 60}")
    print(f"  JIOMART DIRECT SEARCH — {pincode}")
    print(f"{'=' * 60}\n")

    # ── 1. Load AM master ──
    am_map = json.load(open(DATA / "am_product_master.json"))
    mrp_path = DATA / "latest_mrp_wrhs_1.json"
    mrp_map = json.load(open(mrp_path)) if mrp_path.exists() else {}

    # ── 2. Load target items from CSV ──
    target_items = {}
    if csv_path:
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                if state_filter and row.get("State", "").strip() != state_filter:
                    continue
                ic = str(row.get("Item Code", "")).strip()
                if ic:
                    am = am_map.get(ic, {})
                    target_items[ic] = {
                        "item_code": ic,
                        "display_name": row.get("Item Name", "") or am.get("display_name", ""),
                        "brand": row.get("Brand", "") or am.get("brand", ""),
                        "master_category": row.get("Master Catgory", "").strip() or am.get("master_category", ""),
                        "product_type": row.get("Product Type", "") or am.get("product_type", ""),
                        "unit": am.get("unit", ""),
                        "unit_value": am.get("unit_value"),
                        "mrp": am.get("mrp"),
                    }
    print(f"Target items: {len(target_items)}")

    # ── 3. Pool-based matching FIRST (existing jiomart_product_master) ──
    from unified_matcher import UnifiedMatchingEngine, build_pool

    ean_map = json.load(open(DATA / "ean_map.json")) if (DATA / "ean_map.json").exists() else {}
    engine = UnifiedMatchingEngine(ean_map)

    print(f"\n--- Step 1: Pool-based matching (existing data) ---", flush=True)
    pool = build_pool(pincode, "jiomart")
    engine.set_pool(pool)

    results = {}
    need_search = {}  # items that need Jiomart search

    for ic, am in target_items.items():
        mrp_rec = mrp_map.get(ic)
        if mrp_rec:
            am["mrp"] = mrp_rec.get("mrp", am.get("mrp"))
        am_mrp_val = parse_num(am.get("mrp"))

        result = engine.match(am, am_mrp_val)

        if result.status == "COMPLETE MATCH" and result.product:
            # Already matched from pool — keep as-is, NO search needed
            sp = parse_num(result.product.get("price")) or parse_num(result.product.get("sp"))
            results[ic] = {
                "item_code": ic,
                "am_name": am.get("display_name"),
                "am_brand": am.get("brand"),
                "am_unit": f"{am.get('unit_value', '')} {am.get('unit', '')}".strip(),
                "am_mrp": am_mrp_val,
                "jio_name": result.product.get("product_name"),
                "jio_brand": result.product.get("brand"),
                "jio_sp": sp,
                "jio_mrp": result.product.get("mrp"),
                "jio_product_id": result.product.get("product_id"),
                "jio_slug": result.product.get("slug", ""),
                "jio_url": result.product.get("product_url", ""),
                "match_status": result.status,
                "match_score": round(result.score, 3),
                "match_reason": f"pool_{result.reason}",
                "match_flags": "",
            }
        else:
            need_search[ic] = am

    pool_matched = len(results)
    print(f"  Pool COMPLETE: {pool_matched} (kept as-is, no search)", flush=True)
    print(f"  Need search:        {len(need_search)}", flush=True)

    # ── 4. Jiomart search for remaining items ──
    print(f"\n--- Step 2: Jiomart search for {len(need_search)} items ---", flush=True)

    from playwright.async_api import async_playwright
    pw = await async_playwright().start()
    browser, context = await init_browser(pw, pincode)
    page = await context.new_page()

    try:
        await page.goto("https://www.jiomart.com", wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(2)
    except Exception:
        pass

    import io as _io

    total = len(target_items)
    searched = 0
    errors = 0

    # Group ONLY need_search items by brand
    brand_groups = defaultdict(list)
    for ic, am in need_search.items():
        brand = (am.get("brand") or "").strip()
        brand_key = normalize(brand) if brand else f"_no_brand_{ic}"
        brand_groups[brand_key].append((ic, am))

    sorted_groups = sorted(brand_groups.items(), key=lambda x: -len(x[1]))
    search_cache = {}

    for gi, (brand_key, items) in enumerate(sorted_groups, 1):
        brand_name = items[0][1].get("brand", "").strip()

        # Search brand
        if brand_name and brand_name not in search_cache:
            try:
                brand_results = await search_jiomart(page, brand_name, max_pages=3)
                if not brand_results:
                    clean_brand = re.sub(r'[-–\'\"&/]', ' ', brand_name).strip()
                    clean_brand = re.sub(r'\s+', ' ', clean_brand)
                    if clean_brand != brand_name:
                        brand_results = await search_jiomart(page, clean_brand, max_pages=3)
                        searched += 1
                search_cache[brand_name] = brand_results
                searched += 1
                print(f"[{gi}/{len(sorted_groups)}] Brand: \"{brand_name}\" → {len(brand_results)} products "
                      f"({len(items)} items)", flush=True)
            except Exception as e:
                search_cache[brand_name] = []
                errors += 1
                print(f"[{gi}/{len(sorted_groups)}] Brand: \"{brand_name}\" → ERROR: {str(e)[:50]}", flush=True)
        brand_results = search_cache.get(brand_name, [])

        # Set engine pool for brand results
        set_engine_pool(engine, brand_results)

        for ic, am in items:
            am_mrp_val = parse_num(am.get("mrp"))

            matched_p, score, status, reason, flags = match_with_engine(engine, am, am_mrp_val)

            # Name search fallback for all unmatched items (including loose/ASM)
            if status == "NA" or (status == "PARTIAL MATCH" and score < 0.55):
                search_term = build_search_term(am)
                if search_term not in search_cache:
                    try:
                        name_results = await search_jiomart(page, search_term, max_pages=2)
                        search_cache[search_term] = name_results
                        searched += 1
                    except Exception:
                        search_cache[search_term] = []
                        errors += 1
                name_results = search_cache.get(search_term, [])
                set_engine_pool(engine, name_results)
                matched_p2, score2, status2, reason2, flags2 = match_with_engine(engine, am, am_mrp_val)
                if score2 > score:
                    matched_p, score, status, reason, flags = matched_p2, score2, status2, reason2, flags2

            result = {
                "item_code": ic,
                "am_name": am.get("display_name"),
                "am_brand": am.get("brand"),
                "am_unit": f"{am.get('unit_value', '')} {am.get('unit', '')}".strip(),
                "am_mrp": am_mrp_val,
                "match_status": status,
                "match_score": round(score, 3),
                "match_reason": reason,
                "match_flags": ", ".join(flags) if flags else "",
            }

            if matched_p:
                result.update({
                    "jio_name": matched_p.get("name"),
                    "jio_brand": matched_p.get("brand"),
                    "jio_sp": matched_p.get("sp"),
                    "jio_mrp": matched_p.get("mrp"),
                    "jio_product_id": matched_p.get("product_id"),
                    "jio_slug": matched_p.get("slug"),
                    "jio_url": f"https://www.jiomart.com/product/{matched_p['slug']}" if matched_p.get("slug") else "",
                })

            results[ic] = result

        # Save checkpoint every 30 groups
        if gi % 30 == 0:
            _save_json(results, pincode)
            matched_so_far = sum(1 for r in results.values() if r["match_status"] != "NA")
            print(f"  [checkpoint] {len(results)}/{total} done, "
                  f"{matched_so_far} matched, {searched} searches", flush=True)

    await browser.close()
    await pw.stop()

    # ── 5. Save results ──
    _save_json(results, pincode)
    _save_excel(results, target_items, pincode)

    # ── 6. Summary ──
    status_counts = defaultdict(int)
    for r in results.values():
        status_counts[r["match_status"]] += 1

    print(f"\n{'=' * 60}")
    print(f"  DIRECT SEARCH COMPLETE — {pincode}")
    print(f"{'=' * 60}")
    print(f"  Total items:     {total}")
    print(f"  From pool:       {pool_matched} (COMPLETE/SEMI — no search)")
    print(f"  Searched:        {len(need_search)} items, {searched} searches")
    print(f"  Errors:          {errors}")
    print()
    print(f"  Status breakdown:")
    for s, c in sorted(status_counts.items(), key=lambda x: -x[1]):
        pct = c / total * 100
        print(f"    {s:25s} {c:5d}  ({pct:.1f}%)")

    matched = sum(1 for r in results.values() if r["match_status"] != "NA")
    print(f"\n  Matched: {matched}/{total} ({matched/total*100:.1f}%)")
    print(f"  NA:      {status_counts.get('NA', 0)}/{total}")

    # Top matches
    top = sorted([r for r in results.values() if r.get("jio_name")],
                 key=lambda x: -x["match_score"])[:10]
    if top:
        print(f"\n  Top 10 matches:")
        for r in top:
            print(f"    [{r['match_score']:.2f}] {r['am_name']}")
            print(f"           → {r['jio_name']}  SP:{r['jio_sp']} MRP:{r['jio_mrp']}  [{r['match_status']}]")


def _save_json(results, pincode):
    """Save all results to JSON."""
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    out_path = DATA / f"jiomart_direct_search_{pincode}_{ts}.json"
    with open(out_path, "w") as f:
        json.dump({
            "pincode": pincode,
            "searched_at": datetime.now().isoformat(),
            "total": len(results),
            "matched": sum(1 for r in results.values() if r["match_status"] != "NA"),
            "results": list(results.values()),
        }, f, indent=2, default=str, ensure_ascii=False)
    print(f"  JSON: {out_path}", flush=True)


def _save_excel(results, target_items, pincode):
    """Save results as Excel with color-coded status."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("  openpyxl not installed — skipping Excel", flush=True)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Jiomart_Direct_Search"

    headers = [
        "Item Code", "AM Name", "Brand", "AM Unit", "AM MRP",
        "JIO Name", "JIO SP", "JIO MRP", "JIO URL",
        "Match Status", "Match Score", "Match Reason", "Match Flags",
    ]
    am_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
    jio_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, size=10)
        if i <= 5:
            c.fill = am_fill
        elif i <= 9:
            c.fill = jio_fill

    row_idx = 2
    for ic in sorted(results.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        r = results[ic]
        ws.cell(row=row_idx, column=1, value=int(ic) if ic.isdigit() else ic)
        ws.cell(row=row_idx, column=2, value=r.get("am_name"))
        ws.cell(row=row_idx, column=3, value=r.get("am_brand"))
        ws.cell(row=row_idx, column=4, value=r.get("am_unit"))
        ws.cell(row=row_idx, column=5, value=r.get("am_mrp"))
        ws.cell(row=row_idx, column=6, value=r.get("jio_name"))
        ws.cell(row=row_idx, column=7, value=r.get("jio_sp"))
        ws.cell(row=row_idx, column=8, value=r.get("jio_mrp"))
        ws.cell(row=row_idx, column=9, value=r.get("jio_url"))
        ws.cell(row=row_idx, column=10, value=r.get("match_status"))
        ws.cell(row=row_idx, column=11, value=r.get("match_score"))
        ws.cell(row=row_idx, column=12, value=r.get("match_reason"))
        ws.cell(row=row_idx, column=13, value=r.get("match_flags"))

        # Color status
        sc = ws.cell(row=row_idx, column=10)
        if r["match_status"] == "COMPLETE MATCH":
            sc.fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
        elif r["match_status"] == "PARTIAL MATCH":
            sc.fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")

        row_idx += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    ws.freeze_panes = "A2"

    out_path = Path("/Users/satyam/Desktop/price csv") / f"Jiomart_Direct_{pincode}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Excel: {out_path}", flush=True)


if __name__ == "__main__":
    asyncio.run(main())

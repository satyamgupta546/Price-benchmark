"""
Export SAM PDP scrape data and comparison results to CSV/Excel for review.

Produces 3 files in data/exports/:
  1. blinkit_pdp_<pincode>_<ts>.csv            — raw SAM PDP scrape (one row per SKU)
  2. blinkit_pdp_vs_anakin_<pincode>_<ts>.csv  — side-by-side SAM vs Anakin
  3. blinkit_pdp_vs_anakin_<pincode>_<ts>.xlsx — same as #2 but formatted Excel

Usage:
    python3 scripts/export_pdp_csv.py 834002
"""
import csv
import json
import sys
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent


def latest_file(subdir: str, pattern: str) -> Path | None:
    cands = sorted((PROJECT_ROOT / "data" / subdir).glob(pattern))
    return cands[-1] if cands else None


def latest_pdp_sam(pincode: str) -> Path | None:
    # Prefer the latest non-partial final file, otherwise use partial snapshot
    final = sorted((PROJECT_ROOT / "data" / "sam").glob(f"blinkit_pdp_{pincode}_2*.json"))
    if final:
        return final[-1]
    partial = PROJECT_ROOT / "data" / "sam" / f"blinkit_pdp_{pincode}_latest_partial.json"
    if partial.exists():
        return partial
    return None


def parse_num(v):
    if v is None or str(v).strip().lower() in ("", "na", "nan", "null", "none"):
        return None
    try:
        s = str(v).replace("₹", "").replace("Rs.", "").replace("Rs", "").replace(",", "").strip()
        s = s.rstrip("/-").strip()
        return float(s)
    except (ValueError, TypeError):
        return None


def clean_str(v) -> str:
    """Return empty string for sentinel missing values (NA, nan, null, empty)."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("", "na", "nan", "null", "none"):
        return ""
    return s


def main(pincode: str):
    pdp_path = latest_pdp_sam(pincode)
    ana_path = latest_file("anakin", f"blinkit_{pincode}_*.json")

    if not pdp_path:
        print(f"[export] ERROR: no SAM PDP file for {pincode}", file=sys.stderr)
        sys.exit(1)
    if not ana_path:
        print(f"[export] ERROR: no Anakin file for {pincode}", file=sys.stderr)
        sys.exit(1)

    print(f"[export] SAM PDP: {pdp_path.name}")
    print(f"[export] Anakin:   {ana_path.name}")

    pdp = json.load(open(pdp_path))
    ana = json.load(open(ana_path))

    # Index Anakin by item_code
    ana_by_code = {}
    for rec in ana["records"]:
        ic = (rec.get("Item_Code") or "").strip()
        if ic:
            ana_by_code[ic] = rec

    out_dir = PROJECT_ROOT / "data" / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")

    # ── File 1: Raw SAM PDP scrape ──
    raw_csv = out_dir / f"blinkit_pdp_{pincode}_{ts}.csv"
    with open(raw_csv, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "item_code", "blinkit_product_id", "blinkit_product_url",
            "sam_product_name", "sam_selling_price", "sam_mrp",
            "sam_in_stock", "sam_unit", "status", "error",
        ])
        w.writeheader()
        for p in pdp["products"]:
            w.writerow({k: p.get(k, "") for k in w.fieldnames})
    print(f"[export] Raw CSV: {raw_csv}")

    # ── File 2: Side-by-side SAM vs Anakin (CSV) ──
    side_csv = out_dir / f"blinkit_pdp_vs_anakin_{pincode}_{ts}.csv"
    rows = []
    for p in pdp["products"]:
        ic = p.get("item_code")
        ana_rec = ana_by_code.get(ic, {}) if ic else {}
        ana_sp = parse_num(ana_rec.get("Blinkit_Selling_Price"))
        ana_mrp = parse_num(ana_rec.get("Blinkit_Mrp_Price"))
        sam_sp = parse_num(p.get("sam_selling_price"))
        sam_mrp = parse_num(p.get("sam_mrp"))
        diff_pct = None
        if ana_sp and sam_sp:
            diff_pct = round(abs(sam_sp - ana_sp) / ana_sp * 100, 2)
        ana_uv_clean = clean_str(ana_rec.get("Unit_Value"))
        ana_unit_clean = clean_str(ana_rec.get("Unit"))
        weight_txt = f"{ana_uv_clean} {ana_unit_clean}".strip()
        rows.append({
            "item_code": ic,
            "anakin_name": clean_str(ana_rec.get("Item_Name")),
            "anakin_brand": clean_str(ana_rec.get("Brand")),
            "anakin_weight": weight_txt,
            "anakin_mrp_ref": clean_str(ana_rec.get("Mrp")),
            "anakin_blinkit_name": clean_str(ana_rec.get("Blinkit_Item_Name")),
            "anakin_blinkit_mrp": ana_mrp if ana_mrp is not None else "",
            "anakin_blinkit_sp": ana_sp if ana_sp is not None else "",
            "anakin_in_stock": clean_str(ana_rec.get("Blinkit_In_Stock_Remark")),
            "anakin_status": clean_str(ana_rec.get("Blinkit_Status")),
            "blinkit_product_id": p.get("blinkit_product_id", ""),
            "blinkit_product_url": p.get("blinkit_product_url", ""),
            "sam_product_name": p.get("sam_product_name", ""),
            "sam_unit": p.get("sam_unit", ""),
            "sam_mrp": sam_mrp if sam_mrp is not None else "",
            "sam_selling_price": sam_sp if sam_sp is not None else "",
            "sam_in_stock": p.get("sam_in_stock", ""),
            "sam_scrape_status": p.get("status", ""),
            "price_diff_pct": diff_pct if diff_pct is not None else "",
            "within_5pct": "YES" if (diff_pct is not None and diff_pct <= 5) else "",
            "within_10pct": "YES" if (diff_pct is not None and diff_pct <= 10) else "",
        })

    if not rows:
        print("[export] No rows to export (empty SAM PDP file).", file=sys.stderr)
        return
    with open(side_csv, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"[export] Side-by-side CSV: {side_csv}")

    # ── File 3: Excel version with styling ──
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[export] openpyxl not available — skipping Excel export", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "SAM vs Anakin"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")

    headers = list(rows[0].keys())
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(headers, 1):
            val = row.get(key, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if key == "price_diff_pct" and isinstance(val, (int, float)):
                if val <= 5:
                    cell.fill = green_fill
                elif val <= 10:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill
            if key == "sam_scrape_status" and val == "error":
                cell.fill = red_fill
            if key == "sam_scrape_status" and val == "no_price":
                cell.fill = yellow_fill

    # Column widths
    col_widths = {
        "item_code": 10, "anakin_name": 40, "anakin_brand": 15,
        "anakin_weight": 12, "anakin_mrp_ref": 10,
        "anakin_blinkit_name": 40, "anakin_blinkit_mrp": 10, "anakin_blinkit_sp": 10,
        "anakin_in_stock": 14, "anakin_status": 14,
        "blinkit_product_id": 14, "blinkit_product_url": 60,
        "sam_product_name": 40, "sam_unit": 12, "sam_mrp": 10,
        "sam_selling_price": 12, "sam_in_stock": 10, "sam_scrape_status": 12,
        "price_diff_pct": 12, "within_5pct": 10, "within_10pct": 10,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(h, 15)

    ws.freeze_panes = "C2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ok = sum(1 for r in rows if r["sam_scrape_status"] == "ok")
    err = sum(1 for r in rows if r["sam_scrape_status"] == "error")
    noprice = sum(1 for r in rows if r["sam_scrape_status"] == "no_price")
    priced = [r for r in rows if r["price_diff_pct"] != "" and isinstance(r["price_diff_pct"], (int, float))]
    in5 = sum(1 for r in priced if r["price_diff_pct"] <= 5)
    in10 = sum(1 for r in priced if r["price_diff_pct"] <= 10)

    summary = [
        ("Total rows",                len(rows)),
        ("Scrape OK with price",      ok),
        ("No price on PDP",           noprice),
        ("Scrape errors",             err),
        ("", ""),
        ("Price-comparable pool",     len(priced)),
        ("Within ±5%",                f"{in5} ({in5/max(len(priced),1)*100:.1f}%)"),
        ("Within ±10%",               f"{in10} ({in10/max(len(priced),1)*100:.1f}%)"),
    ]
    for i, (k, v) in enumerate(summary, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 25

    excel_path = out_dir / f"blinkit_pdp_vs_anakin_{pincode}_{ts}.xlsx"
    wb.save(excel_path)
    print(f"[export] Excel: {excel_path}")
    print()
    print(f"Total rows exported: {len(rows)}")


if __name__ == "__main__":
    pincode = sys.argv[1] if len(sys.argv) > 1 else "834002"
    main(pincode)

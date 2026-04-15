"""Test script for daily_report.py — verifies all 12 test points."""
import sys
sys.path.insert(0, "scripts")
from pathlib import Path
from daily_report import (
    load_sam_prices, load_cascade_matches, load_anakin, step3_generate_excel,
    clean, DATA, DATE
)

PASS = 0
FAIL = 0

def check(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS: {name}")
    else:
        FAIL += 1
        print(f"  FAIL: {name} — {detail}")


print("=" * 60)
print("  TESTING daily_report.py")
print("=" * 60)

# --- Test 1: Coverage capped at 100% ---
print("\n[Test 1] Coverage capped at 100%")
import daily_report
# Simulate: total_priced > usable
# The formula: min(round(total_priced * 100 / max(usable, 1), 1), 100.0)
# Check the source code has 'min(' in coverage line
import inspect
src = inspect.getsource(daily_report.step3_generate_excel)
check("coverage_pct uses min(..., 100.0)", "min(" in src and "100.0" in src)

# --- Test 2: Anakin SP written as float ---
print("\n[Test 2] Anakin SP written as float in Excel")
from openpyxl import load_workbook
out = step3_generate_excel("834002", "Ranchi")
wb = load_workbook(out)
# Find blinkit sheet
blinkit_sheet = None
for name in wb.sheetnames:
    if "blinkit" in name:
        blinkit_sheet = wb[name]
        break
if blinkit_sheet:
    # Col 7 = Anakin SP (1-indexed)
    found_numeric = False
    found_string = False
    for row in blinkit_sheet.iter_rows(min_row=2, max_row=50, min_col=7, max_col=7):
        val = row[0].value
        if val is not None:
            if isinstance(val, (int, float)):
                found_numeric = True
            elif isinstance(val, str) and val.replace('.','').isdigit():
                found_string = True
    check("Anakin SP is numeric (not string)", found_numeric and not found_string,
          f"numeric={found_numeric}, string_num={found_string}")
else:
    check("Anakin SP is float type", False, "no blinkit sheet found")

# --- Test 3: SAM_Name in headers and row_data ---
print("\n[Test 3] SAM_Name shown in Excel headers and row_data")
if blinkit_sheet:
    headers = [cell.value for cell in blinkit_sheet[1]]
    has_sam_name_header = any("SAM" in str(h) and "Name" in str(h) for h in headers if h)
    check("SAM Name header exists", has_sam_name_header, f"headers={headers}")

    # Check col 11 has some non-empty data
    sam_name_col_idx = None
    for i, h in enumerate(headers):
        if h and "SAM" in str(h) and "Name" in str(h):
            sam_name_col_idx = i
            break
    if sam_name_col_idx is not None:
        has_data = False
        for row in blinkit_sheet.iter_rows(min_row=2, max_row=100, min_col=sam_name_col_idx+1, max_col=sam_name_col_idx+1):
            if row[0].value:
                has_data = True
                break
        check("SAM Name has data in rows", has_data)
    else:
        check("SAM Name has data in rows", False, "header not found")
else:
    check("SAM Name header exists", False, "no blinkit sheet")
    check("SAM Name has data in rows", False, "no blinkit sheet")

# --- Test 4: or-chain fixed for score=0 and price=0 ---
print("\n[Test 4] Explicit None checks (not truthiness) for score=0 and price=0")
src_cascade = inspect.getsource(daily_report.load_cascade_matches)
check("load_cascade_matches uses 'is None' for ic/sp", "ic is None or sp is None" in src_cascade)
check("load_cascade_matches uses 'if score is None'", "if score is None:" in src_cascade)

src_excel = inspect.getsource(daily_report.step3_generate_excel)
check("Excel diff uses 'is not None' check", "ana_sp is not None and sam_sp is not None" in src_excel)
check("PDP fallback uses 'sam_sp is None'", "if sam_sp is None and ic in cascade_matched" in src_excel)
check("match_method uses 'sam_sp is not None'", 'if sam_sp is not None' in src_excel)

# --- Test 5: Sheet names within 31-char limit ---
print("\n[Test 5] Sheet names within 31-char Excel limit")
all_ok = True
for name in wb.sheetnames:
    if len(name) > 31:
        all_ok = False
        check(f"Sheet '{name}' <= 31 chars", False, f"len={len(name)}")
check("All sheet names <= 31 chars", all_ok, f"names={wb.sheetnames}")

# --- Test 6: Dead variables removed ---
print("\n[Test 6] Dead variables removed")
full_src = inspect.getsource(daily_report)
check("No 'matched_count' variable", "matched_count" not in full_src)
check("No 'within_5' variable", "within_5" not in full_src)
# 'total' is used as total_priced, skip checking bare 'total'

# --- Test 7: Unused imports removed ---
print("\n[Test 7] Unused imports removed")
import_lines = [line.strip() for line in inspect.getsource(daily_report).split('\n') if line.strip().startswith('import ') or line.strip().startswith('from ')]
has_csv = any('import csv' in l for l in import_lines)
has_os = any('import os' in l and 'from os' not in l for l in import_lines)
check("No 'import csv'", not has_csv)
check("No 'import os'", not has_os)

# --- Test 8: projects/ PDP items skipped ---
print("\n[Test 8] projects/ PDP items skipped in load_sam_prices()")
src_sam = inspect.getsource(daily_report.load_sam_prices)
check("Skips projects/ names", 'startswith("projects/")' in src_sam)

# --- Test 9: load_cascade_matches keeps higher-scored match ---
print("\n[Test 9] Higher-scored match preserved in load_cascade_matches()")
check("Has score comparison guard", 'matched[ic]["match_score"] >= score' in src_cascade)

# --- Test 10: image_match + barcode_match patterns ---
print("\n[Test 10] image_match + barcode_match glob patterns")
check("Has image_match pattern", "image_match" in src_cascade)
check("Has barcode_match pattern", "barcode_match" in src_cascade)

# --- Test 11: EAN fetch in step1 ---
print("\n[Test 11] EAN fetch added to step1_fetch_anakin()")
src_step1 = inspect.getsource(daily_report.step1_fetch_anakin)
check("Calls fetch_ean_map.py", "fetch_ean_map.py" in src_step1)

# --- Test 12: Subprocess errors logged ---
print("\n[Test 12] Subprocess errors logged")
src_run = inspect.getsource(daily_report.run)
check("Logs on non-zero returncode", "returncode" in src_run and "stderr" in src_run)

# --- Excel data type checks ---
print("\n[Excel Data Checks]")
if blinkit_sheet:
    # SAM SP column = 12
    found_sam_float = False
    for row in blinkit_sheet.iter_rows(min_row=2, max_row=50, min_col=12, max_col=12):
        val = row[0].value
        if val is not None:
            if isinstance(val, (int, float)):
                found_sam_float = True
                break
    check("SAM SP column has numeric values", found_sam_float)

    # Column count = 18
    header_count = sum(1 for cell in blinkit_sheet[1] if cell.value is not None)
    check(f"Column count is 18", header_count == 18, f"got {header_count}")

# Coverage check
summary_sheet = wb[wb.sheetnames[0]]
for row in summary_sheet.iter_rows(min_row=1, max_row=30, values_only=True):
    if row[0] and "Coverage" in str(row[0]):
        cov_val = str(row[1]).replace('%', '')
        try:
            cov = float(cov_val)
            check(f"Coverage {cov}% <= 100", cov <= 100.0, f"got {cov}%")
        except ValueError:
            check(f"Coverage is numeric", False, f"got {row[1]}")


print(f"\n{'=' * 60}")
print(f"  RESULTS: {PASS} passed, {FAIL} failed")
print(f"{'=' * 60}")

if FAIL > 0:
    sys.exit(1)

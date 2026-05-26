"""
Jiomart Price Fetcher — Mapping-based daily price fetch.

Flow:
  1. Load am_jiomart_mapping.json (AM item_code → Jiomart product_id/URL)
  2. Mapped items → Open URL directly → fetch latest SP/MRP (fast, no search)
  3. Unmapped items → Search on Jiomart → match → save mapping + SP/MRP
  4. Save updated mapping + results JSON + Excel

Daily: mapped items just need URL fetch (2-3 sec each). Only NEW items need search.

Usage:
    backend/venv/bin/python scripts/jiomart_fetch_prices.py --pincode 834002 \
        --csv /path/to/assortment.csv --state JH
"""
import asyncio
import contextlib
import csv
import io
import json
import random
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import normalize, parse_num, UNIT_ALIASES, to_base_unit
from unified_matcher import UnifiedMatchingEngine

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA = PROJECT_ROOT / "data"
PINCODE_LOCATION = {
    "834002": {"pincode": "834002", "city": "RANCHI", "state": "JHARKHAND", "lat": "23.3441", "lon": "85.3096"},
    "712232": {"pincode": "712232", "city": "KOLKATA", "state": "WEST BENGAL", "lat": "22.5726", "lon": "88.3639"},
    "492001": {"pincode": "492001", "city": "RAIPUR", "state": "CHHATTISGARH", "lat": "21.2514", "lon": "81.6296"},
    "825301": {"pincode": "825301", "city": "HAZARIBAGH", "state": "JHARKHAND", "lat": "23.9925", "lon": "85.3637"},
    "495001": {"pincode": "495001", "city": "BILASPUR", "state": "CHHATTISGARH", "lat": "22.0797", "lon": "82.1409"},
    "831001": {"pincode": "831001", "city": "JAMSHEDPUR", "state": "JHARKHAND", "lat": "22.8046", "lon": "86.2029"},
}


MAPPING_PATH = DATA / "am_jiomart_mapping.json"


def load_mapping(pincode=None):
    """Load shared mapping (URLs same across cities)."""
    if MAPPING_PATH.exists():
        return json.load(open(MAPPING_PATH))
    return {}


def save_mapping(mapping, pincode=None):
    """Save shared mapping."""
    MAPPING_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(MAPPING_PATH, "w") as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False, default=str)


async def init_browser(pw, pincode):
    loc = PINCODE_LOCATION.get(pincode, PINCODE_LOCATION["834002"])
    browser = await pw.firefox.launch(headless=True)
    context = await browser.new_context(
        user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:128.0) Gecko/20100101 Firefox/128.0",
        viewport={"width": 1366, "height": 768}, locale="en-IN", timezone_id="Asia/Kolkata",
    )
    loc_json = json.dumps({
        "country": "INDIA", "country_iso_code": "IN",
        "city": loc["city"], "pincode": loc["pincode"], "state": loc["state"],
    })
    await context.add_cookies([
        {"name": "pincode", "value": loc["pincode"], "domain": ".jiomart.com", "path": "/"},
        {"name": "app_location_details", "value": quote(loc_json), "domain": "www.jiomart.com", "path": "/"},
    ])
    await context.add_init_script(f"""
        localStorage.setItem('pin', JSON.stringify({{
            country: "INDIA", pincode: "{loc['pincode']}", city: "{loc['city']}", state: "{loc['state']}"
        }}));
        localStorage.setItem('jio_lat_long', JSON.stringify({{
            latitude: "{loc['lat']}", longitude: "{loc['lon']}"
        }}));
        Object.defineProperty(navigator, 'webdriver', {{get: () => undefined}});
    """)
    return browser, context


# ══════════════════════════════════════════════════════════════
#  PDP Fetch — open saved URL, get latest SP/MRP
# ══════════════════════════════════════════════════════════════

async def fetch_price_from_url(page, url):
    """Open a Jiomart product URL, extract latest SP and MRP."""
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
    except Exception:
        try:
            await asyncio.sleep(2)
            await page.goto(url, wait_until="domcontentloaded", timeout=25000)
        except Exception:
            return None, None, None, False

    await asyncio.sleep(2)

    # Extract from DOM
    data = await page.evaluate(r'''() => {
        const sp_el = document.querySelector('[class*=currentPrice], [class*=sellingPrice]');
        const mrp_el = document.querySelector('[class*=originalPrice], [class*=strikePrice]');
        const name_el = document.querySelector('h1, [class*=productTitle]');
        const sp = sp_el ? parseFloat(sp_el.textContent.replace(/[₹,\s]/g, '')) : null;
        const mrp = mrp_el ? parseFloat(mrp_el.textContent.replace(/[₹,\s]/g, '')) : null;
        const name = name_el ? name_el.textContent.trim() : '';
        // Check if product is available (no "out of stock" message)
        const oos = document.querySelector('[class*=outOfStock], [class*=soldOut]');
        return {sp: sp || null, mrp: mrp || sp || null, name, in_stock: !oos};
    }''')

    return data.get("sp"), data.get("mrp"), data.get("name"), data.get("in_stock", True)


# ══════════════════════════════════════════════════════════════
#  Search — find product on Jiomart
# ══════════════════════════════════════════════════════════════

async def search_jiomart(page, search_term, max_pages=3):
    """Search on Jiomart, scroll + paginate, return all products."""
    all_products = []
    seen = set()
    clean_term = re.sub(r'[-–/|&]', ' ', search_term)
    clean_term = re.sub(r'\s+', ' ', clean_term).strip()

    for pg in range(1, max_pages + 1):
        url = f"https://www.jiomart.com/products?q={quote(clean_term)}"
        if pg > 1:
            url += f"&page={pg}"
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=25000)
        except Exception:
            try:
                await asyncio.sleep(2)
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            except Exception:
                break
        await asyncio.sleep(2 + random.random() * 1.5)  # 2-3.5s random delay

        prev = 0
        for _ in range(10):
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(0.8 + random.random() * 0.5)
            count = await page.evaluate("document.querySelectorAll('[class*=productCard]').length")
            if count == prev:
                break
            prev = count

        dl = await page.evaluate('''() => {
            const p = [], s = new Set();
            if (window.dataLayer) for (const e of window.dataLayer)
                for (const i of (e?.ecommerce?.items || []))
                    if (i.product_url && !s.has(i.product_url)) {
                        s.add(i.product_url);
                        p.push({slug: i.product_url, product_id: String(i.item_id || ''),
                                name: i.item_name || '', brand: i.item_brand || '', sp: i.price || null});
                    }
            return p;
        }''')

        dom = await page.evaluate(r'''() => {
            const p = [];
            document.querySelectorAll('.productCard__productDescription, [class*=productCard]').forEach(card => {
                const nameEl = card.querySelector('.productCard__productTitle, h3, [class*=Title]');
                const spEl = card.querySelector('[class*=currentPrice]');
                const mrpEl = card.querySelector('[class*=originalPrice]');
                const linkEl = card.querySelector('a[href*="/product/"]') || card.closest('a[href*="/product/"]');
                const name = nameEl ? nameEl.textContent.trim() : '';
                const sp = spEl ? parseFloat(spEl.textContent.replace(/[₹,\s]/g, '')) : null;
                const mrp = mrpEl ? parseFloat(mrpEl.textContent.replace(/[₹,\s]/g, '')) : null;
                let slug = '';
                if (linkEl) {
                    const href = linkEl.getAttribute('href') || '';
                    const m = href.match(/\/product\/(.+?)(?:\?|$)/);
                    if (m) slug = m[1];
                }
                if (name && sp) p.push({name, sp, mrp: mrp || sp, slug});
            });
            return p;
        }''')

        for dl_p in dl:
            for dom_p in dom:
                if dl_p["name"] == dom_p["name"] and dl_p.get("sp") == dom_p.get("sp"):
                    dl_p["mrp"] = dom_p.get("mrp")
                    # Pick slug from DOM if dataLayer slug is missing
                    if not dl_p.get("slug") and dom_p.get("slug"):
                        dl_p["slug"] = dom_p["slug"]
                    break
            if "mrp" not in dl_p:
                dl_p["mrp"] = dl_p.get("sp")

        page_new = 0
        for p in dl:
            key = p.get("name", "")
            if key and key not in seen:
                seen.add(key)
                all_products.append(p)
                page_new += 1
        for p in dom:
            key = p.get("name", "")
            if key and key not in seen:
                seen.add(key)
                slug = p.get("slug", "")
                all_products.append({"slug": slug, "product_id": "",
                                     "name": key, "brand": "",
                                     "sp": p["sp"], "mrp": p["mrp"]})
                page_new += 1
        if page_new == 0:
            break

    return all_products


def build_search_term(am):
    """
    Build smart search term from AM product.
    Rules:
      - Remove weight, pack info, pipes (| 500g | Pack of 1)
      - Remove hyphens before weight (- 500g)
      - NO brand repeat — if brand already in name, don't prepend
      - Loose items: strip "Loose" prefix, keep product + weight
      - Max 5 words
    Examples:
      "Cadbury Dairy Milk Chocolate Bar | 11g | Pack of 1" → "Cadbury Dairy Milk Chocolate Bar"
      "Everest Chaat Masala - 50g" → "Everest Chaat Masala"
      "Haldirams Salted Peanuts 200g" → "Haldirams Salted Peanuts"
      "Loose Sugar - 1Kg" → "Sugar 1 kg"
    """
    name = am.get("display_name", "")
    brand = (am.get("brand") or "").strip()

    # Loose items: strip "Loose" prefix, keep product name with weight
    is_loose = "loose" in name.lower() or "asm" in brand.lower() or \
               brand.strip().upper() in ("ASM", "LOOSE CONTAINER")
    if is_loose:
        clean = re.sub(r'\s*\|.*$', '', name).strip()
        clean = re.sub(r'(?i)^(loose\s+)', '', clean).strip()
        clean = re.sub(r'[-–]', ' ', clean).strip()
        clean = re.sub(r'\s+', ' ', clean)
        return clean[:60]

    # Branded items: clean product name
    clean = re.sub(r'\s*\|.*$', '', name).strip()                    # Remove | suffix
    clean = re.sub(r'\s*[-–]\s*\d+\.?\d*\s*'                        # Remove "- 500g" weight suffix
                   r'(g|gm|gms|kg|kgs|ml|mls|l|ltr|ltrs|pcs?)\b.*$',
                   '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s*\d+\.?\d*\s*'                                # Remove trailing "500g"
                   r'(g|gm|gms|kg|kgs|ml|mls|l|ltr|ltrs)\s*$',
                   '', clean, flags=re.IGNORECASE).strip()
    clean = re.sub(r'[-–]', ' ', clean).strip()                      # Hyphens to spaces
    clean = re.sub(r'\s+', ' ', clean)                                # Collapse spaces

    # Take max 5 words
    words = clean.split()[:5]
    term = " ".join(words)

    # DON'T prepend brand if already in name (avoid "Everest Everest Chaat Masala")
    return term


def search_results_to_pool(search_results):
    """Convert search results to engine pool format."""
    pool = []
    for p in search_results:
        name = p.get("name", "")
        if not name:
            continue
        unit = ""
        m = re.search(r"(\d+\.?\d*)\s*(g|gm|gms|kg|kgs|ml|mls|l|ltr|ltrs|pc|pcs)\b", name.lower())
        if m:
            unit = f"{m.group(1)} {m.group(2)}"
        pool.append({
            "product_id": p.get("product_id", ""),
            "product_url": f"https://www.jiomart.com/product/{p['slug']}" if p.get("slug") else "",
            "product_name": name,
            "brand": p.get("brand", ""),
            "price": p.get("sp"),
            "mrp": p.get("mrp"),
            "unit": unit,
            "in_stock": True,
            "slug": p.get("slug", ""),
        })
    return pool


# ══════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════

async def main():
    args = sys.argv[1:]
    pincode = "834002"
    csv_path = None
    state_filter = None
    retry_na = False      # --retry-na: only re-search NA items from previous run
    mapped_only = False   # --mapped-only: skip search, only URL fetch (daily mode)
    prev_results_path = None

    for i, a in enumerate(args):
        if a == "--pincode" and i + 1 < len(args):
            pincode = args[i + 1]
        elif a == "--csv" and i + 1 < len(args):
            csv_path = args[i + 1]
        elif a == "--state" and i + 1 < len(args):
            state_filter = args[i + 1]
        elif a == "--retry-na":
            retry_na = True
        elif a == "--mapped-only":
            mapped_only = True
        elif a == "--prev" and i + 1 < len(args):
            prev_results_path = args[i + 1]
        elif a == "--tabs" and i + 1 < len(args):
            pass  # handled below
        elif a == "--limit" and i + 1 < len(args):
            pass  # handled below

    # Parse tabs and limit
    NUM_TABS_OVERRIDE = None
    LIMIT_NA = None
    for i, a in enumerate(args):
        if a == "--tabs" and i + 1 < len(args):
            NUM_TABS_OVERRIDE = int(args[i + 1])
        elif a == "--limit" and i + 1 < len(args):
            LIMIT_NA = int(args[i + 1])

    print(f"\n{'=' * 60}")
    print(f"  JIOMART PRICE FETCHER — {pincode}")
    print(f"{'=' * 60}\n")

    # ── Load data ──
    am_map = json.load(open(DATA / "am_product_master.json"))
    # MRP source: latestproductpricingtracker (fresh) → fallback to model 1808
    pricing_path = DATA / "am_pricing_wrhs_1.json"
    mrp_1808_path = DATA / "latest_mrp_wrhs_1.json"
    mrp_map = {}
    if pricing_path.exists():
        mrp_map = json.load(open(pricing_path))
        print(f"MRP source: am_pricing (latestproductpricingtracker) — {len(mrp_map)} items")
    elif mrp_1808_path.exists():
        mrp_map = json.load(open(mrp_1808_path))
        print(f"MRP source: model 1808 (fallback) — {len(mrp_map)} items")
    mapping = load_mapping(pincode)
    ean_map = json.load(open(DATA / "ean_map.json")) if (DATA / "ean_map.json").exists() else {}
    now = datetime.now().isoformat()

    print(f"AM master: {len(am_map)}")
    print(f"MRP data: {len(mrp_map)}")
    print(f"Existing mapping: {len(mapping)}")

    # ── Load target items ──
    target_items = {}
    if csv_path:
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                if state_filter and row.get("State", "").strip() != state_filter:
                    continue
                ic = str(row.get("Item Code", "")).strip()
                if ic:
                    am = am_map.get(ic, {})
                    target_items[ic] = {
                        "item_code": ic,
                        "display_name": row.get("Item Name", "") or am.get("display_name", ""),
                        "brand": row.get("Brand", "") or am.get("brand", ""),
                        "master_category": row.get("Master Catgory", "").strip() or am.get("master_category", ""),
                        "product_type": row.get("Product Type", "") or am.get("product_type", ""),
                        "unit": am.get("unit", ""),
                        "unit_value": am.get("unit_value"),
                        "mrp": am.get("mrp"),
                    }

    # If no CSV, use ALL AM master items (Docker/daily mode)
    if not target_items:
        valid_cats = {"STPLS", "FMCG", "FMCGF", "FMCGNF", "GM"}
        for ic, am in am_map.items():
            if am.get("master_category") in valid_cats:
                target_items[ic] = {
                    "item_code": ic,
                    "display_name": am.get("display_name", ""),
                    "brand": am.get("brand", ""),
                    "master_category": am.get("master_category", ""),
                    "product_type": am.get("product_type", ""),
                    "unit": am.get("unit", ""),
                    "unit_value": am.get("unit_value"),
                    "mrp": am.get("mrp"),
                }
        print(f"No CSV — using AM master ({len(target_items)} items)")

    total = len(target_items)
    print(f"Target items: {total}")

    # ── Retry NA mode: load previous results, keep matched, only search NA ──
    prev_matched = {}
    if retry_na:
        # Find latest previous results
        if not prev_results_path:
            import glob
            prev_files = sorted(glob.glob(str(DATA / f"jiomart_prices_{pincode}_*.json")))
            if prev_files:
                prev_results_path = prev_files[-1]
        if prev_results_path:
            prev_data = json.load(open(prev_results_path))
            for r in prev_data.get("results", []):
                ic = r.get("item_code", "")
                if r.get("match_status") != "NA" and r.get("jio_sp"):
                    prev_matched[ic] = r
            print(f"Retry NA mode: {len(prev_matched)} previously matched (kept as-is)")
            # Remove already matched from target
            na_only = {ic: am for ic, am in target_items.items() if ic not in prev_matched}
            print(f"NA to re-search: {len(na_only)}")
            target_items = na_only

    # Apply limit if set
    if LIMIT_NA and len(target_items) > LIMIT_NA:
        limited = dict(list(target_items.items())[:LIMIT_NA])
        print(f"Limited to {LIMIT_NA} items (from {len(target_items)})")
        target_items = limited

    # ── Load URLs from Excel (user-provided) ──
    url_excel = Path("/Users/satyam/Documents/url of jio mart.xlsx")
    if url_excel.exists():
        try:
            from openpyxl import load_workbook
            wb_urls = load_workbook(url_excel, read_only=True)
            ws_urls = wb_urls.active
            headers_url = [c.value for c in next(ws_urls.iter_rows(min_row=1, max_row=1))]
            ic_col = headers_url.index("Item Code") if "Item Code" in headers_url else 0
            url_col = headers_url.index("JIO URL") if "JIO URL" in headers_url else -1
            name_col = headers_url.index("JIO Name") if "JIO Name" in headers_url else -1
            sp_col = headers_url.index("JIO SP") if "JIO SP" in headers_url else -1
            mrp_col = headers_url.index("JIO MRP") if "JIO MRP" in headers_url else -1
            score_col = headers_url.index("Score") if "Score" in headers_url else -1

            excel_urls = 0
            for row in ws_urls.iter_rows(min_row=2, values_only=True):
                ic = str(row[ic_col]).strip() if row[ic_col] else ""
                url = str(row[url_col]).strip() if url_col >= 0 and row[url_col] else ""
                if ic and url and "jiomart.com" in url and ic not in mapping:
                    mapping[ic] = {
                        "jio_url": url,
                        "jio_name": str(row[name_col]).strip() if name_col >= 0 and row[name_col] else "",
                        "jio_brand": "",
                        "match_score": float(row[score_col]) if score_col >= 0 and row[score_col] else 0,
                        "matched_at": now,
                    }
                    excel_urls += 1
            print(f"URLs from Excel: {excel_urls} new")
        except Exception as e:
            print(f"Excel URL load error: {e}")

    # ── Split: mapped vs unmapped ──
    mapped_items = {}    # have saved URL → just fetch price
    unmapped_items = {}  # no URL → need search

    for ic, am in target_items.items():
        if ic in mapping and mapping[ic].get("jio_url"):
            mapped_items[ic] = am
        else:
            unmapped_items[ic] = am

    print(f"Mapped (URL fetch):   {len(mapped_items)}")
    print(f"Unmapped (search):    {len(unmapped_items)}")

    # ── Init browser ──
    NUM_TABS = NUM_TABS_OVERRIDE or 4
    from playwright.async_api import async_playwright
    pw = await async_playwright().start()
    browser, context = await init_browser(pw, pincode)

    pages = []
    for _ in range(NUM_TABS):
        p = await context.new_page()
        pages.append(p)

    # Set location on first tab
    try:
        await pages[0].goto("https://www.jiomart.com", wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(2)
    except Exception:
        pass

    results = {}
    engine = UnifiedMatchingEngine(ean_map)
    _lock = asyncio.Lock()  # protect shared results/mapping

    # ══════════════════════════════════════════════════════════
    #  PART 1: Fetch prices for mapped items (4 tabs parallel)
    # ══════════════════════════════════════════════════════════

    if mapped_items:
        print(f"\n--- Part 1: URL fetch for {len(mapped_items)} mapped items ({NUM_TABS} tabs) ---", flush=True)

    fetched = 0
    fetch_errors = 0
    url_queue = asyncio.Queue()
    for ic, am in mapped_items.items():
        await url_queue.put((ic, am))

    async def url_fetch_worker(page, worker_id):
        nonlocal fetched, fetch_errors
        while not url_queue.empty():
            try:
                ic, am = url_queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            m = mapping[ic]
            url = m["jio_url"]
            sp, mrp, name, in_stock = await fetch_price_from_url(page, url)
            await asyncio.sleep(random.uniform(0.5, 1.5))  # anti-block

            mrp_rec = mrp_map.get(ic)
            am_mrp = parse_num(mrp_rec.get("mrp")) if mrp_rec else parse_num(am.get("mrp"))

            async with _lock:
                if sp is not None:
                    status = engine.compute_status(am, am_mrp, sp, mrp, name or m.get("jio_name"))
                    results[ic] = {
                        "item_code": ic,
                        "am_name": am.get("display_name"),
                        "am_brand": am.get("brand"),
                        "am_unit": f"{am.get('unit_value', '')} {am.get('unit', '')}".strip(),
                        "am_mrp": am_mrp,
                        "jio_name": name or m.get("jio_name"),
                        "jio_brand": m.get("jio_brand", ""),
                        "jio_sp": sp, "jio_mrp": mrp,
                        "jio_product_id": m.get("jio_product_id"),
                        "jio_url": url,
                        "match_status": status,
                        "match_score": m.get("match_score", 0),
                        "match_reason": "url_fetch", "match_flags": "",
                        "in_stock": in_stock,
                    }
                    mapping[ic]["last_sp"] = sp
                    mapping[ic]["last_mrp"] = mrp
                    mapping[ic]["last_fetched"] = now
                    fetched += 1
                else:
                    fetch_errors += 1
                    unmapped_items[ic] = am

    if mapped_items:
        await asyncio.gather(*[url_fetch_worker(pages[i], i) for i in range(NUM_TABS)])
        print(f"  URL fetch done: {fetched} ok, {fetch_errors} errors", flush=True)

    # ══════════════════════════════════════════════════════════
    #  PART 2: Search for unmapped items (4 tabs parallel)
    # ══════════════════════════════════════════════════════════

    if mapped_only:
        # Daily mode: skip search entirely, mark unmapped as NA
        print(f"\n--- Part 2: SKIPPED (--mapped-only, {len(unmapped_items)} unmapped = NA) ---", flush=True)
        for ic, am in unmapped_items.items():
            am_mrp_val = parse_num(mrp_map.get(ic, {}).get("mrp")) or parse_num(am.get("mrp"))
            results[ic] = {
                "item_code": ic,
                "am_name": am.get("display_name"),
                "am_brand": am.get("brand"),
                "am_unit": f"{am.get('unit_value', '')} {am.get('unit', '')}".strip(),
                "am_mrp": am_mrp_val,
                "match_status": "NA",
                "match_score": 0,
                "match_reason": "unmapped_daily_skip",
                "match_flags": "",
            }
        unmapped_items = {}  # skip search below

    if unmapped_items:
        print(f"\n--- Part 2: Search for {len(unmapped_items)} unmapped items ({NUM_TABS} tabs) ---", flush=True)

    searched = 0
    search_errors = 0
    search_cache = {}
    _search_cache_lock = asyncio.Lock()

    # Build search tasks: (search_term, [(ic, am), ...])
    brand_groups = defaultdict(list)
    for ic, am in unmapped_items.items():
        brand = (am.get("brand") or "").strip()
        brand_key = normalize(brand) if brand else f"_no_brand_{ic}"
        brand_groups[brand_key].append((ic, am))

    sorted_groups = sorted(brand_groups.items(), key=lambda x: -len(x[1]))
    search_queue = asyncio.Queue()
    for gi, (brand_key, items) in enumerate(sorted_groups):
        brand_name = items[0][1].get("brand", "").strip()
        await search_queue.put((gi, brand_name, items))

    total_groups = len(sorted_groups)

    async def search_worker(page, worker_id):
        nonlocal searched, search_errors

        while not search_queue.empty():
            try:
                gi, brand_name, items = search_queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            # Brand search (with rate-limit retry)
            brand_results = []
            async with _search_cache_lock:
                cached = search_cache.get(brand_name)
            if cached is not None:
                brand_results = cached
            elif brand_name:
                try:
                    brand_results = await search_jiomart(page, brand_name, max_pages=3)
                    # Clean brand retry if special chars
                    if not brand_results:
                        clean_brand = re.sub(r'[-–\'\"&/]', ' ', brand_name).strip()
                        clean_brand = re.sub(r'\s+', ' ', clean_brand)
                        if clean_brand != brand_name:
                            brand_results = await search_jiomart(page, clean_brand, max_pages=3)
                            searched += 1
                    # Rate limit retry: 0 results = likely blocked, wait and retry
                    if not brand_results:
                        await asyncio.sleep(random.uniform(5, 10))
                        brand_results = await search_jiomart(page, brand_name, max_pages=3)
                        if not brand_results:
                            # Final retry with longer wait
                            await asyncio.sleep(random.uniform(10, 15))
                            brand_results = await search_jiomart(page, brand_name, max_pages=2)
                    searched += 1
                    print(f"[{gi+1}/{total_groups}] T{worker_id} Brand: \"{brand_name}\" → "
                          f"{len(brand_results)} ({len(items)} items)", flush=True)
                except Exception as e:
                    search_errors += 1
                    print(f"[{gi+1}/{total_groups}] T{worker_id} Brand: \"{brand_name}\" → "
                          f"ERROR: {str(e)[:50]}", flush=True)
                async with _search_cache_lock:
                    search_cache[brand_name] = brand_results

            # Set engine pool for this worker's brand
            local_engine = UnifiedMatchingEngine(ean_map)
            pool = search_results_to_pool(brand_results)
            with contextlib.redirect_stdout(io.StringIO()):
                local_engine.set_pool(pool)

            for ic, am in items:
                mrp_rec = mrp_map.get(ic)
                if mrp_rec:
                    am["mrp"] = mrp_rec.get("mrp", am.get("mrp"))
                am_mrp = parse_num(am.get("mrp"))

                result = local_engine.match(am, am_mrp)
                matched_p = result.product
                score = result.score
                status = result.status

                # Name search fallback
                if status == "NA" or (status == "PARTIAL MATCH" and score < 0.55):
                    search_term = build_search_term(am)
                    name_results = None
                    async with _search_cache_lock:
                        name_results = search_cache.get(search_term)
                    if name_results is None:
                        try:
                            name_results = await search_jiomart(page, search_term, max_pages=2)
                            # Rate limit retry
                            if not name_results:
                                await asyncio.sleep(random.uniform(5, 8))
                                name_results = await search_jiomart(page, search_term, max_pages=2)
                            searched += 1
                        except Exception:
                            name_results = []
                            search_errors += 1
                        async with _search_cache_lock:
                            search_cache[search_term] = name_results

                    if name_results:
                        pool2 = search_results_to_pool(name_results)
                        with contextlib.redirect_stdout(io.StringIO()):
                            local_engine.set_pool(pool2)
                        result2 = local_engine.match(am, am_mrp)
                        if result2.score > score:
                            matched_p = result2.product
                            score = result2.score
                            status = result2.status
                            result = result2

                r = {
                    "item_code": ic,
                    "am_name": am.get("display_name"),
                    "am_brand": am.get("brand"),
                    "am_unit": f"{am.get('unit_value', '')} {am.get('unit', '')}".strip(),
                    "am_mrp": am_mrp,
                    "match_status": status,
                    "match_score": round(score, 3),
                    "match_reason": result.reason,
                    "match_flags": ", ".join(result.flags) if result.flags else "",
                }

                if matched_p:
                    sp = parse_num(matched_p.get("price")) or parse_num(matched_p.get("sp"))
                    slug = matched_p.get("slug", "")
                    jio_url = matched_p.get("product_url", "")
                    if not jio_url and slug:
                        jio_url = f"https://www.jiomart.com/product/{slug}"

                    r.update({
                        "jio_name": matched_p.get("product_name"),
                        "jio_brand": matched_p.get("brand"),
                        "jio_sp": sp,
                        "jio_mrp": matched_p.get("mrp"),
                        "jio_product_id": matched_p.get("product_id"),
                        "jio_url": jio_url,
                    })

                    async with _lock:
                        mapping[ic] = {
                            "jio_product_id": matched_p.get("product_id", ""),
                            "jio_slug": slug,
                            "jio_url": jio_url,
                            "jio_name": matched_p.get("product_name"),
                            "jio_brand": matched_p.get("brand", ""),
                            "match_status": status,
                            "match_score": round(score, 3),
                            "matched_at": now,
                            "last_sp": sp,
                            "last_mrp": matched_p.get("mrp"),
                            "last_fetched": now,
                        }

                async with _lock:
                    results[ic] = r

            # Anti-block: small delay between brand groups
            await asyncio.sleep(random.uniform(1, 3))

        # Checkpoint at end of worker
        async with _lock:
            save_mapping(mapping, pincode)
            matched_so_far = sum(1 for r in results.values() if r["match_status"] != "NA")
            print(f"  [worker {worker_id} done] {len(results)}/{total}, "
                  f"{matched_so_far} matched", flush=True)

    if unmapped_items:
        await asyncio.gather(*[search_worker(pages[i], i) for i in range(NUM_TABS)])

    await browser.close()
    await pw.stop()

    # ── Merge prev_matched + new results ──
    if prev_matched:
        for ic, r in prev_matched.items():
            if ic not in results:
                results[ic] = r
        print(f"\nMerged: {len(prev_matched)} prev matched + {len(results) - len(prev_matched)} new")

    # ── Save everything ──
    save_mapping(mapping, pincode)
    print(f"Mapping saved: {len(mapping)} entries → am_jiomart_mapping_{pincode}.json")

    # Save results JSON (single file, all results)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    json_path = DATA / f"jiomart_prices_{pincode}_{ts}.json"
    all_total = len(results)
    with open(json_path, "w") as f:
        json.dump({
            "pincode": pincode, "fetched_at": now,
            "total": all_total, "results": list(results.values()),
        }, f, indent=2, default=str, ensure_ascii=False)
    print(f"Results JSON: {json_path.name}")

    # Save PDP-compatible file (for sam_daily_run.py load_pdp compatibility)
    pdp_products = []
    for r in results.values():
        if r.get("jio_sp") is not None:
            pdp_products.append({
                "item_code": r["item_code"],
                "status": "ok",
                "sam_product_name": r.get("jio_name"),
                "sam_selling_price": r.get("jio_sp"),
                "sam_mrp": r.get("jio_mrp"),
                "sam_in_stock": r.get("in_stock", True),
                "sam_unit": r.get("am_unit", ""),
                "jiomart_product_url": r.get("jio_url", ""),
            })
        else:
            pdp_products.append({
                "item_code": r["item_code"],
                "status": "not_available",
            })
    pdp_path = DATA / "sam" / f"jiomart_pdp_{pincode}_{ts}.json"
    pdp_path.parent.mkdir(parents=True, exist_ok=True)
    with open(pdp_path, "w") as f:
        json.dump({"products": pdp_products, "pincode": pincode, "fetched_at": now},
                  f, indent=2, default=str, ensure_ascii=False)
    print(f"PDP compat: {pdp_path.name}")

    # Save Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Jiomart_Prices"
        headers = ["Item Code", "AM Name", "Brand", "AM Unit", "AM MRP",
                    "JIO Name", "JIO SP", "JIO MRP", "JIO URL",
                    "Match Status", "Score", "Reason", "Flags"]
        am_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
        jio_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, size=10)
            if i <= 5: c.fill = am_fill
            elif i <= 9: c.fill = jio_fill

        for idx, r in enumerate(sorted(results.values(),
                key=lambda x: int(x["item_code"]) if x["item_code"].isdigit() else 0), 2):
            ic = r["item_code"]
            ws.cell(row=idx, column=1, value=int(ic) if ic.isdigit() else ic)
            ws.cell(row=idx, column=2, value=r.get("am_name"))
            ws.cell(row=idx, column=3, value=r.get("am_brand"))
            ws.cell(row=idx, column=4, value=r.get("am_unit"))
            ws.cell(row=idx, column=5, value=r.get("am_mrp"))
            ws.cell(row=idx, column=6, value=r.get("jio_name"))
            ws.cell(row=idx, column=7, value=r.get("jio_sp"))
            ws.cell(row=idx, column=8, value=r.get("jio_mrp"))
            ws.cell(row=idx, column=9, value=r.get("jio_url"))
            ws.cell(row=idx, column=10, value=r.get("match_status"))
            ws.cell(row=idx, column=11, value=r.get("match_score"))
            ws.cell(row=idx, column=12, value=r.get("match_reason"))
            ws.cell(row=idx, column=13, value=r.get("match_flags"))
            sc = ws.cell(row=idx, column=10)
            if r["match_status"] == "COMPLETE MATCH":
                sc.fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
            elif r["match_status"] == "SEMI COMPLETE MATCH":
                sc.fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
            elif r["match_status"] == "PARTIAL MATCH":
                sc.fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")

        for col in ws.columns:
            mx = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 50)
        ws.freeze_panes = "A2"

        xlsx_path = Path("/Users/satyam/Desktop/price csv") / f"Jiomart_Prices_{pincode}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(xlsx_path)
        print(f"Excel: {xlsx_path}")
    except ImportError:
        print("openpyxl not installed — skipping Excel")

    # ── Summary ──
    status_counts = defaultdict(int)
    for r in results.values():
        status_counts[r["match_status"]] += 1

    print(f"\n{'=' * 60}")
    print(f"  RESULTS — {pincode}")
    print(f"{'=' * 60}")
    print(f"  Total:          {total}")
    print(f"  URL fetched:    {fetched}")
    print(f"  Searched:       {searched}")
    print(f"  Mapping saved:  {len(mapping)}")
    print()
    for s, c in sorted(status_counts.items(), key=lambda x: -x[1]):
        print(f"    {s:25s} {c:5d}  ({c / total * 100:.1f}%)")

    matched = sum(1 for r in results.values() if r["match_status"] != "NA")
    print(f"\n  Matched: {matched}/{total} ({matched / total * 100:.1f}%)")
    print(f"\n  Next run: mapped items = URL fetch only (fast). No re-search.")


if __name__ == "__main__":
    asyncio.run(main())

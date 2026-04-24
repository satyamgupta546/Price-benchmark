"""
Export Excel from BigQuery — pull SAM data and generate formatted Excel.

Usage:
    python scripts/export_excel_from_bq.py 834002              # today's data for Ranchi
    python scripts/export_excel_from_bq.py 834002 2026-04-24   # specific date
    python scripts/export_excel_from_bq.py all                  # all cities, today
    python scripts/export_excel_from_bq.py all 2026-04-24       # all cities, specific date
"""
import json
import subprocess
import sys
import csv
import io
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = Path("/Users/satyam/Desktop/price csv")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

_cities_config = json.load(open(PROJECT_ROOT / "config" / "cities.json"))
CITIES = {pin: cfg["name"] for pin, cfg in _cities_config["cities"].items()}

BQ_PROJECT = "apna-mart-data"
BQ_DATASET = "googlesheet"

HEADERS = [
    "DATE", "TIME", "CITY", "PINCODE",
    "AM ITEM CODE", "AM ITEM NAME", "AM master cat", "AM BRAND", "AM MARKETED BY",
    "AM PRODUCT TYPE", "AM UNIT", "AM UNIT VALUE", "AM MRP", "IMAGE LINK",
    "BLINKIT URL", "BLINKIT ITEM NAME", "BLINKIT UNIT", "BLINKIT MRP", "BLINKIT SP",
    "BLINKIT IN STOCK REMARK", "BLINKIT STATUS",
    "JIO URL", "JIO ITEM NAME", "JIO UNIT", "JIO MRP", "JIO SP",
    "JIO IN STOCK REMARK", "JIO STATUS",
    "DMART URL", "DMART ITEM NAME", "DMART UNIT", "DMART MRP", "DMART SP",
    "DMART IN STOCK REMARK", "DMART STATUS",
]


def bq_query(sql):
    """Run BQ query, return rows as list of lists."""
    r = subprocess.run(
        ["bq", "query", "--use_legacy_sql=false", "--format=csv", "--max_rows=100000", sql],
        capture_output=True, text=True, timeout=120,
    )
    if r.returncode != 0:
        print(f"BQ error: {r.stderr[:300]}", flush=True)
        return []
    reader = csv.reader(io.StringIO(r.stdout))
    rows = list(reader)
    if len(rows) <= 1:
        return []
    return rows[1:]  # skip header


def generate_excel(rows, city, pincode, date):
    """Generate formatted Excel from BQ rows."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl", flush=True)
        return None

    am_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
    blinkit_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    jio_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    dmart_fill = PatternFill(start_color="FFE6CCFF", end_color="FFE6CCFF", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = f"SAM_{city}_{date}"

    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, size=10)
        if 1 <= i <= 14: c.fill = am_fill
        elif 15 <= i <= 21: c.fill = blinkit_fill
        elif 22 <= i <= 28: c.fill = jio_fill
        elif 29 <= i <= 35: c.fill = dmart_fill

    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            if val == "" or val == "null":
                val = None
            elif c_idx in (5, 12, 13, 18, 19, 25, 26, 32, 33):
                try:
                    val = float(val) if "." in str(val) else int(val)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=r_idx, column=c_idx, value=val)

    widths = [10, 18, 10, 8, 10, 40, 8, 15, 20, 20, 5, 8, 8, 30,
              35, 40, 10, 8, 8, 12, 20, 35, 40, 10, 8, 8, 12, 20,
              35, 40, 10, 8, 8, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"

    out_path = OUTPUT_DIR / f"SAM_{city}_{pincode}_{date}.xlsx"
    wb.save(out_path)
    return out_path


def main():
    pincode_arg = sys.argv[1] if len(sys.argv) > 1 else "all"
    date = sys.argv[2] if len(sys.argv) > 2 else datetime.now().strftime("%Y-%m-%d")

    pincodes = CITIES if pincode_arg == "all" else {pincode_arg: CITIES.get(pincode_arg, pincode_arg)}

    print(f"Exporting Excel from BigQuery — date={date}, cities={', '.join(pincodes.values())}", flush=True)

    for pin, city in pincodes.items():
        sql = (
            f"SELECT * FROM `{BQ_PROJECT}.{BQ_DATASET}.sam_price_history` "
            f"WHERE date = '{date}' AND pincode = '{pin}' "
            f"ORDER BY item_code"
        )
        rows = bq_query(sql)
        if not rows:
            print(f"  {city} ({pin}): no data for {date}", flush=True)
            continue

        out = generate_excel(rows, city, pin, date)
        if out:
            print(f"  {city}: {len(rows)} rows -> {out}", flush=True)

    print("Done!", flush=True)


if __name__ == "__main__":
    main()

"""
Daily Report Generator — THE master script. Run this once, get everything.

What it does:
  1. Fetches fresh Anakin data (all cities, both platforms)
  2. Runs full 6-stage pipeline (PDP → Brand → Type/MRP → Search → Image → Barcode)
  3. Generates Excel per city with sheets:
     - Summary (stats)
     - Blinkit (Anakin data + SAM data + difference)
     - Jiomart (Anakin data + SAM data + difference)
  4. Copies to ~/Downloads/

Usage:
    export METABASE_API_KEY=...
    cd backend && ./venv/bin/python ../scripts/daily_report.py
    cd backend && ./venv/bin/python ../scripts/daily_report.py 834002   # single city
"""
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(line_buffering=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
VENV_PYTHON = str(PROJECT_ROOT / "backend" / "venv" / "bin" / "python")
SCRIPTS = PROJECT_ROOT / "scripts"
DATA = PROJECT_ROOT / "data"

CITIES = {"834002": "Ranchi", "712232": "Kolkata", "492001": "Raipur", "825301": "Hazaribagh"}
DATE = datetime.now().strftime("%Y-%m-%d")


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("na", "nan", "null", "none"):
        return ""
    return s


def run(script, args=[], use_venv=False):
    python = VENV_PYTHON if use_venv else sys.executable
    cmd = [python, str(SCRIPTS / script)] + args
    print(f"  ▶ {script} {' '.join(args)}", flush=True)
    result = subprocess.run(cmd, cwd=str(PROJECT_ROOT), capture_output=True, text=True)
    if result.returncode != 0:
        print(f"    ⚠️ {script} failed (exit {result.returncode}): {(result.stderr or '')[:200]}", flush=True)


def step1_fetch_anakin(pincode):
    """Fetch fresh Anakin data + EAN map for both platforms."""
    print(f"\n📥 Step 1: Fetch Anakin data + EAN", flush=True)
    run("fetch_anakin_blinkit.py", [pincode])
    run("fetch_anakin_jiomart.py", [pincode])
    run("fetch_ean_map.py")


def step2_run_pipeline(pincode, platform):
    """Run full 6-stage pipeline for one platform."""
    print(f"\n⚙️  Step 2: {platform} pipeline", flush=True)

    if platform == "blinkit":
        # Stage 1: PDP
        run("scrape_blinkit_pdps.py", [pincode, "2"], use_venv=True)
        # Clean partial
        partial = DATA / "sam" / f"blinkit_pdp_{pincode}_latest_partial.json"
        if partial.exists():
            partial.unlink()
        run("compare_pdp.py", [pincode])
    elif platform == "jiomart":
        run("scrape_jiomart_pdps.py", [pincode, "2"], use_venv=True)
        partial = DATA / "sam" / f"jiomart_pdp_{pincode}_latest_partial.json"
        if partial.exists():
            partial.unlink()
        run("compare_pdp_jiomart.py", [pincode])

    # Stage 2-3
    run("cascade_match.py", [pincode, platform])
    run("stage3_match.py", [pincode, platform])

    # Stage 4: Jiomart search
    if platform == "jiomart":
        run("jiomart_search_match.py", [pincode], use_venv=True)

    # Stage 5: Image + Barcode
    run("stage4_image_match.py", [pincode, platform])
    run("stage5_barcode_match.py", [pincode, platform])


def load_anakin(platform, pincode) -> dict:
    """Load latest Anakin data, return dict keyed by item_code."""
    files = sorted((DATA / "anakin").glob(f"{platform}_{pincode}_*.json"))
    if not files:
        return {}
    d = json.load(open(files[-1]))
    return {r.get("Item_Code"): r for r in d.get("records", []) if r.get("Item_Code")}


def load_sam_prices(platform, pincode) -> dict:
    """Load latest SAM PDP prices, keyed by item_code."""
    prefix = platform
    files = sorted((DATA / "sam").glob(f"{prefix}_pdp_{pincode}_*.json"))
    files = [f for f in files if "partial" not in f.name]
    if not files:
        return {}
    d = json.load(open(files[-1]))
    prices = {}
    for p in d.get("products", []):
        ic = p.get("item_code")
        if ic and p.get("status") == "ok":
            name = p.get("sam_product_name") or p.get("hmlg_product_name") or ""
            # Skip PDP items with broken names (Google Retail raw IDs) — search will replace
            if name.startswith("projects/"):
                continue
            sam_sp_val = p.get("sam_selling_price")
            if sam_sp_val is None:
                sam_sp_val = p.get("hmlg_selling_price")
            sam_mrp_val = p.get("sam_mrp")
            if sam_mrp_val is None:
                sam_mrp_val = p.get("hmlg_mrp")
            prices[ic] = {
                "sam_name": name,
                "sam_sp": sam_sp_val,
                "sam_mrp": sam_mrp_val,
                "sam_stock": "available" if (p.get("sam_in_stock") or p.get("hmlg_in_stock")) else "out_of_stock",
            }
    return prices


def load_cascade_matches(platform, pincode) -> dict:
    """Load cascade/stage3/search/image/barcode matches with prices, keyed by item_code.
    Returns dict: {item_code: {sam_name, sam_sp, sam_mrp, match_method, score}}
    Only overwrites if new match has a HIGHER score (prevents stage3 replacing good cascade match)."""
    matched = {}
    stages = [
        (f"{platform}_cascade_{pincode}_*.json", "Cascade (Stage 2)"),
        (f"{platform}_stage3_{pincode}_*.json", "Type/MRP (Stage 3)"),
        (f"{platform}_image_match_{pincode}_*.json", "Image (Stage 5)"),
        (f"{platform}_barcode_match_{pincode}_*.json", "Barcode (Stage 6)"),
    ]
    if platform == "jiomart":
        stages.insert(2, (f"jiomart_search_match_{pincode}_*.json", "Search (Stage 4)"))

    for pat, method in stages:
        files = sorted((DATA / "comparisons").glob(pat))
        if not files:
            continue
        d = json.load(open(files[-1]))  # latest file only
        for m in d.get("new_mappings", []):
            ic = m.get("item_code")
            sp = m.get("sam_price")
            if ic is None or sp is None:
                continue
            score = m.get("cascade_score")
            if score is None:
                score = m.get("stage3_score")
            if score is None:
                score = m.get("match_score", 0)
            # Only overwrite if this match has a higher score
            if ic in matched and matched[ic]["match_score"] >= score:
                continue
            matched[ic] = {
                "sam_name": m.get("sam_product_name", ""),
                "sam_sp": sp,
                "sam_mrp": m.get("sam_mrp"),
                "sam_stock": "available",
                "match_method": method,
                "match_score": score,
            }
    return matched


def step3_generate_excel(pincode, city):
    """Generate Excel with Anakin + SAM + Diff side by side."""
    print(f"\n📊 Step 3: Generate Excel for {city}", flush=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  ⚠️ openpyxl not installed — CSV only")
        return

    wb = Workbook()

    for platform in ["blinkit", "jiomart"]:
        if pincode == "825301" and platform == "jiomart":
            continue

        anakin = load_anakin(platform, pincode)
        sam_prices = load_sam_prices(platform, pincode)
        cascade_matched = load_cascade_matches(platform, pincode)

        if not anakin:
            continue

        # Platform field mapping
        if platform == "blinkit":
            pf = {"sp": "Blinkit_Selling_Price", "mrp": "Blinkit_Mrp_Price",
                  "name": "Blinkit_Item_Name", "stock": "Blinkit_In_Stock_Remark",
                  "status": "Blinkit_Status", "pid": "Blinkit_Product_Id"}
        else:
            pf = {"sp": "Jiomart_Selling_Price", "mrp": "Jiomart_Mrp_Price",
                  "name": "Jiomart_Item_Name", "stock": "Jiomart_In_Stock_Remark",
                  "status": "Jiomart_Status", "pid": "Jiomart_Product_Id"}

        sheet_name = f"sam_{city.lower()[:6]}_{platform}_{DATE}"
        ws = wb.create_sheet(sheet_name)

        # Headers
        headers = [
            "Item_Code", "Apna Name", "Brand", "Unit", "MRP",
            f"Anakin {platform.capitalize()} Name", f"Anakin SP", f"Anakin MRP", "Anakin Stock", "Anakin Status",
            f"SAM {platform.capitalize()} Name", f"SAM {platform.capitalize()} SP", f"SAM MRP", "SAM Stock", "SAM Match Method",
            "Price Diff", "Diff %", "Match?"
        ]

        # Header styling
        anakin_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
        sam_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
        diff_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
        green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, size=10)
            if 6 <= i <= 10:
                c.fill = anakin_fill
            elif 11 <= i <= 15:
                c.fill = sam_fill
            elif 16 <= i <= 18:
                c.fill = diff_fill

        # Data rows
        row_idx = 2

        for ic, rec in sorted(anakin.items()):
            ana_sp_raw = clean(rec.get(pf["sp"]))
            ana_sp = None
            try:
                ana_sp = float(ana_sp_raw) if ana_sp_raw else None
            except ValueError:
                pass

            # Stage 1: PDP prices
            sam = sam_prices.get(ic, {})
            sam_sp = sam.get("sam_sp")
            sam_mrp = sam.get("sam_mrp")
            sam_name = sam.get("sam_name", "")
            sam_stock = sam.get("sam_stock", "")
            match_method = "PDP (Stage 1)" if sam_sp is not None else ""

            # Stage 2/3/4: Cascade/Stage3/Search prices (fallback when PDP has no price)
            if sam_sp is None and ic in cascade_matched:
                cm = cascade_matched[ic]
                sam_sp = cm.get("sam_sp")
                sam_mrp = cm.get("sam_mrp")
                sam_name = cm.get("sam_name", "")
                sam_stock = cm.get("sam_stock", "")
                match_method = cm.get("match_method", "Cascade")

            # Price diff
            diff = None
            diff_pct = None
            match_ok = ""
            if ana_sp is not None and sam_sp is not None:
                diff = round(sam_sp - ana_sp, 2)
                if ana_sp != 0:
                    diff_pct = round(abs(diff) / ana_sp * 100, 1)
                    match_ok = "✅" if diff_pct <= 5 else ("🟡" if diff_pct <= 10 else "❌")

            ana_mrp_raw = clean(rec.get(pf["mrp"]))
            ana_mrp = None
            try:
                ana_mrp = float(ana_mrp_raw) if ana_mrp_raw else None
            except ValueError:
                pass

            row_data = [
                ic, clean(rec.get("Item_Name")), clean(rec.get("Brand")),
                f"{clean(rec.get('Unit_Value'))} {clean(rec.get('Unit'))}".strip(),
                clean(rec.get("Mrp")),
                clean(rec.get(pf["name"])), ana_sp, ana_mrp,
                clean(rec.get(pf["stock"])), clean(rec.get(pf["status"])),
                sam_name, sam_sp, sam_mrp, sam_stock, match_method,
                diff, f"{diff_pct}%" if diff_pct is not None else "", match_ok,
            ]

            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=c_idx, value=val)
                if c_idx == 18:  # Match column
                    if val == "✅":
                        cell.fill = green_fill
                    elif val == "❌":
                        cell.fill = red_fill

            row_idx += 1

        # Column widths
        widths = [10, 40, 15, 10, 8, 40, 8, 8, 12, 15, 40, 8, 8, 12, 18, 8, 8, 6]
        for i, w in enumerate(widths, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "B2"

    # Summary sheet
    ws_s = wb.create_sheet(f"sam_{city.lower()[:6]}_{DATE}", 0)
    ws_s.cell(row=1, column=1, value="SAM Daily Report").font = Font(bold=True, size=14)
    ws_s.cell(row=2, column=1, value=f"City: {city} ({pincode})")
    ws_s.cell(row=3, column=1, value=f"Date: {DATE}")
    ws_s.cell(row=4, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M IST')}")

    row = 6
    for platform in ["blinkit", "jiomart"]:
        if pincode == "825301" and platform == "jiomart":
            continue
        anakin_s = load_anakin(platform, pincode)
        sam_pdp = load_sam_prices(platform, pincode)
        cascade_s = load_cascade_matches(platform, pincode)
        pf_sp = "Blinkit_Selling_Price" if platform == "blinkit" else "Jiomart_Selling_Price"
        usable_codes = set(
            ic for ic, r in anakin_s.items()
            if clean(r.get(pf_sp)) and "loose" not in (r.get("Item_Name") or "").lower()
        )
        usable = len(usable_codes)
        pdp_priced = sum(1 for ic in usable_codes if ic in sam_pdp)
        cascade_priced = sum(1 for ic in usable_codes if ic not in sam_pdp and ic in cascade_s)
        total_priced = pdp_priced + cascade_priced
        coverage_pct = min(round(total_priced * 100 / max(usable, 1), 1), 100.0)

        ws_s.cell(row=row, column=1, value=platform.capitalize()).font = Font(bold=True, size=12)
        ws_s.cell(row=row + 1, column=1, value="  Anakin usable (non-loose)")
        ws_s.cell(row=row + 1, column=2, value=usable)
        ws_s.cell(row=row + 2, column=1, value="  Stage 1 (PDP)")
        ws_s.cell(row=row + 2, column=2, value=pdp_priced)
        ws_s.cell(row=row + 3, column=1, value="  Stage 2/3/4 (Cascade+Search)")
        ws_s.cell(row=row + 3, column=2, value=cascade_priced)
        ws_s.cell(row=row + 4, column=1, value="  Total SAM priced")
        ws_s.cell(row=row + 4, column=2, value=total_priced)
        ws_s.cell(row=row + 5, column=1, value="  Coverage")
        ws_s.cell(row=row + 5, column=2, value=f"{coverage_pct}%")
        row += 7

    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 15

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save
    out_dir = DATA / "sam_output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"SAM_{city}_{pincode}_{DATE}.xlsx"
    wb.save(out_path)
    print(f"  ✅ {out_path.name}", flush=True)
    return out_path


def _process_one_city(pincode, city, skip_scrape):
    """Process a single city — fetch + scrape + excel. Called per-thread."""
    print(f"\n{'─' * 60}", flush=True)
    print(f"  {city} ({pincode})", flush=True)
    print(f"{'─' * 60}", flush=True)

    step1_fetch_anakin(pincode)

    if not skip_scrape:
        for platform in ["blinkit", "jiomart"]:
            if pincode == "825301" and platform == "jiomart":
                continue
            step2_run_pipeline(pincode, platform)

    return step3_generate_excel(pincode, city)


def main():
    pincode_arg = sys.argv[1] if len(sys.argv) > 1 else "all"
    skip_scrape = "--no-scrape" in sys.argv

    pincodes = CITIES if pincode_arg == "all" else {pincode_arg: CITIES.get(pincode_arg, pincode_arg)}

    print(f"{'═' * 60}")
    print(f"  SAM DAILY REPORT — {DATE}")
    print(f"  Cities: {', '.join(pincodes.values())}")
    print(f"  Scrape: {'skip' if skip_scrape else 'yes'}")
    print(f"  Mode: {'PARALLEL' if len(pincodes) > 1 else 'single'}")
    print(f"{'═' * 60}")

    output_files = []

    if len(pincodes) > 1 and not skip_scrape:
        # PARALLEL — all cities at once (each city in its own thread)
        from concurrent.futures import ThreadPoolExecutor, as_completed
        print(f"\n🚀 Running {len(pincodes)} cities in PARALLEL...", flush=True)

        with ThreadPoolExecutor(max_workers=len(pincodes)) as executor:
            futures = {
                executor.submit(_process_one_city, pin, city, skip_scrape): (pin, city)
                for pin, city in pincodes.items()
            }
            for future in as_completed(futures):
                pin, city = futures[future]
                try:
                    out = future.result()
                    if out:
                        output_files.append(out)
                    print(f"  ✅ {city} complete", flush=True)
                except Exception as e:
                    print(f"  ❌ {city} failed: {e}", flush=True)
    else:
        # Sequential (single city or --no-scrape)
        for pincode, city in pincodes.items():
            out = _process_one_city(pincode, city, skip_scrape)
            if out:
                output_files.append(out)

    # Copy to Downloads
    import shutil
    downloads = Path.home() / "Downloads"
    print(f"\n📁 Copying to ~/Downloads/")
    for f in output_files:
        dest = downloads / f.name
        shutil.copy(f, dest)
        print(f"  ✅ {dest.name}")

    # Also copy CSVs
    for f in (DATA / "sam_output").glob(f"sam_competitor_prices_*_{DATE}.csv"):
        dest = downloads / f.name
        shutil.copy(f, dest)

    print(f"\n{'═' * 60}")
    print(f"  DONE! {len(output_files)} Excel files in ~/Downloads/")
    print(f"{'═' * 60}")


if __name__ == "__main__":
    main()

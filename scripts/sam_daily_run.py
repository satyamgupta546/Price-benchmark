"""
SAM Daily Run — THE master script. Cron this at 10:30 AM IST.

What it does:
  1. Fetch AM product master (smpcm_product) + latest MRP (model 1808)
  2. Fetch EAN map
  3. Fetch Anakin data (all cities, both platforms)
  4. Scrape Blinkit + Jiomart in PARALLEL (all stages per city)
  5. Compute match status (COMPLETE/SEMI COMPLETE/PARTIAL/NA)
  6. Generate Excel per city → /Users/satyam/Desktop/price csv/
  7. Push to BigQuery (sam_price_live = replace, sam_price_history = append)

Schedule: Daily 8:30 AM IST via Cloud Run Job

Usage:
    export METABASE_API_KEY=...
    python3 scripts/sam_daily_run.py              # all cities
    python3 scripts/sam_daily_run.py 834002       # single city
    python3 scripts/sam_daily_run.py --no-scrape  # skip scrape, just regenerate Excel + push
"""
import csv
import json
import os
import re
import subprocess
import sys
import threading
import time
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(line_buffering=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
_venv_path = PROJECT_ROOT / "backend" / "venv" / "bin" / "python"
VENV_PYTHON = str(_venv_path) if _venv_path.exists() else sys.executable
SCRIPTS = PROJECT_ROOT / "scripts"
DATA = PROJECT_ROOT / "data"
OUTPUT_DIR = Path(os.environ.get("SAM_OUTPUT_DIR", str(PROJECT_ROOT / "output")))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

VALID_MASTER_CATEGORIES = {"STPLS", "FMCG", "FMCGF", "FMCGNF", "GM"}

# Load city + platform config from config/cities.json
_cities_config = json.load(open(PROJECT_ROOT / "config" / "cities.json"))
CITIES = {pin: cfg["name"] for pin, cfg in _cities_config["cities"].items()}
WAREHOUSE_MAP = {pin: cfg["warehouse"] for pin, cfg in _cities_config["cities"].items()}
CITY_PLATFORMS = {pin: set(cfg["platforms"]) for pin, cfg in _cities_config["cities"].items()}
DMART_STORE_IDS = _cities_config.get("dmart_store_ids", {})
DATE = datetime.now().strftime("%Y-%m-%d")
NOW = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

URL_DATABASE_PATH = DATA / "mappings" / "url_database.json"

METABASE_API = "https://mirror.apnamart.in/api"
METABASE_KEY = os.environ.get("METABASE_API_KEY", "")

BQ_PROJECT = "apna-mart-data"
BQ_DATASET = "googlesheet"
BQ_LIVE_TABLE = f"{BQ_PROJECT}:{BQ_DATASET}.sam_price_live"
GCS_BUCKET = "sam-price-data"


RUN_ERRORS = []  # Collect errors across the entire run


def run(script, args=[], use_venv=False, retries=2, critical=False, timeout=7200):
    """Run a script with retry logic. Collects errors for end-of-run alert."""
    python = VENV_PYTHON if use_venv else sys.executable
    cmd = [python, str(SCRIPTS / script)] + args
    print(f"  ▶ {script} {' '.join(args)}", flush=True)

    for attempt in range(retries + 1):
        try:
            r = subprocess.run(cmd, cwd=str(PROJECT_ROOT), capture_output=True, text=True,
                               timeout=timeout)
        except subprocess.TimeoutExpired:
            err_msg = f"Timed out after 120 min"
            print(f"    ⚠️ {script} timed out (attempt {attempt+1}/{retries+1})", flush=True)
            if attempt < retries:
                wait = 10 * (attempt + 1)
                time.sleep(wait)
                continue
            else:
                error = f"{script} {' '.join(args)} failed after {retries+1} attempts: {err_msg}"
                print(f"    ❌ {error}", flush=True)
                RUN_ERRORS.append(error)
                if critical:
                    from alert import send_alert, AlertLevel
                    send_alert(AlertLevel.ERROR, f"{script} timed out", details=err_msg)
                return False
        if r.returncode == 0:
            return True
        err_msg = (r.stderr or r.stdout or "")[:200]
        if attempt < retries:
            wait = 10 * (attempt + 1)
            print(f"    ⚠️ {script} failed (attempt {attempt+1}/{retries+1}), retry in {wait}s: {err_msg}", flush=True)
            time.sleep(wait)
        else:
            error = f"{script} {' '.join(args)} failed after {retries+1} attempts: {err_msg}"
            print(f"    ❌ {error}", flush=True)
            RUN_ERRORS.append(error)
            if critical:
                from alert import send_alert, AlertLevel
                send_alert(AlertLevel.ERROR, f"{script} failed", details=err_msg)
    return False


def metabase_query(payload):
    import urllib.request
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        f"{METABASE_API}/dataset", data=data,
        headers={"x-api-key": METABASE_KEY, "Content-Type": "application/json"},
    )
    resp = urllib.request.urlopen(req, timeout=60)
    return json.loads(resp.read())


# ── URL Database (fallback when Anakin is removed) ──

_url_db_cache = None
_url_db_lock = threading.Lock()


def load_url_database() -> dict:
    """Load url_database.json. Returns dict keyed by '{platform}_{pincode}_{item_code}'."""
    global _url_db_cache
    if _url_db_cache is not None:
        return _url_db_cache
    with _url_db_lock:
        if _url_db_cache is not None:
            return _url_db_cache
        if URL_DATABASE_PATH.exists():
            try:
                _url_db_cache = json.load(open(URL_DATABASE_PATH))
                print(f"  📂 URL database: {len(_url_db_cache)} entries", flush=True)
            except Exception as e:
                print(f"  ⚠️ URL database load error: {e}", flush=True)
                _url_db_cache = {}
        else:
            _url_db_cache = {}
    return _url_db_cache


def save_urls_to_database(pincode):
    """Save new URLs from PDP results to url_database.json. Only adds, never removes."""
    url_db = {}
    if URL_DATABASE_PATH.exists():
        try:
            url_db = json.load(open(URL_DATABASE_PATH))
        except Exception:
            url_db = {}

    added = 0
    for platform in ["blinkit", "jiomart"]:
        # Find latest PDP file for this pincode (any date)
        files = sorted([f for f in DATA.glob(f"sam/{platform}_pdp_{pincode}_*.json")
                        if "partial" not in f.name])
        if not files:
            continue
        try:
            data = json.load(open(files[-1]))
        except Exception:
            continue

        for p in data.get("products", []):
            ic = p.get("item_code")
            if not ic:
                continue
            url_key = f"{platform}_{pincode}_{ic}"
            product_url = p.get(f"{platform}_product_url")
            product_id = p.get(f"{platform}_product_id")
            if not product_url:
                continue

            # Only add if key is new or URL has changed
            existing = url_db.get(url_key)
            if existing and existing.get("product_url") == product_url:
                continue

            url_db[url_key] = {
                "item_code": ic,
                "platform": platform,
                "pincode": pincode,
                "product_id": product_id,
                "product_url": product_url,
                "platform_item_name": p.get("sam_product_name") or "",
                "apna_name": "",
                "brand": "",
                "updated_at": NOW,
            }
            added += 1

    if added > 0:
        URL_DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(URL_DATABASE_PATH, "w") as f:
            json.dump(url_db, f, indent=2)
        # Invalidate cache so next load picks up new data
        global _url_db_cache
        _url_db_cache = None
        print(f"  💾 URL database: +{added} URLs for {pincode} (total {len(url_db)})", flush=True)


# ── Data Validation ──

def validate_data(rows, pincodes):
    """Validate data before BQ push. Returns (ok: bool, messages: list[str])."""
    errors = []

    # Total rows > 0
    if len(rows) == 0:
        errors.append("FAIL: 0 total rows")
        return False, errors

    # Each pincode has at least 100 rows (pincode is index 3 in row)
    pincode_counts = {}
    for row in rows:
        pin = str(row[3])
        pincode_counts[pin] = pincode_counts.get(pin, 0) + 1
    for pin in pincodes:
        count = pincode_counts.get(pin, 0)
        if count < 100:
            errors.append(f"FAIL: pincode {pin} has only {count} rows (need >= 100)")

    # No row has item_code = None (item_code is index 4)
    none_ic = sum(1 for row in rows if row[4] is None)
    if none_ic > 0:
        errors.append(f"FAIL: {none_ic} rows have item_code = None")

    # blinkit_sp (index 18) and jio_sp (index 25) in range 0-50000 when not None
    for idx, label in [(18, "blinkit_sp"), (25, "jio_sp")]:
        bad = 0
        for row in rows:
            val = row[idx]
            if val is not None:
                try:
                    v = float(val)
                    if v < 0 or v > 50000:
                        bad += 1
                except (ValueError, TypeError):
                    bad += 1
        if bad > 0:
            errors.append(f"FAIL: {bad} rows have {label} outside 0-50000 range")

    # At least 10% of rows have blinkit_sp (not all blank)
    blinkit_filled = sum(1 for row in rows if row[18] is not None)
    pct = blinkit_filled / len(rows) * 100 if rows else 0
    if pct < 10:
        errors.append(f"FAIL: only {pct:.1f}% of rows have blinkit_sp ({blinkit_filled}/{len(rows)})")

    ok = len(errors) == 0
    return ok, errors


# ── Old File Cleanup ──

def cleanup_old_files():
    """Delete files older than 7 days from data/sam/, data/comparisons/, data/anakin/.
    Keeps data/mappings/ and data/ean_map.json. Only deletes if >20 files in directory."""
    from datetime import timedelta

    cutoff = datetime.now() - timedelta(days=7)
    dirs_to_clean = [DATA / "sam", DATA / "comparisons", DATA / "anakin"]
    total_deleted = 0

    for dir_path in dirs_to_clean:
        if not dir_path.exists():
            continue
        files = [f for f in dir_path.iterdir() if f.is_file()]
        if len(files) <= 20:
            continue
        deleted = 0
        for f in files:
            try:
                mtime = datetime.fromtimestamp(f.stat().st_mtime)
                if mtime < cutoff:
                    f.unlink()
                    deleted += 1
            except Exception:
                pass
        if deleted > 0:
            print(f"  🧹 {dir_path.name}/: deleted {deleted} old files", flush=True)
        total_deleted += deleted

    if total_deleted > 0:
        print(f"  🧹 Total cleanup: {total_deleted} files removed", flush=True)
    else:
        print(f"  🧹 Cleanup: nothing to remove", flush=True)


# ── Step 1: Fetch AM product master + MRP ──

def fetch_am_master(item_codes: list[str]) -> dict:
    """Fetch product data from smpcm_product for given item_codes."""
    print("📥 Fetching AM product master...", flush=True)
    am_map = {}
    batch_size = 100
    for i in range(0, len(item_codes), batch_size):
        batch = [int(ic) for ic in item_codes[i:i+batch_size] if ic.isdigit()]
        if not batch:
            continue
        try:
            r = metabase_query({
                "database": 5, "type": "query",
                "query": {
                    "source-table": 578,
                    "fields": [["field", 7191], ["field", 7118], ["field", 8935], ["field", 7113],
                               ["field", 7133], ["field", 7131], ["field", 7176], ["field", 7193],
                               ["field", 7158], ["field", 7149], ["field", 7126], ["field", 7137],
                               ["field", 15241]],
                    "filter": ["=", ["field", 7191]] + batch,
                    "limit": 200,
                }
            })
            cols = ["item_code", "display_name", "master_category", "brand", "marketed_by",
                    "product_type", "unit", "unit_value", "mrp", "main_image",
                    "sub_variant", "variant", "pack_size"]
            for row in r.get("data", {}).get("rows", []):
                rec = dict(zip(cols, row))
                ic = str(rec.get("item_code", "")).strip()
                if ic:
                    am_map[ic] = rec
        except Exception as e:
            print(f"    AM batch error: {e}", flush=True)

    print(f"  ✅ AM master: {len(am_map)} products", flush=True)
    with open(DATA / "am_product_master.json", "w") as f:
        json.dump(am_map, f, default=str)
    return am_map


def fetch_latest_mrp(warehouse_id: str) -> dict:
    """Fetch latest inward MRP from model 1808 for a warehouse."""
    print(f"📥 Fetching latest MRP ({warehouse_id})...", flush=True)
    mrp_rows = []
    page = 1
    while True:
        try:
            r = metabase_query({
                "database": 3, "type": "query",
                "query": {
                    "source-table": "card__1808",
                    "filter": ["=", ["field", "warehouse_id", {"base-type": "type/Text"}], warehouse_id],
                    "limit": 2000, "offset": (page - 1) * 2000,
                }
            })
            rows = r.get("data", {}).get("rows", [])
            if not rows:
                break
            mrp_rows.extend(rows)
            if len(rows) < 2000:
                break
            page += 1
            if page > 15:
                break
        except Exception as e:
            print(f"    MRP fetch error: {e}", flush=True)
            break

    mrp_cols = ["warehouse_id", "grn_date", "pricing_approv_date", "product_id",
                "item_code", "cost", "mrp", "display_name", "master_category"]
    mrp_map = {}
    for row in mrp_rows:
        rec = dict(zip(mrp_cols, row))
        ic = str(rec.get("item_code", "")).strip()
        if ic:
            mrp_map[ic] = rec

    print(f"  ✅ MRP: {len(mrp_map)} items for {warehouse_id}", flush=True)
    safe_name = warehouse_id.lower().replace(" ", "_")
    with open(DATA / f"latest_mrp_{safe_name}.json", "w") as f:
        json.dump(mrp_map, f, default=str)
    return mrp_map


# ── Step 2: Scrape one city ──

def fetch_all_anakin(pincodes):
    """Fetch Anakin data for all cities (lightweight API calls, no browser)."""
    print(f"\n📥 Fetching Anakin data for {len(pincodes)} cities...", flush=True)
    for pin, city in pincodes.items():
        run("fetch_anakin_blinkit.py", [pin])
        run("fetch_anakin_jiomart.py", [pin])
        print(f"  ✅ Anakin: {city} ({pin})", flush=True)


def scrape_blinkit_city(pincode, city):
    """Run Blinkit pipeline for one city (fast, ~30 min)."""
    city_platforms = CITY_PLATFORMS.get(pincode, {"blinkit"})
    if "blinkit" not in city_platforms:
        print(f"  ⏭️  blinkit not available in {city}", flush=True)
        return
    try:
        print(f"\n⚙️  {city} — blinkit pipeline", flush=True)
        run("scrape_blinkit_pdps.py", [pincode, "8"], use_venv=True, retries=1, critical=True)
        partial = DATA / "sam" / f"blinkit_pdp_{pincode}_latest_partial.json"
        if partial.exists():
            partial.unlink()
        run("compare_pdp.py", [pincode])
        run("unified_matcher.py", [pincode, "blinkit"], timeout=600)
        run("stage4_image_match.py", [pincode, "blinkit"])
        run("stage5_barcode_match.py", [pincode, "blinkit"], timeout=300, retries=0)
        print(f"  ✅ {city} blinkit complete", flush=True)
    except Exception as e:
        print(f"  ❌ {city} blinkit crashed: {str(e)[:200]}", flush=True)
        RUN_ERRORS.append(f"{city} blinkit: {str(e)[:150]}")
    save_urls_to_database(pincode)


def scrape_jiomart_city(pincode, city):
    """Run Jiomart pipeline for one city (slow, search + fetch)."""
    city_platforms = CITY_PLATFORMS.get(pincode, {"blinkit"})
    if "jiomart" not in city_platforms:
        print(f"  ⏭️  jiomart not available in {city}", flush=True)
        return
    try:
        print(f"\n⚙️  {city} — jiomart pipeline (fetch prices)", flush=True)
        city_state = _cities_config["cities"].get(pincode, {}).get("state", "")
        hd_csv = str(DATA / "hd_assortment.csv")
        jm_args = ["--pincode", pincode, "--tabs", "4", "--csv", hd_csv]
        if city_state:
            jm_args += ["--state", city_state]
        run("jiomart_fetch_prices.py", jm_args, use_venv=True, retries=0, timeout=18000)
        print(f"  ✅ {city} jiomart complete", flush=True)
    except Exception as e:
        print(f"  ❌ {city} jiomart crashed: {str(e)[:200]}", flush=True)
        RUN_ERRORS.append(f"{city} jiomart: {str(e)[:150]}")


# Flipkart Minutes pincode mapping (FK needs specific pincode for geolocation)
FK_PINCODE_MAP = {
    "834002": "834008",  # Ranchi → 834008 (Gandhi Nagar, Kanke)
}


def scrape_flipkart_minutes_city(pincode, city):
    """Run Flipkart Minutes pipeline (Chromium, cookies, ~5 min)."""
    city_platforms = CITY_PLATFORMS.get(pincode, {"blinkit"})
    if "flipkart_minutes" not in city_platforms:
        return
    try:
        fk_pin = FK_PINCODE_MAP.get(pincode, pincode)
        print(f"\n⚙️  {city} — flipkart minutes pipeline (pin:{fk_pin})", flush=True)
        run("scrape_flipkart_minutes.py", [fk_pin, "--match"], use_venv=False, retries=0, timeout=600)
        # Copy match file to original pincode name (for generate_city_data)
        if fk_pin != pincode:
            import glob
            fk_files = sorted(glob.glob(str(DATA / f"flipkart_minutes_match_{fk_pin}_*.json")))
            if fk_files:
                import shutil
                dest = str(fk_files[-1]).replace(f"_{fk_pin}_", f"_{pincode}_")
                shutil.copy2(fk_files[-1], dest)
        print(f"  ✅ {city} flipkart minutes complete", flush=True)
    except Exception as e:
        print(f"  ❌ {city} flipkart minutes: {str(e)[:100]}", flush=True)
        RUN_ERRORS.append(f"{city} flipkart_minutes: {str(e)[:100]}")


def scrape_dealshare_city(pincode, city):
    """Run DealShare pipeline for one city (fast, API only, ~30 sec)."""
    city_platforms = CITY_PLATFORMS.get(pincode, {"blinkit"})
    if "dealshare" not in city_platforms:
        print(f"  ⏭️  dealshare not available in {city}", flush=True)
        return
    try:
        print(f"\n⚙️  {city} — dealshare pipeline (API)", flush=True)
        run("scrape_dealshare.py", [pincode, "--match"], retries=1, timeout=300)
        print(f"  ✅ {city} dealshare complete", flush=True)
    except Exception as e:
        print(f"  ❌ {city} dealshare crashed: {str(e)[:200]}", flush=True)
        RUN_ERRORS.append(f"{city} dealshare: {str(e)[:150]}")


# ── Step 3: Compute status + generate output ──
# Status computation now uses UnifiedMatchingEngine (unified_matcher.py)

from unified_matcher import UnifiedMatchingEngine

_engine = UnifiedMatchingEngine()  # No pool needed — used only for compute_status + heuristics


def compute_status(am, am_mrp, sam_sp, sam_mrp, sam_name, anakin_rec=None, platform=None):
    """Wrapper around UnifiedMatchingEngine.compute_status for backward compat."""
    return _engine.compute_status(am, am_mrp, sam_sp, sam_mrp, sam_name)


def load_pdp(platform, pincode):
    # Try today's file first, fallback to latest available
    files = sorted([f for f in DATA.glob(f"sam/{platform}_pdp_{pincode}_{DATE}*.json") if "partial" not in f.name])
    if not files:
        files = sorted([f for f in DATA.glob(f"sam/{platform}_pdp_{pincode}_*.json") if "partial" not in f.name and "category" not in f.name])
    if not files:
        return {}, set()
    d = json.load(open(files[-1]))
    ok = {p["item_code"]: p for p in d["products"]
          if p.get("item_code") and p.get("status") == "ok"
          and not (p.get("sam_product_name") or "").startswith("projects/")}
    # Track products confirmed not_available on platform (OOS / redirected)
    not_available = {p["item_code"] for p in d["products"]
                     if p.get("item_code") and p.get("status") == "not_available"}
    return ok, not_available


def load_cascade(platform, pincode):
    cm = {}
    # Unified matcher output first (highest priority), then legacy fallbacks
    patterns = [
        f"{platform}_unified_{pincode}_{DATE}*.json",
        f"{platform}_cascade_{pincode}_{DATE}*.json",
        f"{platform}_stage3_{pincode}_{DATE}*.json",
    ]
    if platform == "jiomart":
        patterns.append(f"jiomart_search_match_{pincode}_{DATE}*.json")
    for pat in patterns:
        files = sorted(DATA.glob(f"comparisons/{pat}"))
        # Fallback to latest if today's not found
        if not files:
            fallback_pat = pat.replace(f"_{DATE}", "_")
            files = sorted(DATA.glob(f"comparisons/{fallback_pat}"))
        if files:
            for m in json.load(open(files[-1])).get("new_mappings", []):
                ic = m.get("item_code")
                if ic and m.get("sam_price") and ic not in cm:
                    cm[ic] = m
    return cm


def generate_city_data(pincode, city, am_map, mrp_map):
    """Generate rows for one city. Returns list of CSV rows."""
    blinkit_anakin = {}
    jiomart_anakin = {}
    for f in sorted(DATA.glob(f"anakin/blinkit_{pincode}_*.json")):
        blinkit_anakin = {r["Item_Code"]: r for r in json.load(open(f))["records"] if r.get("Item_Code")}
    for f in sorted(DATA.glob(f"anakin/jiomart_{pincode}_*.json")):
        jiomart_anakin = {r["Item_Code"]: r for r in json.load(open(f))["records"] if r.get("Item_Code")}

    # Load URL database as fallback for missing Anakin URLs
    url_db = load_url_database()

    b_pdp, b_not_avail = load_pdp("blinkit", pincode)
    j_pdp, j_not_avail = load_pdp("jiomart", pincode)
    b_cascade = load_cascade("blinkit", pincode)
    j_cascade = load_cascade("jiomart", pincode)

    # Load DMart data (API-based, no PDP/cascade stages)
    dmart_map = {}
    dmart_files = sorted([f for f in DATA.glob(f"sam/dmart_{pincode}_{DATE}*.json")])
    if dmart_files:
        d = json.load(open(dmart_files[-1]))
        for p in d.get("products", []):
            dmart_map[p.get("product_name", "")] = p

    # Load DealShare match data
    ds_map = {}
    ds_match_files = sorted([f for f in DATA.glob(f"dealshare_match_{pincode}_*.json")])
    if ds_match_files:
        ds_data = json.load(open(ds_match_files[-1]))
        ds_products = json.load(open(DATA / "dealshare_product_master.json")) if (DATA / "dealshare_product_master.json").exists() else {}
        for m in ds_data.get("matches", []):
            ic = m.get("item_code")
            if ic and m.get("ds_sp"):
                # Find URL from product master
                ds_url = ""
                for pid, p in ds_products.items():
                    if p.get("title") == m.get("ds_name"):
                        ds_url = p.get("url", "")
                        break
                ds_map[ic] = {
                    "name": m.get("ds_name"),
                    "sp": m.get("ds_sp"),
                    "mrp": m.get("ds_mrp"),
                    "url": ds_url,
                    "status": m.get("match_status"),
                    "unit": "",
                }

    # Load Flipkart Minutes match data
    fk_map = {}
    fk_match_files = sorted([f for f in DATA.glob(f"flipkart_minutes_match_{pincode}_*.json")])
    if fk_match_files:
        fk_data = json.load(open(fk_match_files[-1]))
        for m in fk_data.get("matches", []):
            ic = m.get("item_code")
            if ic and m.get("fk_sp"):
                fk_map[ic] = {
                    "name": m.get("fk_name"),
                    "sp": m.get("fk_sp"),
                    "mrp": m.get("fk_mrp"),
                    "url": m.get("fk_url", ""),
                    "status": m.get("match_status"),
                    "unit": m.get("fk_unit"),
                }

    rows = []
    for ic in sorted(am_map.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        am = am_map.get(ic, {})
        if am.get("master_category") not in VALID_MASTER_CATEGORIES:
            continue
        mrp_rec = mrp_map.get(ic)
        am_mrp = mrp_rec.get("mrp") if mrp_rec else am.get("mrp")
        b_ana = blinkit_anakin.get(ic, {})
        j_ana = jiomart_anakin.get(ic, {})

        def get_sam(pdp_m, cas_m, ana_r, url_k, platform, not_avail_set=set()):
            sp = mrp = name = stock = unit = None
            url = ana_r.get(url_k)
            if url and str(url).strip().lower() in ("na", "nan", "null", "none", ""):
                url = None
            # Fallback to URL database if Anakin URL is missing
            if not url:
                db_key = f"{platform}_{pincode}_{ic}"
                db_entry = url_db.get(db_key)
                if db_entry:
                    url = db_entry.get("product_url")
            # If PDP confirmed not_available, don't use cascade data
            if ic in not_avail_set:
                return url, None, None, None, None, "out_of_stock"
            if ic in pdp_m:
                p = pdp_m[ic]
                name = p.get("sam_product_name")
                sp = p.get("sam_selling_price")
                mrp = p.get("sam_mrp")
                stock = "available" if p.get("sam_in_stock") else "out_of_stock"
                unit = p.get("sam_unit")
                # Pick URL from PDP if available (new jiomart_fetch_prices URLs)
                pdp_url = p.get("jiomart_product_url") or p.get("blinkit_product_url")
                if pdp_url:
                    url = pdp_url
            elif ic in cas_m:
                m = cas_m[ic]
                name = m.get("sam_product_name")
                sp = m.get("sam_price")
                mrp = m.get("sam_mrp")
                stock = "available" if m.get("sam_in_stock", True) else "out_of_stock"
                unit = m.get("sam_unit")
            # Price sanity: reject bulk/combo prices (SP > 3× AM MRP = clearly wrong product)
            if sp is not None and am_mrp:
                try:
                    if float(sp) > float(am_mrp) * 3:
                        sp = mrp = name = stock = unit = None
                except (ValueError, TypeError):
                    pass
            # Variant check: if weight ratio outside 0.7-1.5, mark as NA (not same product)
            if sp is not None and name:
                from utils import UNIT_ALIASES, to_base_unit
                sam_wt, sam_wu = _engine._parse_weight(name)
                am_u = (am.get("unit") or "").lower().strip()
                am_uv_val = am.get("unit_value")
                if sam_wt and am_uv_val and am_u and sam_wu:
                    try:
                        u_norm = UNIT_ALIASES.get(am_u, am_u)
                        am_bv, am_bu = to_base_unit(float(am_uv_val), u_norm)
                        if am_bu == sam_wu and am_bv > 0 and sam_wt > 0:
                            ratio = sam_wt / am_bv
                            if ratio < 0.7 or ratio > 1.5:
                                sp = mrp = name = stock = unit = None
                    except (ValueError, TypeError):
                        pass
            return url, name, unit, mrp, sp, stock

        b_url, b_name, b_unit, b_mrp, b_sp, b_stock = get_sam(b_pdp, b_cascade, b_ana, "Blinkit_Product_Url", "blinkit", b_not_avail)
        b_status = compute_status(am, am_mrp, b_sp, b_mrp, b_name, b_ana, "blinkit")

        j_url, j_name, j_unit, j_mrp, j_sp, j_stock = get_sam(j_pdp, j_cascade, j_ana, "Jiomart_Product_Url", "jiomart", j_not_avail)
        j_status = compute_status(am, am_mrp, j_sp, j_mrp, j_name, j_ana, "jiomart")

        # DMart: match by product name similarity (no Anakin mapping exists)
        d_url = d_name = d_unit = d_mrp = d_sp = d_stock = d_status = None
        am_display = (am.get("display_name") or "").lower()
        am_brand = (am.get("brand") or "").lower().strip()
        if dmart_map and am_display:
            best_match = None
            best_score = 0
            try:
                from rapidfuzz import fuzz
            except ImportError:
                # Fallback for local dev without rapidfuzz
                from difflib import SequenceMatcher
                class fuzz:
                    @staticmethod
                    def token_sort_ratio(a, b):
                        return SequenceMatcher(None, " ".join(sorted(a.split())), " ".join(sorted(b.split()))).ratio() * 100
            # Strip common prefixes that hurt matching
            am_clean = re.sub(r'\s*[-|]\s*\d+\s*(g|gm|gms|kg|ml|l|ltr|pcs?|n)\b', '', am_display).strip()
            for dname, dp in dmart_map.items():
                dname_low = dname.lower()
                # Strip DMart's SKU suffix (": 500 gms", ": 1 kg")
                dname_clean = re.sub(r'\s*:\s*\d+\s*(g|gm|gms|kg|ml|l|ltr|pcs?)\b.*', '', dname_low).strip()
                # Strip "dmart premia/swaad" brand prefix for better matching
                dname_clean = re.sub(r'^dmart\s+(premia|swaad|kitchen)\s+', '', dname_clean).strip()
                # Brand check: if AM brand exists, DMart product should contain it (or vice versa)
                d_brand = (dp.get("brand") or "").lower().strip()
                if am_brand and d_brand and am_brand not in d_brand and d_brand not in am_brand:
                    # Different brands — skip unless names are very similar
                    score = fuzz.token_sort_ratio(am_clean, dname_clean)
                    if score < 80:
                        continue
                score = fuzz.token_sort_ratio(am_clean, dname_clean)
                if score > best_score:
                    best_score = score
                    best_match = dp
            if best_match and best_score >= 50:
                d_url = best_match.get("product_url")
                d_name = best_match.get("product_name")
                d_unit = best_match.get("unit")
                d_mrp = best_match.get("mrp")
                d_sp = best_match.get("price")
                d_stock = "available" if best_match.get("in_stock") else "out_of_stock"
                d_status = compute_status(am, am_mrp, d_sp, d_mrp, d_name, {}, "dmart")

        # DealShare data
        ds = ds_map.get(ic, {})
        ds_url = ds.get("url")
        ds_name = ds.get("name")
        ds_unit = ds.get("unit")
        ds_mrp = ds.get("mrp")
        ds_sp = ds.get("sp")
        ds_stock = "available" if ds_sp else None
        ds_status = ds.get("status") if ds_sp else None

        # Flipkart Minutes data
        fk = fk_map.get(ic, {})
        fk_url = fk.get("url")
        fk_name = fk.get("name")
        fk_unit = fk.get("unit")
        fk_mrp = fk.get("mrp")
        fk_sp = fk.get("sp")
        fk_stock = "available" if fk_sp else None
        fk_status = fk.get("status") if fk_sp else None

        rows.append([
            DATE, NOW, city, pincode, int(ic) if ic.isdigit() else ic,
            am.get("display_name"), am.get("master_category"), am.get("brand"), am.get("marketed_by"),
            am.get("product_type"), am.get("unit"), am.get("unit_value"), am_mrp, am.get("main_image"),
            b_url, b_name, b_unit, b_mrp, b_sp, b_stock, b_status,
            j_url, j_name, j_unit, j_mrp, j_sp, j_stock, j_status,
            d_url, d_name, d_unit, d_mrp, d_sp, d_stock, d_status,
            am.get("sub_variant"), am.get("variant"), am.get("pack_size"),
            ds_url, ds_name, ds_unit, ds_mrp, ds_sp, ds_stock, ds_status,
            fk_url, fk_name, fk_unit, fk_mrp, fk_sp, fk_stock, fk_status,
        ])
    return rows


def generate_excel(rows, city, pincode):
    """Generate Excel file for one city."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️ openpyxl not installed — skipping Excel", flush=True)
        return

    am_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
    blinkit_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    jio_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = f"SAM_{city}_{DATE}"

    headers = [
        "DATE", "TIME", "CITY", "PINCODE",
        "AM ITEM CODE", "AM ITEM NAME", "AM master cat", "AM BRAND", "AM MARKETED BY",
        "AM PRODUCT TYPE", "AM UNIT", "AM UNIT VALUE", "AM MRP", "IMAGE LINK",
        "BLINKIT URL", "BLINKIT ITEM NAME", "BLINKIT UNIT", "BLINKIT MRP", "BLINKIT SP",
        "BLINKIT IN STOCK REMARK", "BLINKIT STATUS",
        "JIO URL", "JIO ITEM NAME", "JIO UNIT", "JIO MRP", "JIO SP",
        "JIO IN STOCK REMARK", "JIO STATUS",
        "DMART URL", "DMART ITEM NAME", "DMART UNIT", "DMART MRP", "DMART SP",
        "DMART IN STOCK REMARK", "DMART STATUS",
    ]
    dmart_fill = PatternFill(start_color="FFE6CCFF", end_color="FFE6CCFF", fill_type="solid")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, size=10)
        if 1 <= i <= 14: c.fill = am_fill
        elif 15 <= i <= 21: c.fill = blinkit_fill
        elif 22 <= i <= 28: c.fill = jio_fill
        elif 29 <= i <= 35: c.fill = dmart_fill

    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    widths = [10, 18, 10, 8, 10, 40, 8, 15, 20, 20, 5, 8, 8, 30, 35, 40, 10, 8, 8, 12, 20, 35, 40, 10, 8, 8, 12, 20, 35, 40, 10, 8, 8, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"

    out_path = OUTPUT_DIR / f"SAM_{city}_{pincode}_{DATE}.xlsx"
    wb.save(out_path)
    print(f"  📊 {out_path.name}", flush=True)


def process_city(pin, city, am_map, mrp_maps, city_index, total_cities):
    """
    2-step push:
      Step 1: Blinkit scrape → generate → push to BQ (fast, ~30 min)
      Step 2: Jiomart fetch → update BQ with Jiomart data (slow, runs after)
    """
    print(f"\n{'═' * 60}", flush=True)
    print(f"  CITY {city_index}/{total_cities}: {city} ({pin})", flush=True)
    print(f"{'═' * 60}", flush=True)

    wh = WAREHOUSE_MAP.get(pin, "WRHS_1")
    mrp_map = mrp_maps.get(wh, {})

    def _generate_and_push(label):
        """Generate data from all scraped so far + push to BQ."""
        city_rows = generate_city_data(pin, city, am_map, mrp_map)
        city_csv_path = DATA / f"bq_upload_{pin}.csv"
        with open(city_csv_path, "w", newline="") as f:
            w = csv.writer(f)
            for row in city_rows:
                w.writerow(row)
        push_to_bigquery(city_csv_path, [pin])
        b = sum(1 for r in city_rows if r[18])
        j = sum(1 for r in city_rows if r[25])
        print(f"  ✅ {city} {label} → BQ ({len(city_rows)} rows, blinkit:{b} jio:{j})", flush=True)
        return city_rows

    # ── Scrape + push incrementally ──
    scrape_blinkit_city(pin, city)
    _generate_and_push("Blinkit")

    scrape_dealshare_city(pin, city)
    scrape_flipkart_minutes_city(pin, city)
    _generate_and_push("DealShare+Flipkart")

    scrape_jiomart_city(pin, city)
    city_rows = _generate_and_push("Jiomart (final)")

    backup_to_gcs(DATA / f"bq_upload_{pin}.csv", pin)

    return city_rows


def push_to_bigquery(csv_path, pincodes):
    """Push CSV to sam_price_live (dedup: delete today's rows first) + sam_price_history (append)."""
    print("\n📤 Pushing to BigQuery...", flush=True)
    from google.cloud import bigquery
    client = bigquery.Client(project=BQ_PROJECT)

    live_table = f"{BQ_PROJECT}.{BQ_DATASET}.sam_price_live"
    history_table = f"{BQ_PROJECT}.{BQ_DATASET}.sam_price_history"

    # Dedup: delete existing rows for today's pincodes before appending
    try:
        pin_list = ", ".join(f"'{p}'" for p in pincodes)
        delete_sql = f"DELETE FROM `{live_table}` WHERE date = '{DATE}' AND pincode IN ({pin_list})"
        client.query(delete_sql).result()
        print(f"  🗑️  Deleted existing rows for {DATE} [{pin_list}]", flush=True)
    except Exception as e:
        print(f"  ⚠️  Dedup delete failed (continuing): {str(e)[:200]}", flush=True)

    # Push to sam_price_live
    try:
        job_config = bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.CSV,
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            allow_quoted_newlines=True,
            max_bad_records=50,
        )
        with open(csv_path, "rb") as f:
            client.load_table_from_file(f, live_table, job_config=job_config).result()
        print(f"  ✅ sam_price_live (appended)", flush=True)
    except Exception as e:
        print(f"  ❌ sam_price_live: {str(e)[:200]}", flush=True)

    # Push to sam_price_history (append only — permanent record)
    try:
        job_config = bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.CSV,
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            allow_quoted_newlines=True,
            max_bad_records=50,
        )
        with open(csv_path, "rb") as f:
            client.load_table_from_file(f, history_table, job_config=job_config).result()
        print(f"  ✅ sam_price_history (appended)", flush=True)
    except Exception as e:
        print(f"  ❌ sam_price_history: {str(e)[:200]}", flush=True)


def backup_to_gcs(csv_path, pincode):
    """Backup daily CSV to GCS bucket: gs://sam-price-data/{date}/{pincode}.csv"""
    try:
        from google.cloud import storage
        client = storage.Client(project=BQ_PROJECT)
        bucket = client.bucket(GCS_BUCKET)
        blob_name = f"{DATE}/{pincode}.csv"
        blob = bucket.blob(blob_name)
        blob.upload_from_filename(str(csv_path))
        print(f"  ☁️  GCS backup: gs://{GCS_BUCKET}/{blob_name}", flush=True)
    except Exception as e:
        print(f"  ⚠️  GCS backup failed: {str(e)[:200]}", flush=True)


# ── Main ──

def main():
    pincode_arg = sys.argv[1] if len(sys.argv) > 1 and not sys.argv[1].startswith("--") else "all"
    skip_scrape = "--no-scrape" in sys.argv

    # Support state filter: python sam_daily_run.py --state JH
    state_filter = None
    for i, a in enumerate(sys.argv):
        if a == "--state" and i + 1 < len(sys.argv):
            state_filter = sys.argv[i + 1].upper()

    if state_filter:
        pincodes = {pin: cfg["name"] for pin, cfg in _cities_config["cities"].items()
                    if cfg.get("state", "").upper() == state_filter}
    elif pincode_arg == "all":
        pincodes = CITIES
    else:
        pincodes = {pincode_arg: CITIES.get(pincode_arg, pincode_arg)}

    print(f"{'═' * 60}")
    print(f"  SAM DAILY RUN — {DATE} {NOW}")
    print(f"  Cities: {', '.join(pincodes.values())}")
    print(f"  Scrape: {'skip' if skip_scrape else 'yes'}")
    print(f"{'═' * 60}")

    # Step 0: Switch gcloud account (skip in Docker where gcloud may not exist)
    try:
        subprocess.run(["gcloud", "config", "set", "account", "satyam.gupta@apnamart.in"],
                        capture_output=True, timeout=5)
    except FileNotFoundError:
        print("  (gcloud not found — running in Docker, using service account)", flush=True)

    # Step 1: Fetch EAN map
    print("\n📥 Fetching EAN map...", flush=True)
    run("fetch_ean_map.py")

    # Step 2: Collect item_codes from url_database + KVI (Anakin removed — url_database is primary)
    all_item_codes = set()

    # Source 1: URL database (primary — 17,446+ saved URLs)
    url_db = load_url_database()
    for key, entry in url_db.items():
        ic = str(entry.get("item_code", "")).strip()
        if ic:
            all_item_codes.add(ic)
    print(f"  📦 URL database: {len(all_item_codes)} unique item_codes", flush=True)

    # Source 2: KVI master (high-priority items)
    kvi_path = DATA / "kvi_master.json"
    if kvi_path.exists():
        kvi_data = json.load(open(kvi_path))
        kvi_count = 0
        for item in kvi_data.get("kvi", []):
            ic = str(item.get("item_code", "")).strip()
            if ic and ic not in all_item_codes:
                all_item_codes.add(ic)
                kvi_count += 1
        if kvi_count:
            print(f"  📋 KVI: +{kvi_count} items added", flush=True)

    # Source 3: Anakin files if they exist (optional supplement — will be removed)
    for pin in pincodes:
        for platform in ["blinkit", "jiomart"]:
            files = sorted(DATA.glob(f"anakin/{platform}_{pin}_*.json"))
            if files:
                d = json.load(open(files[-1]))
                anakin_added = 0
                for r in d.get("records", []):
                    ic = str(r.get("Item_Code", "")).strip()
                    if ic and ic not in all_item_codes:
                        all_item_codes.add(ic)
                        anakin_added += 1
                if anakin_added:
                    print(f"  📎 Anakin {platform} {pin}: +{anakin_added} items", flush=True)

    print(f"  Total item_codes: {len(all_item_codes)}", flush=True)

    # Step 3: Fetch AM master + MRP
    am_map = fetch_am_master(list(all_item_codes))

    # Fetch MRP per warehouse (deduplicate warehouses)
    mrp_maps = {}
    for pin in pincodes:
        wh = WAREHOUSE_MAP.get(pin, "WRHS_1")
        if wh not in mrp_maps:
            mrp_maps[wh] = fetch_latest_mrp(wh)

    # Step 5: Per-city sequential processing (scrape → generate → validate → push)
    all_rows = []
    cities_summary = {}

    if skip_scrape:
        # --no-scrape mode: just generate + push, no scraping
        for pin, city in pincodes.items():
            wh = WAREHOUSE_MAP.get(pin, "WRHS_1")
            mrp_map = mrp_maps.get(wh, {})
            city_rows = generate_city_data(pin, city, am_map, mrp_map)
            all_rows.extend(city_rows)
            print(f"  ✅ {city}: {len(city_rows)} rows", flush=True)
        # Single push for all cities
        csv_path = DATA / "bq_upload_temp.csv"
        with open(csv_path, "w", newline="") as f:
            w = csv.writer(f)
            for row in all_rows:
                w.writerow(row)
        valid, messages = validate_data(all_rows, list(pincodes.keys()))
        if not valid:
            for msg in messages:
                print(f"  ⚠️ {msg}", flush=True)
        push_to_bigquery(csv_path, list(pincodes.keys()))
    else:
        # Normal mode: per-city scrape → generate → push to BQ
        for i, (pin, city) in enumerate(pincodes.items(), 1):
            city_rows = process_city(pin, city, am_map, mrp_maps, i, len(pincodes))
            all_rows.extend(city_rows)
            b_ok = sum(1 for r in city_rows if r[18])
            j_ok = sum(1 for r in city_rows if r[25])
            d_ok = sum(1 for r in city_rows if len(r) > 32 and r[32])
            cities_summary[pin] = {
                "city": city, "rows": len(city_rows),
                "blinkit_ok": b_ok, "jiomart_ok": j_ok, "dmart_ok": d_ok
            }

    # Step 6: KVI coverage report
    kvi_path = DATA / "kvi_master.json"
    if kvi_path.exists():
        kvi_data = json.load(open(kvi_path))
        state_to_pins = kvi_data.get("state_map", {})
        pin_to_state = {}
        for st, pins in state_to_pins.items():
            for p in pins:
                pin_to_state[p] = st
        kvi_by_state = {}
        for item in kvi_data.get("kvi", []):
            st = {"1": "JH", "2": "WB", "3": "CG"}.get(item.get("state_key"), item.get("state_key"))
            if st not in kvi_by_state:
                kvi_by_state[st] = {}
            kvi_by_state[st][item["item_code"]] = item.get("kvi_tag", "KVI")

        print(f"\n📊 KVI Coverage Report", flush=True)
        for pin, city in pincodes.items():
            state = pin_to_state.get(pin)
            if not state or state not in kvi_by_state:
                continue
            kvi_items = kvi_by_state[state]
            super_kvi = {ic for ic, tag in kvi_items.items() if "Super" in tag}
            city_rows = [r for r in all_rows if str(r[3]) == pin]
            city_ics = {str(r[4]): r for r in city_rows}

            b_hit = sum(1 for ic in kvi_items if ic in city_ics and city_ics[ic][18])
            j_hit = sum(1 for ic in kvi_items if ic in city_ics and city_ics[ic][25])
            either = sum(1 for ic in kvi_items if ic in city_ics and (city_ics[ic][18] or city_ics[ic][25]))
            skvi_b = sum(1 for ic in super_kvi if ic in city_ics and city_ics[ic][18])
            skvi_j = sum(1 for ic in super_kvi if ic in city_ics and city_ics[ic][25])

            n = len(kvi_items)
            sn = len(super_kvi)
            print(f"  {city} ({state}): KVI {either}/{n} = {either/n*100:.0f}%  "
                  f"[B:{b_hit}/{n}={b_hit/n*100:.0f}% J:{j_hit}/{n}={j_hit/n*100:.0f}%]  "
                  f"Super KVI: B:{skvi_b}/{sn} J:{skvi_j}/{sn}", flush=True)

    # Step 7: Cleanup old files
    cleanup_old_files()

    # Step 8: Send summary alert
    end_time = time.time()
    duration = end_time - _run_start_time if '_run_start_time' in globals() else 0
    try:
        from alert import send_daily_summary, send_alert, AlertLevel
        if not cities_summary:
            for pin, city in pincodes.items():
                cr = [r for r in all_rows if str(r[3]) == pin]
                b_ok = sum(1 for r in cr if r[18])
                j_ok = sum(1 for r in cr if r[25])
                d_ok = sum(1 for r in cr if len(r) > 32 and r[32])
                cities_summary[pin] = {"city": city, "rows": len(cr), "blinkit_ok": b_ok, "jiomart_ok": j_ok, "dmart_ok": d_ok}
        send_daily_summary(cities_summary, len(all_rows), duration, errors=RUN_ERRORS or None)
    except Exception as e:
        print(f"  ⚠️ Alert send failed: {e}", flush=True)

    print(f"\n{'═' * 60}")
    print(f"  DONE! {len(pincodes)} cities, {len(all_rows)} rows")
    if RUN_ERRORS:
        print(f"  ⚠️ {len(RUN_ERRORS)} errors during run:")
        for e in RUN_ERRORS:
            print(f"    - {e}")
    print(f"  BigQuery: sam_price_live + sam_price_history")
    print(f"{'═' * 60}")


if __name__ == "__main__":
    _run_start_time = time.time()
    main()

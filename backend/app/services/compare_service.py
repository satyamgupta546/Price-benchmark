import asyncio
import io
import json
import re
from datetime import datetime
from difflib import SequenceMatcher
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings
from app.scrapers.blinkit_scraper import BlinkitScraper
from app.scrapers.zepto_scraper import ZeptoScraper
from app.scrapers.instamart_scraper import InstamartScraper
from app.scrapers.jiomart_scraper import JioMartScraper
from app.scrapers.flipkart_minutes_scraper import FlipkartMinutesScraper

# ── Platform config ──────────────────────────────────────────────

PLATFORM_CONFIG = {
    "blinkit": {
        "class": BlinkitScraper,
        "name": "Blinkit",
        "color": "FFF8C723",
        "text_color": "000000",
        "search_url": lambda base, term: f"{base}/s/?q={quote(term)}",
    },
    "zepto": {
        "class": ZeptoScraper,
        "name": "Zepto",
        "color": "FF8B22CF",
        "text_color": "FFFFFF",
        "search_url": lambda base, term: f"{base}/search?query={quote(term)}",
    },
    "instamart": {
        "class": InstamartScraper,
        "name": "Instamart",
        "color": "FFFC8019",
        "text_color": "FFFFFF",
        "search_url": lambda base, term: f"https://www.swiggy.com/instamart/search?custom_back=true&query={quote(term)}",
    },
    "jiomart": {
        "class": JioMartScraper,
        "name": "JioMart",
        "color": "FF0078AD",
        "text_color": "FFFFFF",
        "search_url": lambda base, term: f"{base}/search/{quote(term)}",
    },
    "flipkart_minutes": {
        "class": FlipkartMinutesScraper,
        "name": "Flipkart Min",
        "color": "FF2874F0",
        "text_color": "FFFFFF",
        "search_url": lambda base, term: f"{base}/search?q={quote(term)}&marketplace=GROCERY",
    },
}

ALL_PLATFORM_IDS = list(PLATFORM_CONFIG.keys())

# ── Excel styles ─────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
GRAY_FILL = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
GRAY_FONT = Font(color="808080", italic=True)


def _style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


# ── Reference Excel parser ───────────────────────────────────────

def parse_reference_excel(file_bytes: bytes) -> list[dict]:
    """Parse the reference Excel file and extract product rows from 'anaken' sheet."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Try to find the 'anaken' sheet (case-insensitive)
    sheet = None
    for name in wb.sheetnames:
        if name.lower().strip() == "anaken":
            sheet = wb[name]
            break

    if sheet is None:
        sheet = wb.active

    products = []
    rows = list(sheet.iter_rows(min_row=1, values_only=True))
    if not rows:
        wb.close()
        return products

    # Find header row — look for 'Item_Name' or similar in first 5 rows
    header_row_idx = 0
    headers = []
    for i, row in enumerate(rows[:5]):
        row_strs = [str(c).strip().lower() if c else "" for c in row]
        if any("item_name" in s or "item name" in s for s in row_strs):
            header_row_idx = i
            headers = [str(c).strip() if c else "" for c in row]
            break

    if not headers:
        headers = [str(c).strip() if c else "" for c in rows[0]]

    # Build column index map (case-insensitive)
    col_map = {}
    for idx, h in enumerate(headers):
        col_map[h.lower().replace(" ", "_")] = idx

    def find_col(*candidates):
        for c in candidates:
            c_lower = c.lower().replace(" ", "_")
            if c_lower in col_map:
                return col_map[c_lower]
            for key, idx in col_map.items():
                if c_lower in key or key in c_lower:
                    return idx
        return None

    item_name_col = find_col("item_name", "item name", "product_name", "name")
    brand_col = find_col("brand", "brand_name")
    jio_name_col = find_col("jiomart_item_name", "jiomart item name", "jio_item_name")
    jio_mrp_col = find_col("jiomart_mrp_price", "jiomart mrp price", "jio_mrp")
    jio_sp_col = find_col("jiomart_selling_price", "jiomart selling price", "jio_selling", "jiomart_sp")

    if item_name_col is None:
        wb.close()
        raise ValueError(f"Could not find 'Item_Name' column. Headers found: {headers}")

    for row in rows[header_row_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        def safe_get(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        name = safe_get(item_name_col)
        if not name or not str(name).strip():
            continue

        brand = safe_get(brand_col)
        jio_name = safe_get(jio_name_col)
        jio_mrp = safe_get(jio_mrp_col)
        jio_sp = safe_get(jio_sp_col)

        def parse_price(val):
            if val is None:
                return None
            try:
                return float(str(val).replace(",", "").replace("₹", "").strip())
            except (ValueError, TypeError):
                return None

        products.append({
            "name": str(name).strip(),
            "brand": str(brand).strip() if brand else "",
            "jio_name": str(jio_name).strip() if jio_name else "",
            "jio_mrp": parse_price(jio_mrp),
            "jio_sp": parse_price(jio_sp),
        })

    wb.close()
    return products


def _name_similarity(a: str, b: str) -> float:
    """Compare two product names using SequenceMatcher. Returns 0-1 score."""
    if not a or not b:
        return 0.0
    a_clean = re.sub(r'[^\w\s]', '', a.lower()).strip()
    b_clean = re.sub(r'[^\w\s]', '', b.lower()).strip()
    return SequenceMatcher(None, a_clean, b_clean).ratio()


# ── Platform browser init ────────────────────────────────────────

async def _init_platform_browser(platform_id: str, pincode: str):
    """Initialize a platform scraper with browser + location. Returns scraper or None."""
    config = PLATFORM_CONFIG.get(platform_id)
    if not config:
        return None

    scraper = config["class"](pincode=pincode, max_products=10000)

    try:
        await scraper.init_browser()

        if platform_id == "blinkit":
            await _init_blinkit(scraper)
        elif platform_id == "zepto":
            await _init_zepto(scraper)
        elif platform_id == "instamart":
            await _init_instamart(scraper)
        elif platform_id == "jiomart":
            await _init_jiomart(scraper)
        elif platform_id == "flipkart_minutes":
            await _init_flipkart(scraper)

        return scraper
    except Exception as e:
        print(f"[compare] Failed to init {platform_id}: {e}")
        try:
            await scraper.close()
        except Exception:
            pass
        return None


async def _init_blinkit(scraper):
    """Blinkit: Chromium + localStorage location + cookies."""
    await scraper.page.goto(scraper.base_url, wait_until="domcontentloaded", timeout=20000)
    await asyncio.sleep(1)

    location_data = json.dumps({
        "coords": {
            "isDefault": False, "lat": scraper.lat, "lon": scraper.lng,
            "locality": "Selected Location", "cityName": "Selected",
            "id": None, "isTopCity": False, "landmark": None, "addressId": None,
        }
    })
    await scraper.page.evaluate(f"() => localStorage.setItem('location', {json.dumps(location_data)})")

    await scraper.context.add_cookies([
        {"name": "__pincode", "value": scraper.pincode, "domain": ".blinkit.com", "path": "/"},
        {"name": "gr_1_lat", "value": str(scraper.lat), "domain": ".blinkit.com", "path": "/"},
        {"name": "gr_1_lon", "value": str(scraper.lng), "domain": ".blinkit.com", "path": "/"},
    ])

    await scraper.page.goto(scraper.base_url, wait_until="domcontentloaded", timeout=20000)
    await asyncio.sleep(2)


async def _init_zepto(scraper):
    """Zepto: Chromium + cookies + localStorage + UI fallback."""
    await scraper.page.goto(scraper.base_url, wait_until="domcontentloaded", timeout=20000)
    await asyncio.sleep(1.5)
    await scraper._set_location()
    await scraper._set_local_storage()
    await scraper.page.goto(scraper.base_url, wait_until="domcontentloaded", timeout=20000)
    await asyncio.sleep(2)
    await scraper._try_ui_location()


async def _init_instamart(scraper):
    """Instamart: Chromium + Swiggy cookies + localStorage + WAF bypass."""
    location_data = json.dumps({
        "lat": scraper.lat, "lng": scraper.lng,
        "address": f"Pincode {scraper.pincode}", "pincode": scraper.pincode,
    })
    await scraper.context.add_cookies([
        {"name": "lat", "value": str(scraper.lat), "domain": ".swiggy.com", "path": "/"},
        {"name": "lng", "value": str(scraper.lng), "domain": ".swiggy.com", "path": "/"},
        {"name": "userLocation", "value": location_data, "domain": ".swiggy.com", "path": "/"},
        {"name": "addressId", "value": "", "domain": ".swiggy.com", "path": "/"},
    ])

    try:
        await scraper.page.goto("https://www.swiggy.com", wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(3)
    except Exception:
        pass

    await scraper._set_swiggy_location()

    try:
        await scraper.page.goto(scraper.base_url, wait_until="domcontentloaded", timeout=25000)
    except Exception:
        await asyncio.sleep(2)
        if "/instamart" not in scraper.page.url:
            await scraper.page.goto(scraper.base_url, wait_until="domcontentloaded", timeout=25000)
    await asyncio.sleep(3)


async def _init_jiomart(scraper):
    """JioMart: Firefox + pincode cookies."""
    await scraper.context.add_cookies([
        {"name": "pincode", "value": scraper.pincode, "domain": ".jiomart.com", "path": "/"},
        {"name": "address_pincode", "value": scraper.pincode, "domain": ".jiomart.com", "path": "/"},
    ])
    try:
        await scraper.page.goto(scraper.base_url, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(2)
    except Exception:
        pass

    # Try UI location picker
    try:
        sel_loc = scraper.page.locator('text="Select Location Manually"').first
        if await sel_loc.is_visible(timeout=3000):
            await sel_loc.click()
            await asyncio.sleep(2)
            inp = scraper.page.locator('input[placeholder*="incode"], input[type="text"]').first
            if await inp.is_visible(timeout=3000):
                await inp.fill(scraper.pincode)
                await asyncio.sleep(2)
                try:
                    sug = scraper.page.locator('li, [role="option"]').first
                    if await sug.is_visible(timeout=3000):
                        await sug.click()
                        await asyncio.sleep(2)
                except Exception:
                    await inp.press("Enter")
                    await asyncio.sleep(2)
    except Exception:
        pass


async def _init_flipkart(scraper):
    """Flipkart Minutes: Chromium + pincode form (press_sequentially)."""
    await scraper.page.goto(scraper.base_url, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(1)

    # Dismiss login popup
    try:
        close = scraper.page.locator('button:has-text("✕")').first
        if await close.is_visible(timeout=3000):
            await close.click()
            await asyncio.sleep(1)
    except Exception:
        pass

    await scraper._set_delivery_location()
    await scraper._set_grocery_pincode()


# ── Single product search ────────────────────────────────────────

async def _search_one_product(scraper, config, search_term, ref_name):
    """Search for a single product on a platform. Returns match dict or None."""
    # Clear state for fresh search
    scraper.products.clear()
    scraper._captured_responses.clear()
    scraper._processed_urls.clear()
    scraper._seen_ids.clear()

    try:
        search_url = config["search_url"](scraper.base_url, search_term)
        await scraper.page.goto(search_url, wait_until="domcontentloaded", timeout=15000)
        await scraper._wait_for_network_settle(0.5, 2.0)

        # Process API responses
        scraper._process_responses()

        # DOM extraction fallback
        await scraper._extract_products_from_dom()

        candidates = scraper.products
        if not candidates:
            return None

        best_match = None
        best_score = 0.0

        for p in candidates:
            score = _name_similarity(search_term, p.product_name)
            if ref_name != search_term:
                score = max(score, _name_similarity(ref_name, p.product_name))
            if score > best_score:
                best_score = score
                best_match = p

        if best_match and best_score >= 0.35:
            return {
                "name": best_match.product_name,
                "price": best_match.price,
                "mrp": best_match.mrp,
                "match_score": round(best_score, 2),
            }
    except Exception as e:
        print(f"[compare] Search error on {scraper.platform_name} for '{search_term[:40]}': {e}")

    return None


# ── Main compare function ────────────────────────────────────────

async def compare_with_reference(file_bytes: bytes, pincode: str, platforms: list[str],
                                  progress_callback=None) -> tuple[bytes, str]:
    """
    Compare reference Excel products against live prices on selected platforms.
    Platforms run in parallel (one browser each), products searched sequentially per platform.
    Returns (excel_bytes, filename).
    """
    ref_products = parse_reference_excel(file_bytes)
    if not ref_products:
        raise ValueError("No products found in the uploaded Excel file")

    total = len(ref_products)
    if progress_callback:
        await progress_callback("started", {"total": total, "platforms": platforms})

    # Init all platform browsers in parallel
    print(f"[compare] Initializing {len(platforms)} platform browsers for pincode {pincode}...")
    init_results = await asyncio.gather(
        *[_init_platform_browser(p, pincode) for p in platforms],
        return_exceptions=True,
    )

    scrapers = {}
    for platform_id, result in zip(platforms, init_results):
        if isinstance(result, Exception) or result is None:
            print(f"[compare] {platform_id} init failed: {result}")
            if progress_callback:
                await progress_callback("platform_error", {
                    "platform": platform_id,
                    "error": str(result) if isinstance(result, Exception) else "Init failed",
                })
        else:
            scrapers[platform_id] = result
            print(f"[compare] {platform_id} browser ready")

    if not scrapers:
        raise ValueError("No platforms could be initialized. Check browser installation.")

    # Per-platform search (platforms parallel, products sequential per platform)
    platform_results = {}

    async def search_platform(platform_id, scraper):
        config = PLATFORM_CONFIG[platform_id]
        results = []
        matched = 0
        not_found = 0

        for i, ref in enumerate(ref_products):
            # Use jio_name as alternate search term for JioMart only
            search_name = ref["name"]
            alt_name = None
            if platform_id == "jiomart" and ref.get("jio_name"):
                search_name = ref["jio_name"]
                alt_name = ref["name"]

            result = await _search_one_product(scraper, config, search_name, ref["name"])

            # Retry with original name if JioMart jio_name search failed
            if result is None and alt_name and alt_name != search_name:
                result = await _search_one_product(scraper, config, alt_name, alt_name)

            if result:
                matched += 1
            else:
                not_found += 1
            results.append(result)

            if progress_callback:
                await progress_callback("progress", {
                    "platform": platform_id,
                    "current": i + 1,
                    "total": total,
                    "product": ref["name"][:50],
                    "status": "matched" if result else "not_found",
                    "matched": matched,
                    "not_found": not_found,
                })

        print(f"[compare] {platform_id} done: {matched} matched, {not_found} not found")
        return results, matched, not_found

    # Run all platforms in parallel
    gather_results = await asyncio.gather(
        *[search_platform(p, s) for p, s in scrapers.items()],
        return_exceptions=True,
    )

    for platform_id, result in zip(scrapers.keys(), gather_results):
        if isinstance(result, Exception):
            print(f"[compare] {platform_id} search failed: {result}")
            platform_results[platform_id] = {
                "results": [None] * total, "matched": 0, "not_found": total,
            }
        else:
            results_list, matched, not_found = result
            platform_results[platform_id] = {
                "results": results_list, "matched": matched, "not_found": not_found,
            }

    # Cleanup browsers
    for scraper in scrapers.values():
        try:
            await scraper.close()
        except Exception:
            pass

    # Done event
    if progress_callback:
        platform_stats = {}
        for p in platforms:
            if p in platform_results:
                platform_stats[p] = {
                    "matched": platform_results[p]["matched"],
                    "not_found": platform_results[p]["not_found"],
                }
            else:
                platform_stats[p] = {"matched": 0, "not_found": 0, "error": True}
        await progress_callback("done", {
            "total": total,
            "platforms": platform_stats,
            "file_ready": True,
        })

    # Generate Excel
    active_platforms = [p for p in platforms if p in platform_results]
    excel_bytes = _generate_delta_excel(ref_products, platform_results, active_platforms, pincode)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"Delta_{pincode}_{date_str}.xlsx"

    return excel_bytes, filename


# ── Excel generation ─────────────────────────────────────────────

def _generate_delta_excel(ref_products, platform_results, active_platforms, pincode):
    """Generate styled Excel with per-platform price columns and delta analysis."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Delta"
    ws.sheet_properties.tabColor = "0078AD"

    # Build column layout
    base_headers = ["Sr No", "Item Name", "Brand", "Ref MRP", "Ref SP"]
    col_offset = len(base_headers) + 1  # next 1-indexed column
    platform_price_col = {}
    platform_match_col = {}

    platform_headers = []
    for p in active_platforms:
        cfg = PLATFORM_CONFIG[p]
        platform_price_col[p] = col_offset
        platform_match_col[p] = col_offset + 1
        platform_headers.extend([f"{cfg['name']} Price", f"{cfg['name']} Match%"])
        col_offset += 2

    tail_headers = ["Best Price", "Best Platform", "Delta from Ref"]
    best_price_col = col_offset
    best_platform_col = col_offset + 1
    delta_col = col_offset + 2

    all_headers = base_headers + platform_headers + tail_headers

    # Write headers
    for i, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        _style_header(cell)

    # Color platform headers with platform brand colors
    for p in active_platforms:
        cfg = PLATFORM_CONFIG[p]
        fill = PatternFill(start_color=cfg["color"], end_color=cfg["color"], fill_type="solid")
        font = Font(bold=True, color=cfg["text_color"], size=11)
        for c in [platform_price_col[p], platform_match_col[p]]:
            ws.cell(row=1, column=c).fill = fill
            ws.cell(row=1, column=c).font = font
            ws.cell(row=1, column=c).alignment = HEADER_ALIGN

    # Write data rows
    for idx, ref in enumerate(ref_products):
        row_num = idx + 2
        ref_price = ref.get("jio_sp") or ref.get("jio_mrp")

        # Base columns
        ws.cell(row=row_num, column=1, value=idx + 1).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=ref["name"]).border = THIN_BORDER
        ws.cell(row=row_num, column=3, value=ref["brand"]).border = THIN_BORDER

        for col, val in [(4, ref.get("jio_mrp")), (5, ref.get("jio_sp"))]:
            cell = ws.cell(row=row_num, column=col, value=val if val is not None else "")
            cell.border = THIN_BORDER
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

        # Platform columns
        best_price = None
        best_plat_name = None

        for p in active_platforms:
            p_results = platform_results.get(p, {}).get("results", [])
            p_result = p_results[idx] if idx < len(p_results) else None

            price_cell = ws.cell(row=row_num, column=platform_price_col[p])
            match_cell = ws.cell(row=row_num, column=platform_match_col[p])
            price_cell.border = THIN_BORDER
            match_cell.border = THIN_BORDER

            if p_result:
                price_cell.value = p_result["price"]
                price_cell.number_format = '#,##0.00'
                match_cell.value = p_result["match_score"]
                match_cell.number_format = '0%'

                if best_price is None or p_result["price"] < best_price:
                    best_price = p_result["price"]
                    best_plat_name = PLATFORM_CONFIG[p]["name"]

                # Color: green if cheaper than ref, red if costlier
                if ref_price and ref_price > 0:
                    if p_result["price"] < ref_price - 0.5:
                        price_cell.fill = GREEN_FILL
                        price_cell.font = GREEN_FONT
                    elif p_result["price"] > ref_price + 0.5:
                        price_cell.fill = RED_FILL
                        price_cell.font = RED_FONT
            else:
                price_cell.value = ""
                match_cell.value = ""
                price_cell.fill = GRAY_FILL
                price_cell.font = GRAY_FONT
                match_cell.fill = GRAY_FILL
                match_cell.font = GRAY_FONT

        # Best price / platform / delta
        bp_cell = ws.cell(row=row_num, column=best_price_col)
        bpl_cell = ws.cell(row=row_num, column=best_platform_col)
        d_cell = ws.cell(row=row_num, column=delta_col)
        bp_cell.border = THIN_BORDER
        bpl_cell.border = THIN_BORDER
        d_cell.border = THIN_BORDER

        if best_price is not None:
            bp_cell.value = best_price
            bp_cell.number_format = '#,##0.00'
            bpl_cell.value = best_plat_name

            if ref_price and ref_price > 0:
                delta = best_price - ref_price
                d_cell.value = delta
                d_cell.number_format = '+#,##0.00;-#,##0.00;0.00'
                if delta < -0.5:
                    d_cell.fill = GREEN_FILL
                    d_cell.font = GREEN_FONT
                elif delta > 0.5:
                    d_cell.fill = RED_FILL
                    d_cell.font = RED_FONT
            else:
                d_cell.value = ""
        else:
            bp_cell.value = ""
            bpl_cell.value = ""
            d_cell.value = ""
            for c in [bp_cell, bpl_cell, d_cell]:
                c.fill = GRAY_FILL
                c.font = GRAY_FONT

    # Column widths
    widths = [6, 40, 15, 12, 12]
    for _ in active_platforms:
        widths.extend([12, 8])
    widths.extend([12, 15, 12])
    for i, w in enumerate(widths, 1):
        if i <= len(all_headers):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

    # Summary sheet (insert at front)
    _write_summary_sheet(wb, ref_products, platform_results, active_platforms, pincode)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _write_summary_sheet(wb, ref_products, platform_results, active_platforms, pincode):
    """Write a summary sheet with per-platform match rates and price analysis."""
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_properties.tabColor = "4CAF50"
    total = len(ref_products)

    summary_data = [
        ("Metric", "Value"),
        ("Total Products", total),
        ("Pincode", pincode),
        ("Platforms Compared", len(active_platforms)),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Platform Match Rates", ""),
    ]

    for p in active_platforms:
        cfg = PLATFORM_CONFIG[p]
        pr = platform_results.get(p, {"matched": 0, "not_found": total})
        rate = f"{pr['matched'] / total * 100:.1f}%" if total else "0%"
        summary_data.append((cfg["name"], f"{pr['matched']} matched ({rate})"))

    # Cross-platform coverage
    any_match = sum(
        1 for i in range(total)
        if any(
            platform_results.get(p, {"results": []}).get("results", [None] * total)[i] is not None
            for p in active_platforms
        )
    )
    summary_data.extend([
        ("", ""),
        ("Cross-Platform Coverage", ""),
        ("Found on Any Platform", f"{any_match} ({any_match / total * 100:.1f}%)" if total else "0"),
    ])

    # Price analysis
    ref_prices = [ref.get("jio_sp") or ref.get("jio_mrp") for ref in ref_products]
    cheaper = costlier = same = 0
    deltas = []

    for i in range(total):
        prices = []
        for p in active_platforms:
            r_list = platform_results.get(p, {"results": []}).get("results", [])
            if i < len(r_list) and r_list[i]:
                prices.append(r_list[i]["price"])
        if prices and ref_prices[i] and ref_prices[i] > 0:
            best = min(prices)
            d = best - ref_prices[i]
            deltas.append(d)
            if d < -0.5:
                cheaper += 1
            elif d > 0.5:
                costlier += 1
            else:
                same += 1

    if deltas:
        avg_delta = sum(deltas) / len(deltas)
        summary_data.extend([
            ("", ""),
            ("Price Analysis (Best Live vs Ref)", ""),
            ("Avg Delta", f"{avg_delta:+.2f}"),
            ("Cheaper on Live", cheaper),
            ("Costlier on Live", costlier),
            ("Same Price (within 0.5)", same),
        ])

    # Write cells
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        cell_b = ws.cell(row=row_idx, column=2, value=value)
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER
        if row_idx == 1:
            _style_header(cell_a)
            _style_header(cell_b)
        else:
            cell_a.font = Font(bold=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30

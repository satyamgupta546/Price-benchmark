import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.product import Product

PLATFORM_NAMES = {
    "blinkit": "Blinkit",
    "zepto": "Zepto",
    "instamart": "Swiggy Instamart",
    "jiomart": "JioMart",
    "flipkart_minutes": "Flipkart Minutes",
}

PLATFORM_COLORS = {
    "blinkit": "FFF8C723",
    "zepto": "FF8B22CF",
    "instamart": "FFFC8019",
    "jiomart": "FF0078AD",
    "flipkart_minutes": "FF2874F0",
}

HEADERS = ["Sr No", "Product Name", "Brand", "Category", "Pincode", "Price", "MRP"]

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
GREEN_FONT = Font(bold=True, color="006100")
RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")


def _style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


def _write_sheet(ws, products: list[Product], sheet_color: str = None):
    """Write products to a worksheet. Each product is a separate row — no dedup."""
    if sheet_color:
        ws.sheet_properties.tabColor = sheet_color

    for col_idx, header in enumerate(HEADERS, 1):
        _style_header(ws.cell(row=1, column=col_idx, value=header))

    sorted_products = sorted(products, key=lambda p: (p.pincode, p.product_name.lower()))

    for idx, p in enumerate(sorted_products, 1):
        row = idx + 1
        values = [
            idx,
            p.product_name,
            p.brand or "",
            p.category or "",
            p.pincode,
            p.price,
            p.mrp if p.mrp else "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx in (6, 7) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col_idx == 5:
                cell.alignment = Alignment(horizontal="center")

    col_widths = [8, 50, 20, 20, 10, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _write_comparison_sheet(ws, products: list[Product], active_platforms: list[str]):
    """Price Comparison: same product across platforms in one row.
    Columns: Sr No, Product Name, Brand, Category, Pincode,
             Blinkit Price, Blinkit MRP, JioMart Price, JioMart MRP, ...,
             Cheapest Platform, Cheapest Price, Max Price, Price Diff
    """
    ws.sheet_properties.tabColor = "FF9C27B0"

    # Build headers
    headers = ["Sr No", "Product Name", "Brand", "Category", "Pincode"]
    for plat in active_platforms:
        name = PLATFORM_NAMES.get(plat, plat)
        headers.append(f"{name} Price")
        headers.append(f"{name} MRP")
    headers.extend(["Cheapest Platform", "Cheapest Price", "Highest Price", "Price Diff"])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _style_header(cell)
        # Color platform headers
        for plat in active_platforms:
            name = PLATFORM_NAMES.get(plat, plat)
            if header.startswith(name):
                color = PLATFORM_COLORS.get(plat, "FF666666")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                # Use black text for light backgrounds (Blinkit yellow)
                if plat == "blinkit":
                    cell.font = Font(bold=True, color="000000", size=11)
                break

    # Group products by (normalized_name, pincode) to compare across platforms
    comparison_map: dict[tuple, dict] = {}
    for p in products:
        key = ((p.product_name or "").lower().strip(), p.pincode)
        if key[0] == "":
            continue
        if key not in comparison_map:
            comparison_map[key] = {
                "product_name": p.product_name,
                "brand": p.brand or "",
                "category": p.category or "",
                "pincode": p.pincode,
                "platforms": {},
            }
        # Keep each platform's price — don't overwrite if already set with a valid price
        if p.platform not in comparison_map[key]["platforms"] or p.price > 0:
            comparison_map[key]["platforms"][p.platform] = {
                "price": p.price,
                "mrp": p.mrp,
            }

    rows = sorted(comparison_map.values(), key=lambda x: (x["pincode"], x["product_name"].lower()))

    # Write data
    for idx, item in enumerate(rows, 1):
        row_num = idx + 1
        col = 1

        # Basic info
        for val in [idx, item["product_name"], item["brand"], item["category"], item["pincode"]]:
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 5:
                cell.alignment = Alignment(horizontal="center")
            col += 1

        # Platform prices
        valid_prices = []
        price_cells = {}  # track cells for highlighting

        for plat in active_platforms:
            info = item["platforms"].get(plat)
            price_col = col
            mrp_col = col + 1

            if info and info["price"] > 0:
                cell_price = ws.cell(row=row_num, column=price_col, value=info["price"])
                cell_price.number_format = '#,##0.00'
                valid_prices.append((plat, info["price"]))
                price_cells[plat] = cell_price
            else:
                cell_price = ws.cell(row=row_num, column=price_col, value="")

            if info and info.get("mrp") and info["mrp"] > 0:
                cell_mrp = ws.cell(row=row_num, column=mrp_col, value=info["mrp"])
                cell_mrp.number_format = '#,##0.00'
            else:
                cell_mrp = ws.cell(row=row_num, column=mrp_col, value="")

            cell_price.border = THIN_BORDER
            cell_mrp.border = THIN_BORDER
            col += 2

        # Cheapest/Highest/Diff
        if valid_prices:
            valid_prices.sort(key=lambda x: x[1])
            cheapest_plat, cheapest_price = valid_prices[0]
            highest_plat, highest_price = valid_prices[-1]
            diff = highest_price - cheapest_price

            ws.cell(row=row_num, column=col, value=PLATFORM_NAMES.get(cheapest_plat, cheapest_plat)).border = THIN_BORDER

            cell_cp = ws.cell(row=row_num, column=col + 1, value=cheapest_price)
            cell_cp.number_format = '#,##0.00'
            cell_cp.border = THIN_BORDER

            cell_hp = ws.cell(row=row_num, column=col + 2, value=highest_price)
            cell_hp.number_format = '#,##0.00'
            cell_hp.border = THIN_BORDER

            cell_diff = ws.cell(row=row_num, column=col + 3, value=diff)
            cell_diff.number_format = '#,##0.00'
            cell_diff.border = THIN_BORDER

            # Highlight cheapest in green, highest in red
            if len(valid_prices) > 1:
                if cheapest_plat in price_cells:
                    price_cells[cheapest_plat].fill = GREEN_FILL
                    price_cells[cheapest_plat].font = GREEN_FONT
                if highest_plat in price_cells and highest_plat != cheapest_plat:
                    price_cells[highest_plat].fill = RED_FILL
                    price_cells[highest_plat].font = RED_FONT
                if diff > 0:
                    cell_diff.fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
        else:
            for i in range(4):
                ws.cell(row=row_num, column=col + i, value="").border = THIN_BORDER

    # Column widths
    widths = [8, 50, 18, 18, 10]
    for _ in active_platforms:
        widths.extend([14, 14])
    widths.extend([18, 14, 14, 14])
    for i, w in enumerate(widths, 1):
        if i <= len(headers):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B2"


def generate_excel(products: list[Product], pincodes: str) -> tuple[bytes, str]:
    """Generate Excel with sheets: Price Comparison + All + per-platform."""
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"PriceBenchmark_{pincodes}_{date_str}.xlsx"

    wb = Workbook()
    active_platforms = sorted(set(p.platform for p in products))

    # Sheet 1: Price Comparison (most important — cross-platform view)
    ws_compare = wb.active
    ws_compare.title = "Price Comparison"
    _write_comparison_sheet(ws_compare, products, active_platforms)

    # Sheet 2: All products (flat list)
    ws_all = wb.create_sheet(title="All")
    all_headers = ["Sr No", "Product Name", "Brand", "Category", "Platform", "Pincode", "Price", "MRP"]

    ws_all.sheet_properties.tabColor = "4CAF50"
    for col_idx, header in enumerate(all_headers, 1):
        _style_header(ws_all.cell(row=1, column=col_idx, value=header))

    sorted_all = sorted(products, key=lambda p: (p.pincode, p.platform, p.product_name.lower()))
    for idx, p in enumerate(sorted_all, 1):
        row = idx + 1
        values = [
            idx,
            p.product_name,
            p.brand or "",
            p.category or "",
            PLATFORM_NAMES.get(p.platform, p.platform),
            p.pincode,
            p.price,
            p.mrp if p.mrp else "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_all.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx in (7, 8) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if col_idx == 6:
                cell.alignment = Alignment(horizontal="center")

    all_widths = [8, 50, 20, 20, 18, 10, 12, 12]
    for i, width in enumerate(all_widths, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = width
    ws_all.freeze_panes = "A2"

    # Per-platform sheets
    for platform in active_platforms:
        platform_products = [p for p in products if p.platform == platform]
        if not platform_products:
            continue

        sheet_name = PLATFORM_NAMES.get(platform, platform)
        ws = wb.create_sheet(title=sheet_name[:31])
        color = PLATFORM_COLORS.get(platform, "FF666666").replace("FF", "", 1)
        _write_sheet(ws, platform_products, sheet_color=color)

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue(), filename

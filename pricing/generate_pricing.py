"""
Apnamart Pricing File Generator
================================
Generates Excel pricing files matching the Google Sheets EXACTLY.
Three outputs: FMCGF (59 cols), FMCGNF (59 cols), Staples (57 cols, 2 header rows).

Usage:
    cd "/Users/satyam/Desktop/code/Price benchmark"
    ./backend/venv/bin/python3 pricing/generate_pricing.py
    ./backend/venv/bin/python3 pricing/generate_pricing.py --category FMCGF
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))

from pricing.engine import calculate_sp_batch, safe_float
from pricing.data_loader import load_all_data, create_sample_inputs, STATES

OUTPUT_DIR = Path('/Users/satyam/Desktop/price csv')

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
FILL_GREEN = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
FILL_ORANGE = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
FILL_RED = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
FILL_PURPLE = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
FLAG_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
PCT_FMT = '0.00%'
NUM_FMT = '#,##0'

STATE_NAMES = {'JH': 'Jharkhand', 'CG': 'Chhattisgarh', 'WB': 'West Bengal'}


def _hdr(ws, row, col, fill=None):
    c = ws.cell(row=row, column=col)
    c.font = HEADER_FONT
    c.fill = fill or HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN_BORDER


def _autowidth(ws, mn=8, mx=30):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        w = mn
        for c in col_cells[:50]:
            if c.value:
                w = max(w, min(len(str(c.value)), mx))
        ws.column_dimensions[letter].width = w + 2


def _mrp_bucket(mrp):
    if mrp <= 40: return '<=40'
    if mrp <= 100: return '40-100'
    if mrp <= 200: return '100-200'
    return '>=200'


def _margin_bucket(m):
    if m <= 0.10: return '<=10%'
    if m <= 0.15: return '10%-15%'
    if m <= 0.20: return '15%-20%'
    if m <= 0.25: return '20%-25%'
    if m <= 0.30: return '25%-30%'
    if m <= 0.40: return '30%-40%'
    if m <= 0.54: return '40%-54%'
    return '>54%'


def _disc_bucket(d):
    if d <= 0: return '0%'
    if d <= 0.03: return 'b0%-3%'
    if d <= 0.07: return 'b3%-7%'
    if d <= 0.10: return 'b7%-10%'
    if d <= 0.20: return 'b10%-20%'
    return '>20%'


def _rm_bucket(r):
    if r <= 0: return '<0%'
    if r <= 0.05: return '0-5%'
    if r <= 0.10: return '5-10%'
    if r <= 0.15: return '10-15%'
    if r <= 0.20: return '15-20%'
    return '>20%'


def _filter_state(products, results, state):
    fp, fr = [], []
    for p, r in zip(products, results):
        if p.get('state', '').upper() == state:
            fp.append(p)
            fr.append(r)
    return fp, fr


# ══════════════════════════════════════════════════════════════════════════════
# FMCG MAIN SHEET — 59 columns (exact match with Google Sheet)
# ══════════════════════════════════════════════════════════════════════════════

# (name, fill)
FMCG_COLS = [
    # 1-10: Product info
    ('Key', HEADER_FILL), ('state', HEADER_FILL), ('item_code', HEADER_FILL),
    ('display_name', HEADER_FILL), ('KVI TAG', HEADER_FILL),
    ('category', HEADER_FILL), ('Sub_Category', HEADER_FILL), ('Leaf_Category', HEADER_FILL),
    ('brand', HEADER_FILL), ('marketed_by', HEADER_FILL),
    # 11-15: Sales
    ('LAST 30 DAY QTY SOLD', FILL_GREEN), ('LAST 30 DAY VALUE SOLD', FILL_GREEN),
    ('Last_90day_qty_Sold', FILL_GREEN), ('Last_90day_Sale', FILL_GREEN), ('Demand wt', FILL_GREEN),
    # 16-23: Promos
    ('EXCLUSIONS', FILL_ORANGE), ('OFF INVOICE PROMO %', FILL_ORANGE),
    ('OFF INVOICE PROMO VALUE', FILL_ORANGE), ('ON INVOICE PROMO %', FILL_ORANGE),
    ('ON INVOICE PROMO VALUE', FILL_ORANGE), ('FINAL INVOICE VALUE', FILL_ORANGE),
    ('ON INVOICE % OF MRP', FILL_ORANGE), ('ON INVOICE % / MARGIN %', FILL_ORANGE),
    # 24-30: Cost & Margin
    ('MRP', HEADER_FILL), ('MRP BUCKET', HEADER_FILL), ('LATEST INWARD COST', HEADER_FILL),
    ('MAP', HEADER_FILL), ('OFF INVOICE ADJUSTED COST', HEADER_FILL),
    ('GRN_MARGIN', HEADER_FILL), ('MAP MARGIN', HEADER_FILL),
    # 31-33: Benchmarking
    ('JIOMART MRP', FILL_PURPLE), ('JIOMART SP', FILL_PURPLE), ('equal mrp?', FILL_PURPLE),
    # 34-35: Margin
    ('MARGIN', HEADER_FILL), ('MARGIN BUCKET', HEADER_FILL),
    # 36-39: Old data (placeholders)
    ('OLD MRP', HEADER_FILL), ('OLD SP BAU', HEADER_FILL),
    ('OLD DISCOUNT %', HEADER_FILL), ('ASSORTMENT CHECK', HEADER_FILL),
    # 40-43: SP calc section
    ('MRP', FILL_RED), ('OFF INV ADJ COST', FILL_RED),
    ('Automated MAP', FILL_RED), ('SP BAU', FILL_RED),
    # 44: Changes
    ('CHANGES', FILL_RED),
    # 45-49: Remarks (5→1)
    ('remark 5', FILL_RED), ('Remark 4', FILL_RED), ('Remark 3', FILL_RED),
    ('REMARK 2', FILL_RED), ('REMARK 1', FILL_RED),
    # 50-54: Discount & RM
    ('DISCOUNT %', FILL_RED), ('BAU DISCOUNT', FILL_RED), ('DISCOUNT BUCKET', FILL_RED),
    ('RM %', FILL_RED), ('RM BUCKET', FILL_RED),
    # 55-56: Revenue estimates
    ('SP X QTY', FILL_GREEN), ('CP X QTY', FILL_GREEN),
    # 57-59: Validation
    ('SP = MRP', HEADER_FILL), ('MRP >= BAU', HEADER_FILL), ('BAU >= COST', HEADER_FILL),
]


def write_fmcg_main(wb, products, results):
    """Write FMCG Main sheet — 59 columns matching Google Sheet exactly."""
    ws = wb.active
    ws.title = 'Main'

    # Headers
    for i, (name, fill) in enumerate(FMCG_COLS, 1):
        ws.cell(row=1, column=i, value=name)
        _hdr(ws, 1, i, fill)
    ws.freeze_panes = 'A2'

    # Pre-compute demand weight denominator
    total_90d = sum(safe_float(p.get('sale_90d')) for p in products)

    for row, (prod, res) in enumerate(zip(products, results), 2):
        mrp = safe_float(prod.get('mrp'))
        map_price = safe_float(prod.get('map_price'))
        inward = safe_float(prod.get('latest_inward_cost'))
        margin = res.get('margin', 0)
        disc = res.get('discount_pct', 0)
        rm = res.get('retention_margin', 0)
        off_adj_cost = res.get('off_invoice_adjusted_cost', 0)
        on_inv_val = safe_float(prod.get('on_invoice_value'))
        on_inv_of_mrp = on_inv_val / mrp if mrp > 0 else 0
        on_inv_of_margin = on_inv_val / (mrp * margin) if mrp > 0 and margin > 0 else 0
        map_margin = (mrp - map_price) / mrp if mrp > 0 and map_price > 0 else 0
        jio_mrp = safe_float(prod.get('jio_mrp'))
        jio_sp = safe_float(prod.get('jio_sp'))
        sale_90d = safe_float(prod.get('sale_90d'))
        qty_90d = safe_float(prod.get('qty_90d'))
        demand_wt = sale_90d / total_90d if total_90d > 0 else 0
        sp = res.get('sp', 0)

        # Split remarks into 5 slots (REMARK 1 = primary, remark 5 = least important)
        remarks = res.get('remarks', [])
        remark1 = remarks[0] if len(remarks) > 0 else ''
        remark2 = remarks[1] if len(remarks) > 1 else ''
        remark3 = remarks[2] if len(remarks) > 2 else ''
        remark4 = remarks[3] if len(remarks) > 3 else ''
        remark5 = remarks[4] if len(remarks) > 4 else ''

        vals = [
            # 1-10
            prod.get('key', ''), prod.get('state', ''), prod.get('item_code', ''),
            prod.get('display_name', ''), prod.get('kvi_tag', ''),
            prod.get('category', ''), prod.get('sub_category', ''), prod.get('leaf_category', ''),
            prod.get('brand', ''), prod.get('marketed_by', ''),
            # 11-15
            safe_float(prod.get('qty_30d')), safe_float(prod.get('sale_30d')),
            qty_90d, sale_90d, demand_wt,
            # 16-23
            'yes' if prod.get('is_excluded') else '',
            safe_float(prod.get('off_invoice_pct')), safe_float(prod.get('off_invoice_value')),
            safe_float(prod.get('on_invoice_pct')), on_inv_val,
            res.get('final_invoice_value', 0), on_inv_of_mrp, on_inv_of_margin,
            # 24-30
            mrp, _mrp_bucket(mrp), inward, map_price, off_adj_cost,
            res.get('grn_margin', 0), map_margin,
            # 31-33
            jio_mrp if jio_mrp > 0 else '', jio_sp if jio_sp > 0 else '',
            'TRUE' if jio_mrp > 0 and abs(jio_mrp - mrp) < 1 else ('FALSE' if jio_mrp > 0 else ''),
            # 34-35
            margin, _margin_bucket(margin),
            # 36-39 (OLD data — blank placeholders)
            '', '', '', '',
            # 40-43 (SP calc)
            mrp, off_adj_cost, map_price, sp,
            # 44 (CHANGES — blank)
            '',
            # 45-49 (remarks 5→1)
            remark5, remark4, remark3, remark2, remark1,
            # 50-54
            disc, disc, _disc_bucket(disc), rm, _rm_bucket(rm),
            # 55-56
            sp * qty_90d if qty_90d > 0 else '', off_adj_cost * qty_90d if qty_90d > 0 else '',
            # 57-59
            res.get('sp_equals_mrp', False), res.get('mrp_gte_sp', True), res.get('sp_gte_cost', True),
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = THIN_BORDER
            name = FMCG_COLS[ci - 1][0]
            if any(k in name for k in ['MARGIN', 'RM %', 'DISCOUNT', 'Demand']):
                if isinstance(val, float) and val != 0:
                    cell.number_format = PCT_FMT
            elif any(k in name for k in ['MRP', 'COST', 'MAP', 'SP', 'SALE', 'VALUE', 'INVOICE', 'QTY']):
                if isinstance(val, (int, float)) and val != 0:
                    cell.number_format = NUM_FMT

    _autowidth(ws)
    print(f"  ✓ Main sheet: {len(products)} rows, {len(FMCG_COLS)} cols")


# ══════════════════════════════════════════════════════════════════════════════
# FMCG QUICK_DATA — 11 columns (exact match)
# ══════════════════════════════════════════════════════════════════════════════

FMCG_QD_COLS = [
    'KEY', 'STATE', 'ITEM  CODE', 'DISPLAY NAME', 'KVI / NKVI TAG',
    'CATEGORY', 'SUB CATEGORY', 'LEAF CATEGORY', 'MRP', 'OFF INVOICE ADJ COST', 'SP BAU',
]


def write_fmcg_quick_data(wb, products, results):
    ws = wb.create_sheet('Quick_data')
    for i, h in enumerate(FMCG_QD_COLS, 1):
        ws.cell(row=1, column=i, value=h)
        _hdr(ws, 1, i)
    ws.freeze_panes = 'A2'

    for row, (p, r) in enumerate(zip(products, results), 2):
        vals = [
            p.get('key'), p.get('state'), p.get('item_code'), p.get('display_name'),
            p.get('kvi_tag'), p.get('category'), p.get('sub_category'), p.get('leaf_category'),
            safe_float(p.get('mrp')), r.get('off_invoice_adjusted_cost', 0), r.get('sp', 0),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = THIN_BORDER
    _autowidth(ws)
    print(f"  ✓ Quick_data sheet: 11 cols")


# ══════════════════════════════════════════════════════════════════════════════
# STAPLES MAIN SHEET — 57 columns, 2 header rows (exact match)
# ══════════════════════════════════════════════════════════════════════════════

# Row 1: Category headers (merged cells)
STPLS_ROW1 = [
    ('', 1, 12),  # product info (no header in row 1 for these)
    ('', 13, 15),  # sales
    ('', 16, 20),  # cost
    ('', 21, 24),  # pricing
    ('', 25, 25),  # gap
    ('AM CITY LEVEL PRICES', 26, 31),
    ('ONLINE BENCHMARKING', 32, 34),
    ('', 35, 35),  # gap
    ('', 36, 37),  # changes
    ('OFFLINE BENCHMARKING', 38, 50),
    ('', 51, 57),  # output
]

# Row 2: Column headers
STPLS_COLS = [
    # 1-12: Product info
    ('KEY', HEADER_FILL), ('State', HEADER_FILL), ('Item code', HEADER_FILL),
    ('Item name', HEADER_FILL), ('KVI list', HEADER_FILL), ('Segment', HEADER_FILL),
    ('category', HEADER_FILL), ('sub_category', HEADER_FILL), ('leaf_category', HEADER_FILL),
    ('brand', HEADER_FILL), ('marketed by', HEADER_FILL), ('AM Brand Tagging', HEADER_FILL),
    # 13-15: Sales
    ('LAST 90 DAY QTY SOLD', FILL_GREEN), ('LAST 90 DAY SALES', FILL_GREEN),
    ('DEMAND WEIGHT', FILL_GREEN),
    # 16-20: Cost
    ('MRP', HEADER_FILL), ('LAST INWARD COST', HEADER_FILL), ('MAP', HEADER_FILL),
    ('MAX OF (INWARD AND MAP)', HEADER_FILL), ('GRN MARGIN', HEADER_FILL),
    # 21-24: Pricing
    ('CURRENT SP', HEADER_FILL), ('CURRENT SP', HEADER_FILL),
    ('FINAL SP', FILL_RED), ('RETENTION MARGIN', FILL_RED),
    # 25: gap
    ('', HEADER_FILL),
    # 26-31: City SPs
    ('ASANSOL SP', FILL_PURPLE), ('KOLKATA SP', FILL_PURPLE), ('BILASPUR SP', FILL_PURPLE),
    ('JAMSHEDPUR SP', FILL_PURPLE), ('HAZARIBAGH SP', FILL_PURPLE), ('RANCHI SP', FILL_PURPLE),
    # 32-34: Online benchmarking
    ('BLINKIT SP', FILL_ORANGE), ('JIO SP', FILL_ORANGE), ('MAX OF BOTH', FILL_ORANGE),
    # 35: gap
    ('', HEADER_FILL),
    # 36-37: Changes
    ('CHANGES', HEADER_FILL), ('CHANGES', HEADER_FILL),
    # 38-50: Offline benchmarking (blank placeholders)
    ('RANCHI', FILL_ORANGE), ('RAIPUR', FILL_ORANGE), ('BILASPUR', FILL_ORANGE),
    ('KORBA', FILL_ORANGE), ('KOLKATA (SUMO)', FILL_ORANGE), ('KOLKATA (RELIANCE)', FILL_ORANGE),
    ('RELIANCE', FILL_ORANGE), ('SMART POINT', FILL_ORANGE), ('GT', FILL_ORANGE),
    ('DHANUKA STORE', FILL_ORANGE), ('RELIANCE WB', FILL_ORANGE),
    ('ASANSOL GT', FILL_ORANGE), ('DURGAPUR GT', FILL_ORANGE),
    # 51-54: Output
    ('FINAL SKU LEVEL MARGIN', FILL_RED), ('FINAL SKU LEVEL MARKUP', FILL_RED),
    ('GUARDRAIL SKU PRICE', FILL_RED), ('WSH MARCH SKU', FILL_RED),
    # 55-57: Remarks (3 city groups)
    ('REMARKS (RANCHI, RAIPUR, KOLKATA)', FILL_RED),
    ('REMARKS (JMS , BILASP, ASANSOL)', FILL_RED),
    ('REMARKS (HZB , KORBA)', FILL_RED),
]


def write_staples_main(wb, products, results):
    """Write Staples Main sheet — 57 columns, 2 header rows."""
    ws = wb.active
    ws.title = 'MAIN SHEET'

    # Row 1: Category headers (merged cells)
    for text, start, end in STPLS_ROW1:
        if text:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            cell = ws.cell(row=1, column=start, value=text)
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = FILL_PURPLE
            cell.alignment = Alignment(horizontal='center')

    # Row 2: Column headers
    for i, (name, fill) in enumerate(STPLS_COLS, 1):
        ws.cell(row=2, column=i, value=name)
        _hdr(ws, 2, i, fill)
    ws.freeze_panes = 'A3'

    # Pre-compute demand weight
    total_90d = sum(safe_float(p.get('sale_90d')) for p in products)

    for row, (prod, res) in enumerate(zip(products, results), 3):
        mrp = safe_float(prod.get('mrp'))
        inward = safe_float(prod.get('latest_inward_cost'))
        map_val = safe_float(prod.get('map_price'))
        max_cost = max(inward, map_val) if inward > 0 or map_val > 0 else 0
        grn_margin = (mrp - inward) / mrp if mrp > 0 and inward > 0 else 0
        sp = res.get('sp', 0)
        cost = res.get('cost', 0)
        sku_margin = (sp - cost) / sp if sp > 0 and cost > 0 else 0
        sku_markup = (sp - cost) / cost if cost > 0 else 0
        blinkit_sp = safe_float(prod.get('blinkit_sp'))
        jio_sp = safe_float(prod.get('jio_sp'))
        max_bm = max(blinkit_sp, jio_sp)
        sale_90d = safe_float(prod.get('sale_90d'))
        qty_90d = safe_float(prod.get('qty_90d'))
        demand_wt = sale_90d / total_90d if total_90d > 0 else 0
        current_sp = safe_float(prod.get('current_sp'))
        city = prod.get('city_sps', {})

        # Guardrail check
        gl = safe_float(prod.get('guardrail_lower'))
        gu = safe_float(prod.get('guardrail_upper'))
        guard_price = ''
        if gl > 0 or gu > 0:
            if sku_margin < gl:
                guard_price = f'BELOW ({sku_margin:.1%} < {gl:.0%})'
            elif sku_margin > gu:
                guard_price = f'ABOVE ({sku_margin:.1%} > {gu:.0%})'
            else:
                guard_price = 'OK'

        # Remarks: split into 3 city groups
        remarks = res.get('remarks', [])
        rmk = '; '.join(remarks) if remarks else ''

        vals = [
            # 1-12
            prod.get('key', ''), prod.get('state', ''), prod.get('item_code', ''),
            prod.get('display_name', ''), prod.get('kvi_tag', ''),
            prod.get('segment', ''), prod.get('category', ''),
            prod.get('sub_category', ''), prod.get('leaf_category', ''),
            prod.get('brand', ''), prod.get('marketed_by', ''), '',
            # 13-15
            qty_90d, sale_90d, demand_wt,
            # 16-20
            mrp, inward, map_val, max_cost, grn_margin,
            # 21-24
            current_sp if current_sp > 0 else '', current_sp if current_sp > 0 else '',
            sp, res.get('retention_margin', 0),
            # 25: gap
            '',
            # 26-31: City SPs
            safe_float(city.get('ASANSOL')), safe_float(city.get('KOLKATA')),
            safe_float(city.get('BILASPUR')), safe_float(city.get('JAMSHEDPUR')),
            safe_float(city.get('HAZARIBAGH')), safe_float(city.get('RANCHI')),
            # 32-34: Online BM
            blinkit_sp if blinkit_sp > 0 else '', jio_sp if jio_sp > 0 else '',
            max_bm if max_bm > 0 else '',
            # 35: gap
            '',
            # 36-37: Changes (blank)
            '', '',
            # 38-50: Offline BM (blank)
            '', '', '', '', '', '', '', '', '', '', '', '', '',
            # 51-54
            sku_margin, sku_markup, guard_price, '',
            # 55-57: Remarks (3 city groups — put main remark in first group)
            rmk, '', '',
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = THIN_BORDER
            if ci <= len(STPLS_COLS):
                name = STPLS_COLS[ci - 1][0]
                if 'MARGIN' in name or 'WEIGHT' in name:
                    if isinstance(val, float) and val != 0:
                        cell.number_format = PCT_FMT
                elif any(k in name for k in ['MRP', 'COST', 'MAP', 'SP', 'SALES', 'QTY', 'BLINKIT', 'JIO', 'MAX']):
                    if isinstance(val, (int, float)) and val != 0:
                        cell.number_format = NUM_FMT

    _autowidth(ws)
    print(f"  ✓ MAIN SHEET: {len(products)} rows, {len(STPLS_COLS)} cols")


# ══════════════════════════════════════════════════════════════════════════════
# STAPLES QUICK DATA — 17 columns (exact match)
# ══════════════════════════════════════════════════════════════════════════════

STPLS_QD_COLS = [
    'key', 'STATE', 'ITEM CODE', 'DISPLAY NAME', 'KVI TAG',
    'MRP', 'LAST INWARD COST', 'CURRENT SP', 'OFF INVOICE ADJUSTED COST',
    'ADR SP', 'KOLKATA SP', 'BILASPUR/KORBA SP', 'JAMSHEDPUR SP',
    'HAZARIBAGH SP', 'RANCHI SP', 'LAST 90 DAY SALE VALUE', 'LAST 90 DAY SALE QTY',
]


def write_staples_quick_data(wb, products, results):
    ws = wb.create_sheet('QUICK DATA')
    for i, h in enumerate(STPLS_QD_COLS, 1):
        ws.cell(row=1, column=i, value=h)
        _hdr(ws, 1, i)
    ws.freeze_panes = 'A2'

    for row, (p, r) in enumerate(zip(products, results), 2):
        city = p.get('city_sps', {})
        vals = [
            p.get('key'), p.get('state'), p.get('item_code'), p.get('display_name'),
            p.get('kvi_tag'), safe_float(p.get('mrp')),
            safe_float(p.get('latest_inward_cost')),
            safe_float(p.get('current_sp')) or '',
            r.get('off_invoice_adjusted_cost', 0),
            r.get('sp', 0),  # ADR SP = cluster-level final SP
            safe_float(city.get('KOLKATA')) or '',
            safe_float(city.get('BILASPUR')) or '',
            safe_float(city.get('JAMSHEDPUR')) or '',
            safe_float(city.get('HAZARIBAGH')) or '',
            safe_float(city.get('RANCHI')) or '',
            safe_float(p.get('sale_90d')), safe_float(p.get('qty_90d')),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = THIN_BORDER
    _autowidth(ws)
    print(f"  ✓ QUICK DATA sheet: 17 cols")


# ══════════════════════════════════════════════════════════════════════════════
# STATE-WISE SHEETS (bonus — not in target, but useful)
# ══════════════════════════════════════════════════════════════════════════════

def _write_state_fmcg(wb, products, results, state):
    ws = wb.create_sheet(f'{state} - {STATE_NAMES.get(state, state)}')
    hdrs = ['ITEM CODE', 'DISPLAY NAME', 'KVI TAG', 'BRAND', 'CATEGORY', 'MRP',
            'INWARD COST', 'MAP', 'OFF INV ADJ COST', 'MARGIN',
            'OFF INV PROMO', 'ON INV PROMO', 'FINAL INVOICE',
            'JIOMART SP', 'BLINKIT SP', 'SP BAU', 'DISCOUNT %', 'RM %', 'RULE', 'REMARK']
    for i, h in enumerate(hdrs, 1):
        ws.cell(row=1, column=i, value=h)
        _hdr(ws, 1, i)
    ws.freeze_panes = 'A2'
    for row, (p, r) in enumerate(zip(products, results), 2):
        vals = [
            p.get('item_code'), p.get('display_name'), p.get('kvi_tag'),
            p.get('brand'), p.get('category') or p.get('product_type'),
            safe_float(p.get('mrp')), safe_float(p.get('latest_inward_cost')),
            safe_float(p.get('map_price')), r.get('off_invoice_adjusted_cost', 0),
            r.get('margin', 0), safe_float(p.get('off_invoice_value')),
            safe_float(p.get('on_invoice_value')), r.get('final_invoice_value', 0),
            safe_float(p.get('jio_sp')), safe_float(p.get('blinkit_sp')),
            r.get('sp', 0), r.get('discount_pct', 0), r.get('retention_margin', 0),
            r.get('rule_applied', ''), '; '.join(r.get('remarks', [])),
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v).border = THIN_BORDER
    _autowidth(ws)
    print(f"  ✓ {state} sheet: {len(products)} rows")


def _write_state_staples(wb, products, results, state):
    ws = wb.create_sheet(f'{state} - {STATE_NAMES.get(state, state)}')
    hdrs = ['ITEM CODE', 'ITEM NAME', 'KVI', 'SKU TYPE', 'BRAND', 'CATEGORY',
            'MRP', 'INWARD COST', 'MAP', 'MAX COST', 'GRN MARGIN',
            'OFF INV PROMO', 'ON INV PROMO', 'BLINKIT SP', 'JIO SP',
            'FINAL SP', 'DISCOUNT %', 'RM %', 'SKU MARGIN',
            'GUARDRAIL LOW', 'GUARDRAIL HIGH', 'RULE', 'REMARK']
    for i, h in enumerate(hdrs, 1):
        ws.cell(row=1, column=i, value=h)
        _hdr(ws, 1, i)
    ws.freeze_panes = 'A2'
    for row, (p, r) in enumerate(zip(products, results), 2):
        mrp = safe_float(p.get('mrp'))
        inward = safe_float(p.get('latest_inward_cost'))
        map_v = safe_float(p.get('map_price'))
        sp = r.get('sp', 0)
        cost = r.get('cost', 0)
        vals = [
            p.get('item_code'), p.get('display_name'), p.get('kvi_tag'),
            p.get('sku_type', 'GKM'), p.get('brand'),
            p.get('category') or p.get('product_type'),
            mrp, inward, map_v, max(inward, map_v),
            (mrp - inward) / mrp if mrp > 0 and inward > 0 else 0,
            safe_float(p.get('off_invoice_value')), safe_float(p.get('on_invoice_value')),
            safe_float(p.get('blinkit_sp')), safe_float(p.get('jio_sp')),
            sp, r.get('discount_pct', 0), r.get('retention_margin', 0),
            (sp - cost) / sp if sp > 0 and cost > 0 else 0,
            safe_float(p.get('guardrail_lower')), safe_float(p.get('guardrail_upper')),
            r.get('rule_applied', ''), '; '.join(r.get('remarks', [])),
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v).border = THIN_BORDER
    _autowidth(ws)
    print(f"  ✓ {state} sheet (Staples): {len(products)} rows")


# ══════════════════════════════════════════════════════════════════════════════
# RULES + SUMMARY + FLAGGED (bonus sheets)
# ══════════════════════════════════════════════════════════════════════════════

def _write_rules(wb):
    ws = wb.create_sheet('Rules')
    rules = [
        ('Rule 1', 'MRP <= 40', 'SP = MRP - promo. No promo → SP = MRP'),
        ('Rule 2', 'NON KVI, MRP > 40', 'A: Promo → SP = MRP - promo. B: No promo → Discount formula'),
        ('Rule 3', 'KVI, MRP > 40', 'A: Promo+BM → MIN(promo,BM) w/ cost floor. B: Promo. C: BM. D: Formula'),
        ('Rule 4', 'SP = MRP', 'MRP>=200 → 1.5% min. 100-200 → 1%. 40-100 → no min'),
        ('Rule 5', 'SP=MRP (competitor at MRP)', 'Discount formula for KVI, MRP>40'),
        ('Rule 6', 'Guardrails', 'No negative disc/RM. MRP>=SP. SP>=Cost'),
        ('Rule 7', 'High Margin', 'MRP>10, margin>54% → 50% disc. Promo>50% → keep promo'),
        ('Rule 8', 'Exclusions (FMCGF)', 'Baby food, BOGO, choc<80 → 0% disc, only promo'),
        ('Rule 9', 'On-Invoice Check', '>80% margin → escalate if RM<6-7%'),
        ('Rule 10', 'RM Calc', 'Overall retention margin'),
        ('Rule 11', 'Dual Invoice', 'Pick greater one'),
    ]
    for i, h in enumerate(['Rule', 'Condition', 'Action'], 1):
        ws.cell(row=1, column=i, value=h)
        _hdr(ws, 1, i)
    for r, (rule, cond, act) in enumerate(rules, 2):
        ws.cell(row=r, column=1, value=rule)
        ws.cell(row=r, column=2, value=cond)
        ws.cell(row=r, column=3, value=act)
    # Discount table
    ws.cell(row=14, column=1, value='DISCOUNT TABLE').font = Font(bold=True)
    for r, (m, d) in enumerate([('<=10%','0%'),('<=15%','2%'),('<=20%','3%'),('<=25%','5%'),
                                 ('<=30%','7%'),('<=40%','10%'),('<=54%','20%'),('>54%','50%')], 15):
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=d)
    _autowidth(ws)


def _write_summary(ws, products, results, cat):
    ws.cell(row=1, column=1, value=f'{cat} Pricing Summary').font = Font(bold=True, size=14)
    total = len(results)
    kvi = sum(1 for p in products if p.get('kvi_tag', '').upper() in ('KVI', 'SUPER KVI'))
    promo = sum(1 for r in results if r.get('final_invoice_value', 0) > 0)
    flagged = sum(1 for r in results if r.get('flags'))
    rule_dist = {}
    for r in results:
        rule_dist[r.get('rule_applied', '?')] = rule_dist.get(r.get('rule_applied', '?'), 0) + 1
    stats = [('Total SKUs', total), ('KVI/Super KVI', kvi), ('Non-KVI', total - kvi),
             ('With Promo', promo), ('Flagged', flagged), ('', ''), ('RULE DISTRIBUTION', '')]
    stats += sorted(rule_dist.items(), key=lambda x: -x[1])
    for r, (k, v) in enumerate(stats, 3):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True) if k else Font()
        ws.cell(row=r, column=2, value=v)
    _autowidth(ws)


def _write_flagged(ws, products, results):
    hdrs = ['KEY', 'ITEM CODE', 'DISPLAY NAME', 'KVI', 'MRP', 'SP', 'FLAG']
    for i, h in enumerate(hdrs, 1):
        ws.cell(row=1, column=i, value=h)
        _hdr(ws, 1, i, FILL_RED)
    row = 2
    for p, r in zip(products, results):
        if r.get('flags'):
            ws.cell(row=row, column=1, value=p.get('key'))
            ws.cell(row=row, column=2, value=p.get('item_code'))
            ws.cell(row=row, column=3, value=p.get('display_name'))
            ws.cell(row=row, column=4, value=p.get('kvi_tag'))
            ws.cell(row=row, column=5, value=p.get('mrp', 0))
            ws.cell(row=row, column=6, value=r.get('sp', 0))
            ws.cell(row=row, column=7, value='; '.join(r['flags']))
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = FLAG_FILL
            row += 1
    _autowidth(ws)
    print(f"  ✓ Flagged_Items: {row - 2} items")


# ══════════════════════════════════════════════════════════════════════════════
# GENERATE FILES
# ══════════════════════════════════════════════════════════════════════════════

def generate_fmcg_file(category, products, results):
    today = datetime.now().strftime('%d_%b_%Y')
    filepath = OUTPUT_DIR / f'{category}_PRICING_{today}.xlsx'
    wb = Workbook()
    write_fmcg_main(wb, products, results)
    for st in ['JH', 'CG', 'WB']:
        sp, sr = _filter_state(products, results, st)
        if sp:
            _write_state_fmcg(wb, sp, sr, st)
    write_fmcg_quick_data(wb, products, results)
    _write_rules(wb)
    ws = wb.create_sheet('Summary')
    _write_summary(ws, products, results, category)
    ws2 = wb.create_sheet('Flagged_Items')
    _write_flagged(ws2, products, results)
    wb.save(filepath)
    print(f"\n  ✅ Saved: {filepath}")
    return filepath


def generate_staples_file(products, results):
    today = datetime.now().strftime('%d_%b_%Y')
    filepath = OUTPUT_DIR / f'STPLS_PRICING_{today}.xlsx'
    wb = Workbook()
    write_staples_main(wb, products, results)
    for st in ['JH', 'CG', 'WB']:
        sp, sr = _filter_state(products, results, st)
        if sp:
            _write_state_staples(wb, sp, sr, st)
    write_staples_quick_data(wb, products, results)
    _write_rules(wb)
    ws = wb.create_sheet('Summary')
    _write_summary(ws, products, results, 'STPLS')
    ws2 = wb.create_sheet('Flagged_Items')
    _write_flagged(ws2, products, results)
    wb.save(filepath)
    print(f"\n  ✅ Saved: {filepath}")
    return filepath


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Generate Apnamart pricing files')
    parser.add_argument('--category', choices=['FMCGF', 'FMCGNF', 'STPLS', 'ALL'],
                        default='ALL', help='Which category to generate')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    create_sample_inputs()

    categories = ['FMCGF', 'FMCGNF', 'STPLS'] if args.category == 'ALL' else [args.category]

    for cat in categories:
        print(f"\n{'='*60}")
        print(f"GENERATING {cat} PRICING FILE")
        print(f"{'='*60}")
        products = load_all_data(master_category=cat)
        if not products:
            print(f"  ⚠ No products for {cat}")
            continue
        print(f"\n  Running engine on {len(products)} products...")
        results = calculate_sp_batch(products)
        ok = sum(1 for r in results if not r.get('error'))
        flagged = sum(1 for r in results if r.get('flags'))
        print(f"  ✓ {ok} OK, {flagged} flagged")
        if args.dry_run:
            continue
        if cat in ('FMCGF', 'FMCGNF'):
            generate_fmcg_file(cat, products, results)
        else:
            generate_staples_file(products, results)

    print(f"\n{'='*60}\nDONE!\n{'='*60}")


if __name__ == '__main__':
    main()
